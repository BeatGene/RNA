import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ============ 配置 ============
CSV_PATH = r"C:\Users\49586\Desktop\Learning\Laboratory\Admis\graduate_first\RNA\Data\pdb_data\all_rna_structures.csv"
XLSX_PATH = r"C:\Users\49586\Desktop\Learning\Laboratory\Admis\graduate_first\RNA\Code\Download_PDB_RAW\rna_experimental_pdb_ids.xlsx"
MISSING_OUTPUT = r"C:\Users\49586\Desktop\Learning\Laboratory\Admis\graduate_first\RNA\Code\Download_PDB_RAW\missing_from_api.csv"
# =============================

# 1. 读取两个文件
df_csv = pd.read_csv(CSV_PATH)
df_xlsx = pd.read_excel(XLSX_PATH, engine="openpyxl")

# 提取 PDB ID 集合
set_csv = set(df_csv["PDB_ID"].dropna().unique())
set_xlsx = set(df_xlsx["PDB_ID"].dropna().unique())

print(f"CSV（原下载记录）中的 PDB ID 数: {len(set_csv)}")
print(f"xlsx（API 查询结果）中的 PDB ID 数: {len(set_xlsx)}")

# 2. 交集和差集
in_both = set_xlsx & set_csv
in_csv_only = set_csv - set_xlsx
in_xlsx_only = set_xlsx - set_csv

print(f"两者共有: {len(in_both)} 个")
print(f"仅在 CSV（原下载）中，API 未返回: {len(in_csv_only)} 个")
print(f"仅在 xlsx（API 新返回）中: {len(in_xlsx_only)} 个")

# 3. 打开 xlsx，写入第二列并标绿
wb = load_workbook(XLSX_PATH)
ws = wb.active

# 写表头
ws.cell(row=1, column=2, value="Previously_Downloaded")

green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

# 从第2行开始遍历（第1行是表头）
for row_idx in range(2, ws.max_row + 1):
    pdb_id = ws.cell(row=row_idx, column=1).value
    if pdb_id and str(pdb_id).strip() in in_both:
        cell = ws.cell(row=row_idx, column=2, value="YES")
        cell.fill = green_fill

wb.save(XLSX_PATH)
print(f"已更新 xlsx 文件: {XLSX_PATH}")

# 4. 输出"CSV有但xlsx没有"的PDB ID
if in_csv_only:
    df_missing = pd.DataFrame({"PDB_ID": sorted(in_csv_only)})
    df_missing.to_csv(MISSING_OUTPUT, index=False, encoding="utf-8-sig")
    print(f"仅在原 CSV 中的 PDB ID 已保存至: {MISSING_OUTPUT}")
else:
    print("所有原下载的 PDB ID 都在 API 查询结果中。")
