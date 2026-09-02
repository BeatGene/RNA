#!/usr/bin/env python3
"""Audit and synchronize experimental pure-RNA PDBx/mmCIF files.

The current target list controls what may be downloaded.  The legacy list is
used only to audit the files migrated from the previous server.  No CIF file is
deleted by this program.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import os
import re
import shutil
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

try:
    import pandas as pd
    import requests
    from gemmi import cif
    from requests.adapters import HTTPAdapter
    from tqdm import tqdm
    from urllib3.util.retry import Retry
except ImportError as exc:  # pragma: no cover - exercised only on bad installs
    missing = getattr(exc, "name", str(exc))
    raise SystemExit(
        f"缺少依赖 {missing!r}。请先运行：\n"
        "python3 -m pip install -r requirements_pdb_pipeline.txt"
    ) from exc


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_TARGET_LIST = (
    SCRIPT_DIR
    / "Second_PDB_ID_xlsx_and_InOut"
    / "experimental_pure_rna_pdb_ids.xlsx"
)
DEFAULT_LEGACY_LIST = (
    SCRIPT_DIR / "Original_count" / "rna_experimental_pdb_ids.xlsx"
)
RCSB_DOWNLOAD_URL = "https://files.rcsb.org/download/{pdb_id}.cif"
RNA_POLYMER_TYPE = "polyribonucleotide"
NULL_CIF_VALUES = {"", ".", "?"}
PDB_ID_PATTERN = re.compile(r"^[A-Z0-9_]+$")

MANIFEST_COLUMNS = [
    "PDB_ID",
    "CURRENT_TARGET",
    "LEGACY_1979",
    "FILE_STATUS",
    "SYNC_STATUS",
    "ELIGIBILITY_STATUS",
    "FILE_PATH",
    "SOURCE_URL",
    "DUPLICATE_FILE_COUNT",
    "FILE_SIZE_BYTES",
    "SHA256",
    "ENTRY_ID_IN_CIF",
    "ENTRY_ID_MATCH",
    "EXPERIMENT_METHOD",
    "RESOLUTION_ANGSTROM",
    "POLYMER_TYPES",
    "PURE_RNA_POLYMERS",
    "RNA_ENTITY_COUNT",
    "RNA_CHAIN_COUNT",
    "ATOM_COUNT",
    "VALIDATION_MESSAGE",
    "DOWNLOAD_ATTEMPTS",
    "CHECKED_AT_UTC",
]

CHAIN_COLUMNS = [
    "PDB_ID",
    "ENTITY_ID",
    "CHAIN_ID",
    "POLYMER_TYPE",
    "SEQUENCE_REPORTED",
    "SEQUENCE_CANONICAL",
    "SEQUENCE_LENGTH",
    "HAS_MODIFIED_RESIDUES",
]


@dataclass
class ValidationResult:
    valid: bool
    message: str
    entry_id: str = ""
    methods: list[str] = field(default_factory=list)
    resolution: str = ""
    polymer_types: list[str] = field(default_factory=list)
    pure_rna: bool = False
    rna_entity_count: int = 0
    rna_chain_count: int = 0
    atom_count: int = 0
    file_size: int = 0
    sha256: str = ""
    chains: list[dict[str, Any]] = field(default_factory=list)


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def normalize_pdb_id(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    pdb_id = str(value).strip().upper()
    if not pdb_id or pdb_id in {"PDB_ID", "NAN", "NONE"}:
        return ""
    if not PDB_ID_PATTERN.fullmatch(pdb_id):
        raise ValueError(f"非法 PDB ID：{value!r}")
    return pdb_id


def load_id_list(path: Path, sheet_name: str | None = None) -> list[str]:
    """Read PDB IDs from the first column of xlsx/csv/tsv/txt."""
    path = path.expanduser().resolve()
    if not path.is_file():
        raise FileNotFoundError(f"ID 清单不存在：{path}")

    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xls"}:
        frame = pd.read_excel(path, sheet_name=sheet_name or 0, dtype=str)
        values = frame.iloc[:, 0].tolist()
    elif suffix == ".csv":
        frame = pd.read_csv(path, dtype=str)
        values = frame.iloc[:, 0].tolist()
    elif suffix in {".tsv", ".tab"}:
        frame = pd.read_csv(path, sep="\t", dtype=str)
        values = frame.iloc[:, 0].tolist()
    else:
        values = [
            line.split(",", maxsplit=1)[0].strip()
            for line in path.read_text(encoding="utf-8-sig").splitlines()
        ]

    ids: list[str] = []
    seen: set[str] = set()
    duplicates: set[str] = set()
    for value in values:
        pdb_id = normalize_pdb_id(value)
        if not pdb_id:
            continue
        if pdb_id in seen:
            duplicates.add(pdb_id)
            continue
        ids.append(pdb_id)
        seen.add(pdb_id)

    if not ids:
        raise ValueError(f"清单没有读到任何 PDB ID：{path}")
    if duplicates:
        sample = ", ".join(sorted(duplicates)[:10])
        print(f"警告：{path.name} 中有重复 ID，已去重：{sample}", file=sys.stderr)
    return ids


def _clean_cif_value(value: Any) -> str:
    if value is None or value is False:
        return ""
    text = cif.as_string(str(value)).strip()
    return "" if text in NULL_CIF_VALUES else text


def _category_column(
    category: dict[str, list[Any]], name: str, row_count: int
) -> list[Any]:
    values = category.get(name, [])
    if not values:
        return [""] * row_count
    if len(values) != row_count:
        raise ValueError(
            f"mmCIF category 列长度不一致：{name}={len(values)}, expected={row_count}"
        )
    return values


def _first_value(block: Any, tags: Iterable[str]) -> str:
    for tag in tags:
        values = block.find_values(tag)
        if values:
            value = _clean_cif_value(values[0])
            if value:
                return value
    return ""


def _normalize_sequence(value: Any) -> str:
    return re.sub(r"\s+", "", _clean_cif_value(value))


def _sequence_length(sequence: str) -> int:
    """Count parenthesized modified monomers as one residue."""
    if not sequence:
        return 0
    tokens = re.findall(r"\([^)]*\)|[A-Za-z?]", sequence)
    return len(tokens)


def sha256_file(path: Path, chunk_size: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while chunk := handle.read(chunk_size):
            digest.update(chunk)
    return digest.hexdigest()


def _validate_numeric_coordinates(block: Any) -> int:
    columns = {
        "id": block.find_values("_atom_site.id"),
        "x": block.find_values("_atom_site.Cartn_x"),
        "y": block.find_values("_atom_site.Cartn_y"),
        "z": block.find_values("_atom_site.Cartn_z"),
    }
    lengths = {name: len(values) for name, values in columns.items()}
    atom_count = lengths["id"]
    if atom_count == 0:
        raise ValueError("缺少 _atom_site 原子坐标")
    if len(set(lengths.values())) != 1:
        raise ValueError(f"_atom_site 必需列长度不一致：{lengths}")

    for axis in ("x", "y", "z"):
        for index, raw in enumerate(columns[axis]):
            try:
                number = float(raw)
            except (TypeError, ValueError) as exc:
                raise ValueError(
                    f"_atom_site.Cartn_{axis} 第 {index + 1} 行不是数字：{raw!r}"
                ) from exc
            if not math.isfinite(number):
                raise ValueError(
                    f"_atom_site.Cartn_{axis} 第 {index + 1} 行不是有限数"
                )
    return atom_count


def _extract_polymer_metadata(
    block: Any, expected_id: str
) -> tuple[list[str], bool, list[dict[str, Any]]]:
    category = block.get_mmcif_category("_entity_poly.")
    entity_ids = category.get("entity_id", [])
    if not entity_ids:
        raise ValueError("缺少 _entity_poly，无法确认 RNA polymer")

    row_count = len(entity_ids)
    types = _category_column(category, "type", row_count)
    strands = _category_column(category, "pdbx_strand_id", row_count)
    reported = _category_column(
        category, "pdbx_seq_one_letter_code", row_count
    )
    canonical = _category_column(
        category, "pdbx_seq_one_letter_code_can", row_count
    )

    polymer_types: list[str] = []
    chains: list[dict[str, Any]] = []
    for entity_id, polymer_type, chain_text, raw_seq, canonical_seq in zip(
        entity_ids, types, strands, reported, canonical
    ):
        clean_type = _clean_cif_value(polymer_type)
        lower_type = clean_type.lower()
        if clean_type and clean_type not in polymer_types:
            polymer_types.append(clean_type)
        if lower_type != RNA_POLYMER_TYPE:
            continue

        reported_seq = _normalize_sequence(raw_seq)
        canonical_seq = _normalize_sequence(canonical_seq)
        if not canonical_seq:
            canonical_seq = reported_seq
        if not canonical_seq:
            raise ValueError(
                f"RNA entity {_clean_cif_value(entity_id)} 缺少 polymer sequence"
            )

        chain_ids = [
            item.strip()
            for item in _clean_cif_value(chain_text).split(",")
            if item.strip()
        ] or [""]
        for chain_id in chain_ids:
            chains.append(
                {
                    "PDB_ID": expected_id,
                    "ENTITY_ID": _clean_cif_value(entity_id),
                    "CHAIN_ID": chain_id,
                    "POLYMER_TYPE": clean_type,
                    "SEQUENCE_REPORTED": reported_seq,
                    "SEQUENCE_CANONICAL": canonical_seq,
                    "SEQUENCE_LENGTH": _sequence_length(canonical_seq),
                    "HAS_MODIFIED_RESIDUES": "(" in reported_seq,
                }
            )

    normalized_types = sorted({item.lower() for item in polymer_types if item})
    pure_rna = bool(normalized_types) and normalized_types == [RNA_POLYMER_TYPE]
    return polymer_types, pure_rna, chains


def validate_cif(path: Path, expected_id: str) -> ValidationResult:
    """Parse a CIF and validate syntax, ID, coordinates and RNA metadata."""
    path = path.expanduser().resolve()
    expected_id = normalize_pdb_id(expected_id)
    result = ValidationResult(valid=False, message="未校验")
    try:
        result.file_size = path.stat().st_size
        if result.file_size == 0:
            raise ValueError("文件大小为 0")
        result.sha256 = sha256_file(path)

        document = cif.read_file(str(path), check_level=2)
        if len(document) != 1:
            raise ValueError(f"预期 1 个 data block，实际 {len(document)} 个")
        block = document[0]

        result.entry_id = _first_value(block, ["_entry.id"]).upper()
        if not result.entry_id:
            raise ValueError("缺少 _entry.id")
        if result.entry_id != expected_id:
            raise ValueError(
                f"_entry.id={result.entry_id} 与预期 {expected_id} 不一致"
            )

        result.methods = sorted(
            {
                _clean_cif_value(value)
                for value in block.find_values("_exptl.method")
                if _clean_cif_value(value)
            }
        )
        if not result.methods:
            raise ValueError("缺少 _exptl.method，无法确认实验方法")

        result.resolution = _first_value(
            block,
            [
                "_refine.ls_d_res_high",
                "_em_3d_reconstruction.resolution",
                "_reflns.d_resolution_high",
            ],
        )
        (
            result.polymer_types,
            result.pure_rna,
            result.chains,
        ) = _extract_polymer_metadata(block, expected_id)
        result.rna_entity_count = len(
            {row["ENTITY_ID"] for row in result.chains}
        )
        result.rna_chain_count = len(result.chains)

        result.atom_count = _validate_numeric_coordinates(block)
        result.valid = True
        result.message = "OK"
    except Exception as exc:
        result.message = f"{type(exc).__name__}: {exc}"
    return result


def _filename_aliases(pdb_id: str) -> set[str]:
    aliases = {pdb_id.upper()}
    if len(pdb_id) == 4:
        aliases.add(f"PDB_0000{pdb_id}".upper())
    return aliases


def index_cif_files(scan_root: Path, pdb_ids: set[str]) -> dict[str, list[Path]]:
    """Index matching .cif files recursively, case-insensitively."""
    aliases = {
        alias: pdb_id
        for pdb_id in pdb_ids
        for alias in _filename_aliases(pdb_id)
    }
    indexed: dict[str, list[Path]] = {pdb_id: [] for pdb_id in pdb_ids}
    for path in scan_root.rglob("*"):
        if not path.is_file() or path.suffix.lower() != ".cif":
            continue
        if "quarantine" in {part.lower() for part in path.parts}:
            continue
        pdb_id = aliases.get(path.stem.upper())
        if pdb_id:
            indexed[pdb_id].append(path.resolve())
    return indexed


def choose_candidate(paths: list[Path], preferred: Path) -> Path | None:
    if not paths:
        return None
    preferred_resolved = preferred.resolve()
    for path in paths:
        if path.resolve() == preferred_resolved:
            return path
    return sorted(paths, key=lambda item: (len(item.parts), str(item).lower()))[0]


def build_http_session(retries: int) -> requests.Session:
    retry = Retry(
        total=retries,
        connect=retries,
        read=retries,
        status=retries,
        backoff_factor=1.0,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=frozenset({"GET"}),
        respect_retry_after_header=True,
    )
    session = requests.Session()
    session.headers.update(
        {"User-Agent": "RNA-PDB-CIF-Pipeline/1.0 (research data curation)"}
    )
    session.mount("https://", HTTPAdapter(max_retries=retry))
    return session


def _quarantine_copy(path: Path, quarantine_dir: Path, pdb_id: str) -> Path:
    quarantine_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%dT%H%M%S")
    target = quarantine_dir / f"{pdb_id.lower()}.{timestamp}.invalid.cif"
    counter = 1
    while target.exists():
        target = quarantine_dir / (
            f"{pdb_id.lower()}.{timestamp}.{counter}.invalid.cif"
        )
        counter += 1
    shutil.copy2(path, target)
    return target


def download_and_validate(
    pdb_id: str,
    destination: Path,
    quarantine_dir: Path,
    session: requests.Session,
    timeout: float,
) -> tuple[ValidationResult | None, int, str]:
    """Stream to a temporary file, validate, then atomically install it."""
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary = destination.with_name(
        f".{destination.name}.{os.getpid()}.part"
    )
    url = RCSB_DOWNLOAD_URL.format(pdb_id=pdb_id)
    attempts = 1
    try:
        with session.get(url, stream=True, timeout=(20, timeout)) as response:
            response.raise_for_status()
            with temporary.open("wb") as handle:
                for chunk in response.iter_content(chunk_size=1024 * 1024):
                    if chunk:
                        handle.write(chunk)
                handle.flush()
                os.fsync(handle.fileno())

        validation = validate_cif(temporary, pdb_id)
        if not validation.valid:
            return (
                validation,
                attempts,
                f"下载内容未通过校验：{validation.message}",
            )

        if destination.exists():
            old_validation = validate_cif(destination, pdb_id)
            if not old_validation.valid:
                saved = _quarantine_copy(
                    destination, quarantine_dir, pdb_id
                )
                print(f"\n已备份损坏文件：{saved}")
        os.replace(temporary, destination)
        final_validation = validate_cif(destination, pdb_id)
        if not final_validation.valid:
            return (
                final_validation,
                attempts,
                f"原子落盘后校验失败：{final_validation.message}",
            )
        return final_validation, attempts, "OK"
    except requests.RequestException as exc:
        return None, attempts, f"{type(exc).__name__}: {exc}"
    except OSError as exc:
        return None, attempts, f"{type(exc).__name__}: {exc}"
    finally:
        if temporary.exists():
            temporary.unlink()


def manifest_row(
    pdb_id: str,
    current: bool,
    legacy: bool,
    file_status: str,
    sync_status: str,
    path: Path | None,
    duplicate_count: int,
    validation: ValidationResult | None,
    message: str,
    attempts: int,
    checked_at: str,
) -> dict[str, Any]:
    validation = validation or ValidationResult(valid=False, message=message)
    return {
        "PDB_ID": pdb_id,
        "CURRENT_TARGET": current,
        "LEGACY_1979": legacy,
        "FILE_STATUS": file_status,
        "SYNC_STATUS": sync_status,
        "ELIGIBILITY_STATUS": (
            "ELIGIBLE_CURRENT"
            if current and validation.valid and validation.pure_rna
            else (
                "CURRENT_LIST_POLYMER_MISMATCH"
                if current and validation.valid
                else ("LEGACY_ONLY" if legacy and not current else "NOT_ASSESSED")
            )
        ),
        "FILE_PATH": str(path.resolve()) if path else "",
        "SOURCE_URL": (
            RCSB_DOWNLOAD_URL.format(pdb_id=pdb_id) if current else ""
        ),
        "DUPLICATE_FILE_COUNT": duplicate_count,
        "FILE_SIZE_BYTES": validation.file_size or "",
        "SHA256": validation.sha256,
        "ENTRY_ID_IN_CIF": validation.entry_id,
        "ENTRY_ID_MATCH": bool(
            validation.entry_id and validation.entry_id == pdb_id
        ),
        "EXPERIMENT_METHOD": " | ".join(validation.methods),
        "RESOLUTION_ANGSTROM": validation.resolution,
        "POLYMER_TYPES": " | ".join(validation.polymer_types),
        "PURE_RNA_POLYMERS": validation.pure_rna,
        "RNA_ENTITY_COUNT": validation.rna_entity_count or "",
        "RNA_CHAIN_COUNT": validation.rna_chain_count or "",
        "ATOM_COUNT": validation.atom_count or "",
        "VALIDATION_MESSAGE": message,
        "DOWNLOAD_ATTEMPTS": attempts,
        "CHECKED_AT_UTC": checked_at,
    }


def _atomic_to_csv(frame: pd.DataFrame, path: Path) -> None:
    temporary = path.with_suffix(path.suffix + ".tmp")
    frame.to_csv(temporary, index=False, encoding="utf-8-sig")
    os.replace(temporary, path)


def _excel_safe_sequences(frame: pd.DataFrame) -> pd.DataFrame:
    safe = frame.copy()
    excel_limit = 32_000
    for column in ("SEQUENCE_REPORTED", "SEQUENCE_CANONICAL"):
        if column in safe:
            safe[column] = safe[column].map(
                lambda value: (
                    value
                    if not isinstance(value, str) or len(value) <= excel_limit
                    else value[:excel_limit] + "...[完整序列见 CSV]"
                )
            )
    return safe


def _style_workbook(path: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    good_fill = PatternFill("solid", fgColor="C6EFCE")
    bad_fill = PatternFill("solid", fgColor="FFC7CE")
    warn_fill = PatternFill("solid", fgColor="FFEB9C")
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
        for column_cells in sheet.columns:
            values = [str(cell.value or "") for cell in list(column_cells)[:200]]
            width = min(max(max(map(len, values), default=8) + 2, 10), 45)
            sheet.column_dimensions[
                get_column_letter(column_cells[0].column)
            ].width = width

    manifest = workbook["结构清单"]
    status_column = next(
        cell.column
        for cell in manifest[1]
        if cell.value == "FILE_STATUS"
    )
    letter = get_column_letter(status_column)
    last_row = max(manifest.max_row, 2)
    manifest.conditional_formatting.add(
        f"A2:{get_column_letter(manifest.max_column)}{last_row}",
        FormulaRule(formula=[f'${letter}2="VALID"'], fill=good_fill),
    )
    manifest.conditional_formatting.add(
        f"A2:{get_column_letter(manifest.max_column)}{last_row}",
        FormulaRule(formula=[f'${letter}2="INVALID"'], fill=bad_fill),
    )
    manifest.conditional_formatting.add(
        f"A2:{get_column_letter(manifest.max_column)}{last_row}",
        FormulaRule(formula=[f'${letter}2="MISSING"'], fill=warn_fill),
    )
    workbook.save(path)


def write_reports(
    report_dir: Path,
    manifest_rows: list[dict[str, Any]],
    chain_rows: list[dict[str, Any]],
    summary: dict[str, Any],
) -> None:
    report_dir.mkdir(parents=True, exist_ok=True)
    manifest = pd.DataFrame(manifest_rows, columns=MANIFEST_COLUMNS)
    chains = pd.DataFrame(chain_rows, columns=CHAIN_COLUMNS)
    summary_frame = pd.DataFrame(
        [{"METRIC": key, "VALUE": value} for key, value in summary.items()]
    )

    _atomic_to_csv(manifest, report_dir / "pdb_cif_manifest.csv")
    _atomic_to_csv(chains, report_dir / "rna_chain_sequences.csv")

    summary_path = report_dir / "summary.json"
    temporary_json = summary_path.with_suffix(".json.tmp")
    temporary_json.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    os.replace(temporary_json, summary_path)

    workbook_path = report_dir / "pdb_cif_report.xlsx"
    temporary_xlsx = workbook_path.with_name(
        f".{workbook_path.stem}.tmp.xlsx"
    )
    with pd.ExcelWriter(temporary_xlsx, engine="openpyxl") as writer:
        manifest.to_excel(writer, sheet_name="结构清单", index=False)
        _excel_safe_sequences(chains).to_excel(
            writer, sheet_name="RNA链序列", index=False
        )
        summary_frame.to_excel(writer, sheet_name="汇总", index=False)
    _style_workbook(temporary_xlsx)
    os.replace(temporary_xlsx, workbook_path)


def calculate_summary(
    rows: list[dict[str, Any]],
    target_count: int,
    legacy_count: int,
    started_at: str,
    finished_at: str,
) -> dict[str, Any]:
    current = [row for row in rows if row["CURRENT_TARGET"]]
    legacy = [row for row in rows if row["LEGACY_1979"]]
    return {
        "started_at_utc": started_at,
        "finished_at_utc": finished_at,
        "current_target_expected": target_count,
        "current_target_valid": sum(
            row["FILE_STATUS"] == "VALID" for row in current
        ),
        "current_target_missing": sum(
            row["FILE_STATUS"] == "MISSING" for row in current
        ),
        "current_target_invalid": sum(
            row["FILE_STATUS"] == "INVALID" for row in current
        ),
        "current_target_polymer_mismatch": sum(
            row["ELIGIBILITY_STATUS"] == "CURRENT_LIST_POLYMER_MISMATCH"
            for row in current
        ),
        "downloaded_valid_this_run": sum(
            row["SYNC_STATUS"] == "DOWNLOADED_VALID" for row in current
        ),
        "download_failed_this_run": sum(
            row["SYNC_STATUS"] == "DOWNLOAD_FAILED" for row in current
        ),
        "legacy_expected": legacy_count,
        "legacy_valid": sum(
            row["FILE_STATUS"] == "VALID" for row in legacy
        ),
        "legacy_missing": sum(
            row["FILE_STATUS"] == "MISSING" for row in legacy
        ),
        "legacy_invalid": sum(
            row["FILE_STATUS"] == "INVALID" for row in legacy
        ),
        "legacy_only_not_current": sum(
            row["LEGACY_1979"] and not row["CURRENT_TARGET"] for row in rows
        ),
        "all_current_targets_complete": bool(current)
        and all(
            row["FILE_STATUS"] == "VALID"
            and row["ELIGIBILITY_STATUS"] == "ELIGIBLE_CURRENT"
            for row in current
        ),
        "all_legacy_files_intact": bool(legacy)
        and all(row["FILE_STATUS"] == "VALID" for row in legacy),
    }


def append_event(path: Path, row: dict[str, Any]) -> None:
    with path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(row, ensure_ascii=False, default=str) + "\n")
        handle.flush()


def run_pipeline(args: argparse.Namespace) -> int:
    started_at = utc_now()
    scan_root = args.scan_root.expanduser().resolve()
    if not scan_root.is_dir():
        raise FileNotFoundError(f"扫描目录不存在：{scan_root}")

    download_dir = (
        args.download_dir.expanduser().resolve()
        if args.download_dir
        else (
            scan_root / "01_Pure_RNA"
            if (scan_root / "01_Pure_RNA").is_dir()
            else scan_root
        )
    )
    report_dir = (
        args.report_dir.expanduser().resolve()
        if args.report_dir
        else scan_root / "pipeline_reports"
    )
    report_dir.mkdir(parents=True, exist_ok=True)
    events_path = report_dir / "run_events.jsonl"
    quarantine_dir = report_dir / "quarantine"

    target_ids = load_id_list(args.target_list, args.target_sheet)
    legacy_ids = load_id_list(args.legacy_list, args.legacy_sheet)
    target_set = set(target_ids)
    legacy_set = set(legacy_ids)
    all_ids = target_set | legacy_set

    print(
        f"当前目标 {len(target_set)} 个；旧集合 {len(legacy_set)} 个；"
        f"交集 {len(target_set & legacy_set)} 个；"
        f"新增 {len(target_set - legacy_set)} 个；"
        f"旧集合独有 {len(legacy_set - target_set)} 个。"
    )
    print(f"扫描目录：{scan_root}")
    print(f"新文件目录：{download_dir}")
    print(f"报告目录：{report_dir}")
    print(f"模式：{'校验并下载' if args.mode == 'sync' else '只校验'}")

    indexed = index_cif_files(scan_root, all_ids)
    session = build_http_session(args.retries)
    checked_at = utc_now()
    rows: list[dict[str, Any]] = []
    chain_rows: list[dict[str, Any]] = []

    for pdb_id in tqdm(sorted(all_ids), desc="校验/同步 CIF", unit="个"):
        current = pdb_id in target_set
        legacy = pdb_id in legacy_set
        destination = download_dir / f"{pdb_id.lower()}.cif"
        paths = indexed.get(pdb_id, [])
        selected = choose_candidate(paths, destination)
        duplicate_count = max(0, len(paths) - 1)
        validation: ValidationResult | None = None
        attempts = 0
        message = ""

        if selected:
            validation = validate_cif(selected, pdb_id)
            if validation.valid:
                file_status = "VALID"
                sync_status = "EXISTING_VALID"
                message = "OK"
            else:
                file_status = "INVALID"
                sync_status = "EXISTING_INVALID"
                message = validation.message
        else:
            file_status = "MISSING"
            sync_status = "NOT_DOWNLOADED"
            message = "没有找到 CIF 文件"

        needs_download = (
            args.mode == "sync"
            and current
            and file_status != "VALID"
        )
        if needs_download:
            downloaded, attempts, download_message = download_and_validate(
                pdb_id=pdb_id,
                destination=destination,
                quarantine_dir=quarantine_dir,
                session=session,
                timeout=args.timeout,
            )
            if downloaded and downloaded.valid:
                validation = downloaded
                selected = destination
                file_status = "VALID"
                sync_status = "DOWNLOADED_VALID"
                message = "OK"
            else:
                if downloaded:
                    validation = downloaded
                    file_status = "INVALID"
                sync_status = "DOWNLOAD_FAILED"
                message = download_message
            if args.delay > 0:
                time.sleep(args.delay)

        row = manifest_row(
            pdb_id=pdb_id,
            current=current,
            legacy=legacy,
            file_status=file_status,
            sync_status=sync_status,
            path=selected,
            duplicate_count=duplicate_count,
            validation=validation,
            message=message,
            attempts=attempts,
            checked_at=checked_at,
        )
        rows.append(row)
        if validation and validation.valid:
            chain_rows.extend(validation.chains)
        append_event(events_path, row)

    finished_at = utc_now()
    summary = calculate_summary(
        rows, len(target_set), len(legacy_set), started_at, finished_at
    )
    write_reports(report_dir, rows, chain_rows, summary)

    print("\n运行完成")
    for key, value in summary.items():
        print(f"  {key}: {value}")
    print(f"\n易读报告：{report_dir / 'pdb_cif_report.xlsx'}")
    print(f"结构清单：{report_dir / 'pdb_cif_manifest.csv'}")
    print(f"RNA 链序列：{report_dir / 'rna_chain_sequences.csv'}")

    complete = (
        summary["all_current_targets_complete"]
        and summary["all_legacy_files_intact"]
    )
    return 0 if complete else 2


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "校验旧服务器迁移的 1979 个 CIF，并同步当前实验纯 RNA 清单。"
        )
    )
    parser.add_argument(
        "mode",
        choices=("audit", "sync"),
        help="audit 只校验；sync 校验并下载缺失/损坏的当前目标",
    )
    parser.add_argument(
        "--scan-root",
        type=Path,
        required=True,
        help="递归查找已有 .cif 的根目录，例如 ~/pdb_data",
    )
    parser.add_argument(
        "--download-dir",
        type=Path,
        help=(
            "新 CIF 的落盘目录；省略时优先使用"
            " <scan-root>/01_Pure_RNA，否则使用 scan-root"
        ),
    )
    parser.add_argument(
        "--report-dir",
        type=Path,
        help="报告目录；默认 <scan-root>/pipeline_reports",
    )
    parser.add_argument(
        "--target-list",
        type=Path,
        default=DEFAULT_TARGET_LIST,
        help="当前实验纯 RNA ID 清单",
    )
    parser.add_argument(
        "--target-sheet",
        default="PDB_IDs",
        help="当前清单的 Excel sheet 名",
    )
    parser.add_argument(
        "--legacy-list",
        type=Path,
        default=DEFAULT_LEGACY_LIST,
        help="旧 1979 个纯 RNA ID 清单",
    )
    parser.add_argument(
        "--legacy-sheet",
        default="纯RNA",
        help="旧清单的 Excel sheet 名",
    )
    parser.add_argument(
        "--timeout",
        type=float,
        default=180.0,
        help="单次下载读取超时秒数（默认 180）",
    )
    parser.add_argument(
        "--retries",
        type=int,
        default=4,
        help="HTTP 自动重试次数（默认 4）",
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=0.15,
        help="下载之间的礼貌等待秒数（默认 0.15）",
    )
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    try:
        return run_pipeline(args)
    except (FileNotFoundError, ValueError, OSError) as exc:
        parser.error(str(exc))
    return 2


if __name__ == "__main__":
    raise SystemExit(main())
