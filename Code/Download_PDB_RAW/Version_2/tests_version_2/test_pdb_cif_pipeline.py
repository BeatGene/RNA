from __future__ import annotations

import importlib.util
import sys
import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


MODULE_PATH = Path(__file__).resolve().parents[1] / "pdb_cif_pipeline.py"
SPEC = importlib.util.spec_from_file_location("pdb_cif_pipeline", MODULE_PATH)
pipeline = importlib.util.module_from_spec(SPEC)
assert SPEC.loader is not None
sys.modules[SPEC.name] = pipeline
SPEC.loader.exec_module(pipeline)


VALID_CIF = """\
data_1ABC
_entry.id 1ABC
_struct.title 'Synthetic RNA'
loop_
_exptl.entry_id
_exptl.method
1ABC 'X-RAY DIFFRACTION'
loop_
_entity_poly.entity_id
_entity_poly.type
_entity_poly.pdbx_seq_one_letter_code
_entity_poly.pdbx_seq_one_letter_code_can
_entity_poly.pdbx_strand_id
1 polyribonucleotide 'AC(PSU)G' 'ACUG' 'A,B'
loop_
_atom_site.id
_atom_site.Cartn_x
_atom_site.Cartn_y
_atom_site.Cartn_z
1 1.0 2.0 3.0
2 4.0 5.0 6.0
#
"""


class PdbCifPipelineTests(unittest.TestCase):
    def test_load_excel_ids_deduplicates_and_normalizes(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "ids.xlsx"
            pd.DataFrame({"PDB_ID": ["1abc", "2XYZ", "1ABC", None]}).to_excel(
                path, sheet_name="PDB_IDs", index=False
            )
            self.assertEqual(
                pipeline.load_id_list(path, "PDB_IDs"), ["1ABC", "2XYZ"]
            )

    def test_validate_cif_extracts_method_and_chain_sequences(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "1abc.cif"
            path.write_text(VALID_CIF, encoding="utf-8")
            result = pipeline.validate_cif(path, "1ABC")

            self.assertTrue(result.valid, result.message)
            self.assertEqual(result.methods, ["X-RAY DIFFRACTION"])
            self.assertTrue(result.pure_rna)
            self.assertEqual(result.atom_count, 2)
            self.assertEqual(result.rna_entity_count, 1)
            self.assertEqual(result.rna_chain_count, 2)
            self.assertEqual(
                [row["CHAIN_ID"] for row in result.chains], ["A", "B"]
            )
            self.assertEqual(result.chains[0]["SEQUENCE_CANONICAL"], "ACUG")
            self.assertEqual(result.chains[0]["SEQUENCE_LENGTH"], 4)
            self.assertTrue(result.chains[0]["HAS_MODIFIED_RESIDUES"])

    def test_validate_cif_rejects_wrong_entry_id(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "9xyz.cif"
            path.write_text(VALID_CIF, encoding="utf-8")
            result = pipeline.validate_cif(path, "9XYZ")
            self.assertFalse(result.valid)
            self.assertIn("与预期 9XYZ 不一致", result.message)

    def test_intact_non_rna_legacy_file_is_not_called_corrupt(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "1abc.cif"
            path.write_text(
                VALID_CIF.replace(
                    "polyribonucleotide", "polydeoxyribonucleotide"
                ),
                encoding="utf-8",
            )
            result = pipeline.validate_cif(path, "1ABC")
            self.assertTrue(result.valid, result.message)
            self.assertFalse(result.pure_rna)
            self.assertEqual(result.rna_chain_count, 0)

    def test_index_is_case_insensitive_and_recursive(self):
        with tempfile.TemporaryDirectory() as directory:
            nested = Path(directory) / "01_Pure_RNA"
            nested.mkdir()
            path = nested / "1AbC.CIF"
            path.write_text(VALID_CIF, encoding="utf-8")
            indexed = pipeline.index_cif_files(
                Path(directory), {"1ABC", "2XYZ"}
            )
            self.assertEqual(indexed["1ABC"], [path.resolve()])
            self.assertEqual(indexed["2XYZ"], [])

    def test_reports_include_three_readable_sheets(self):
        with tempfile.TemporaryDirectory() as directory:
            cif_path = Path(directory) / "1abc.cif"
            cif_path.write_text(VALID_CIF, encoding="utf-8")
            validation = pipeline.validate_cif(cif_path, "1ABC")
            row = pipeline.manifest_row(
                pdb_id="1ABC",
                current=True,
                legacy=True,
                file_status="VALID",
                sync_status="EXISTING_VALID",
                path=cif_path,
                duplicate_count=0,
                validation=validation,
                message="OK",
                attempts=0,
                checked_at="2026-07-27T00:00:00+00:00",
            )
            report_dir = Path(directory) / "reports"
            pipeline.write_reports(
                report_dir,
                [row],
                validation.chains,
                {"all_current_targets_complete": True},
            )

            workbook = load_workbook(
                report_dir / "pdb_cif_report.xlsx", read_only=True
            )
            self.assertEqual(
                workbook.sheetnames, ["结构清单", "RNA链序列", "汇总"]
            )
            workbook.close()
            self.assertTrue((report_dir / "pdb_cif_manifest.csv").is_file())
            self.assertTrue((report_dir / "rna_chain_sequences.csv").is_file())
            self.assertTrue((report_dir / "summary.json").is_file())


if __name__ == "__main__":
    unittest.main()
