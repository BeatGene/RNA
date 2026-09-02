#!/usr/bin/env python3
"""Download OUT-labelled PDB entries from the second worksheet as mmCIF files."""

from __future__ import annotations

import argparse
import os
import re
import shutil
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from urllib.request import Request, urlopen

from openpyxl import load_workbook


DEFAULT_XLSX = Path(
    "~/Code/Download_PDB_RAW/Second_PDB_ID_xlsx_and_InOut/"
    "experimental_pure_rna_pdb_ids.xlsx"
).expanduser()
DEFAULT_OUTPUT_DIR = Path("~/pdb_data_mixed").expanduser()
PDB_ID_PATTERN = re.compile(r"[0-9][A-Z0-9]{3}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Read the second worksheet of an XLSX file and download entries whose "
            "fourth column is OUT as PDBx/mmCIF files."
        )
    )
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=DEFAULT_XLSX,
        help=f"input XLSX file (default: {DEFAULT_XLSX})",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help=f"existing output directory (default: {DEFAULT_OUTPUT_DIR})",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=4,
        help="number of parallel downloads (default: 4)",
    )
    parser.add_argument(
        "--retries",
        type=int,
        default=3,
        help="maximum attempts for each download (default: 3)",
    )
    return parser.parse_args()


def read_out_pdb_ids(xlsx_path: Path) -> list[str]:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if len(workbook.worksheets) < 2:
            raise ValueError("Excel 文件少于两张工作表，无法读取第二张表")

        worksheet = workbook.worksheets[1]
        pdb_ids: set[str] = set()

        for row_number, row in enumerate(
            worksheet.iter_rows(
                min_row=2,
                min_col=1,
                max_col=4,
                values_only=True,
            ),
            start=2,
        ):
            pdb_id, in_or_out = row[0], row[3]
            if pdb_id is None or in_or_out is None:
                continue

            pdb_id = str(pdb_id).strip().upper()
            in_or_out = str(in_or_out).strip().upper()

            if in_or_out != "OUT":
                continue

            if PDB_ID_PATTERN.fullmatch(pdb_id):
                pdb_ids.add(pdb_id)
            else:
                print(
                    f"[警告] 第 {row_number} 行的 PDB ID 格式异常，已忽略：{pdb_id}",
                    flush=True,
                )

        return sorted(pdb_ids)
    finally:
        workbook.close()


def download_cif(
    pdb_id: str,
    output_dir: Path,
    retries: int,
) -> tuple[str, int, str | None]:
    url = f"https://files.rcsb.org/download/{pdb_id.lower()}.cif"
    target = output_dir / f"{pdb_id}.cif"
    temporary = output_dir / f".{pdb_id}.cif.part"
    last_error: Exception | None = None

    for attempt in range(1, retries + 1):
        try:
            request = Request(
                url,
                headers={"User-Agent": "RNA-CIF-downloader/1.0"},
            )
            with urlopen(request, timeout=120) as response:
                with temporary.open("wb") as output_file:
                    shutil.copyfileobj(response, output_file)

            if temporary.stat().st_size == 0:
                raise RuntimeError("下载结果为空文件")

            os.replace(temporary, target)
            return pdb_id, target.stat().st_size, None
        except Exception as error:  # report network and filesystem errors together
            last_error = error
            try:
                temporary.unlink()
            except FileNotFoundError:
                pass

            if attempt < retries:
                time.sleep(attempt * 2)

    return pdb_id, 0, str(last_error)


def main() -> int:
    args = parse_args()
    xlsx_path = args.xlsx.expanduser().resolve()
    output_dir = args.output_dir.expanduser().resolve()

    if not xlsx_path.is_file():
        print(f"[错误] Excel 文件不存在：{xlsx_path}", file=sys.stderr, flush=True)
        return 2
    if not output_dir.is_dir():
        print(f"[错误] 目标目录不存在：{output_dir}", file=sys.stderr, flush=True)
        return 2
    if args.workers < 1:
        print("[错误] --workers 必须大于或等于 1", file=sys.stderr, flush=True)
        return 2
    if args.retries < 1:
        print("[错误] --retries 必须大于或等于 1", file=sys.stderr, flush=True)
        return 2

    try:
        pdb_ids = read_out_pdb_ids(xlsx_path)
    except Exception as error:
        print(f"[错误] 读取 Excel 失败：{error}", file=sys.stderr, flush=True)
        return 2

    print(f"Excel 文件：{xlsx_path}", flush=True)
    print("工作表：第二张工作表", flush=True)
    print(f"共找到 {len(pdb_ids)} 个不重复的 OUT PDB ID", flush=True)
    print(f"下载目录：{output_dir}", flush=True)
    print("下载格式：PDBx/mmCIF (.cif)", flush=True)

    id_list = output_dir / "out_pdb_ids.txt"
    id_list.write_text(
        "".join(f"{pdb_id}\n" for pdb_id in pdb_ids),
        encoding="utf-8",
    )

    success: list[str] = []
    failed: list[tuple[str, str]] = []

    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        futures = {
            executor.submit(download_cif, pdb_id, output_dir, args.retries): pdb_id
            for pdb_id in pdb_ids
        }

        for future in as_completed(futures):
            pdb_id, size, error = future.result()
            if error is None:
                success.append(pdb_id)
                print(f"[成功] {pdb_id}.cif ({size:,} bytes)", flush=True)
            else:
                failed.append((pdb_id, error))
                print(f"[失败] {pdb_id}: {error}", flush=True)

    print("\n========== 下载结果 ==========", flush=True)
    print(f"预期：{len(pdb_ids)}", flush=True)
    print(f"成功：{len(success)}", flush=True)
    print(f"失败：{len(failed)}", flush=True)
    print(f"ID 清单：{id_list}", flush=True)

    if failed:
        print("\n下载失败的 ID：", flush=True)
        for pdb_id, error in failed:
            print(f"{pdb_id}\t{error}", flush=True)
        return 1

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
