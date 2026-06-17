import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ============ 配置 ============
PDB_DATA_DIR = "/remote-home/jinxianwang/tinghaoxia/RNA/Data/pdb_data"
XLSX_PATH = "/remote-home/jinxianwang/tinghaoxia/RNA/Code/Download_PDB_RAW/rna_experimental_pdb_ids.xlsx"
CIF_ONLY_OUTPUT = "/remote-home/jinxianwang/tinghaoxia/RNA/Code/Download_PDB_RAW/cif_not_in_api.csv"
# =============================

# 要扫描的子文件夹（以及根目录直接存放的 .cif）
SUBDIRS = [
    "01_Pure_RNA",
    "02_RNA_Protein_Complex",
    "03_Ribosome_Apo",
    "04_Ribosome_Bound_RNA",
    "05_Others_or_Failed",
]

# 1. 扫描所有 .cif 文件，提取 PDB ID
set_cif = set()

# 根目录下的 .cif
for f in os.listdir(PDB_DATA_DIR):
    if f.endswith(".cif"):
        pdb_id = f.replace(".cif", "").upper()
        set_cif.add(pdb_id)

# 子文件夹下的 .cif
for subdir in SUBDIRS:
    subdir_path = os.path.join(PDB_DATA_DIR, subdir)
    if not os.path.exists(subdir_path):
        print(f"警告: 目录不存在 {subdir_path}，跳过。")
        continue
    for f in os.listdir(subdir_path):
        if f.endswith(".cif"):
            pdb_id = f.replace(".cif", "").upper()
            set_cif.add(pdb_id)

print(f"扫描到 .cif 文件（去重后 PDB ID）: {len(set_cif)} 个")

# 2. 读取 xlsx
df_xlsx = pd.read_excel(XLSX_PATH, engine="openpyxl")
set_xlsx = set(df_xlsx["PDB_ID"].dropna().astype(str).str.strip().str.upper())

print(f"xlsx 中的 PDB ID 数: {len(set_xlsx)}")

# 3. 交集与差集
in_both = set_xlsx & set_cif
in_xlsx_only = set_xlsx - set_cif  # API 有但没下载的
in_cif_only = set_cif - set_xlsx   # 本地有但 API 没返回的

print(f"xlsx 中有 cif 文件的: {len(in_both)} 个")
print(f"xlsx 中没有 cif 文件的: {len(in_xlsx_only)} 个")
print(f"本地 cif 文件但不在 xlsx 中: {len(in_cif_only)} 个")

# 4. 打开 xlsx，写入第三列并上色
wb = load_workbook(XLSX_PATH)
ws = wb.active

green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

# 写表头
ws.cell(row=1, column=3, value="Has_CIF_File")

num_rows = len(df_xlsx)
for row_idx in range(2, num_rows + 2):
    cell_a = ws.cell(row=row_idx, column=1)
    if cell_a.value is None:
        continue
    pdb_id = str(cell_a.value).strip().upper()
    cell_c = ws.cell(row=row_idx, column=3)

    if pdb_id in in_both:
        cell_c.value = "YES"
        cell_c.fill = green_fill

wb.save(XLSX_PATH)
print(f"已更新 xlsx 第三列: {XLSX_PATH}")

# 5. 输出本地 cif 有但 xlsx 没有的 PDB ID
if in_cif_only:
    # 尝试获取这些 ID 所在的文件夹
    records = []
    for pdb_id in sorted(in_cif_only):
        location = "unknown"
        # 先在根目录找
        cif_name = f"{pdb_id.lower()}.cif"
        if os.path.exists(os.path.join(PDB_DATA_DIR, cif_name)):
            location = PDB_DATA_DIR
        else:
            for subdir in SUBDIRS:
                subdir_path = os.path.join(PDB_DATA_DIR, subdir)
                if os.path.exists(os.path.join(subdir_path, cif_name)):
                    location = subdir_path
                    break
        records.append({"PDB_ID": pdb_id, "Location": location})

    df_cif_only = pd.DataFrame(records)
    df_cif_only.to_csv(CIF_ONLY_OUTPUT, index=False, encoding="utf-8-sig")
    print(f"本地 cif 但不在 API 中的 PDB ID 已保存至: {CIF_ONLY_OUTPUT}")
    print(f"  （共 {len(records)} 个，含所在文件夹路径）")
else:
    print("所有本地 cif 文件的 PDB ID 都在 xlsx 中。")

# 6. 汇总
print("\n===== 汇总 =====")
print(f"xlsx 总行数: {len(set_xlsx)}")
print(f"  绿色(YES): {len(in_both)} 个 — 本地已有 cif 文件")
print(f"  未标记   : {len(in_xlsx_only)} 个 — 尚未下载")
print(f"本地 cif 但不在 xlsx: {len(in_cif_only)} 个 → {CIF_ONLY_OUTPUT}")
