#!/usr/bin/env python3
"""Build an XLSX metadata table with one row per RNA chain from local mmCIF files.

The two input collections are expected to be:

* ~/pdb_data: 2246 pure-RNA mmCIF files, of which five are excluded by an XLSX list.
* ~/pdb_data_mixed: 77 useful mixed-composition mmCIF files.

For an RNA entity instantiated as multiple chains, one output row is emitted for
each chain.  All such rows share the entity sequence and sequence length.
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
from collections import Counter
from pathlib import Path
from typing import Any, Iterable

try:
    from gemmi import cif
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    missing = getattr(exc, "name", str(exc))
    raise SystemExit(
        f"缺少依赖 {missing!r}。请先安装：\n"
        "python3 -m pip install --user gemmi openpyxl"
    ) from exc


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_PURE_DIR = Path("~/pdb_data").expanduser()
DEFAULT_MIXED_DIR = Path("~/pdb_data_mixed").expanduser()
DEFAULT_EXCLUSION_XLSX = (
    SCRIPT_DIR / "alignment_pdb_ids_not_in_experimental_pure_rna.xlsx"
)
DEFAULT_RECORD_XLSX = SCRIPT_DIR / "experimental_pure_rna_pdb_ids.xlsx"
DEFAULT_OUTPUT_XLSX = SCRIPT_DIR / "rna_entity_metadata.xlsx"

EXPECTED_PURE_TOTAL = 2246
EXPECTED_EXCLUDED = 5
EXPECTED_PURE_USED = 2241
EXPECTED_MIXED = 77

RNA_POLYMER_TYPE = "polyribonucleotide"
NULL_CIF_VALUES = {"", ".", "?"}
PDB_ID_PATTERN = re.compile(r"^[A-Z0-9]{4}$")

OUTPUT_COLUMNS = [
    "TARGET_ID",
    "DATASET",
    "PDB_ID",
    "CHAIN_IDS",
    "SEQUENCE",
    "LENGTH",
    "RELEASE_DATE",
    "EXPERIMENT_METHOD",
    "ROLE",
    "HOMOLOGY_CLUSTER",
    "GROUND_TRUTH_PATH",
    "DECOY_SOURCE",
    "GENERATOR_VERSION",
    "PREP_PROTOCOL",
    "DATABASE_SNAPSHOT",
    "EXCLUSION_REASON",
    "STATUS",
]

MIXED_DATASET_MAP = {
    "1": ("PDB_MIXED_DRFOLD2", "DRFOLD2"),
    "2": ("EMRNA", "EMRNA"),
    "3": ("CASP16", "CASP16"),
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate an XLSX metadata table from pure and mixed RNA mmCIF files."
    )
    parser.add_argument(
        "--pure-dir",
        type=Path,
        default=DEFAULT_PURE_DIR,
        help=f"directory containing pure-RNA CIF files (default: {DEFAULT_PURE_DIR})",
    )
    parser.add_argument(
        "--mixed-dir",
        type=Path,
        default=DEFAULT_MIXED_DIR,
        help=f"directory containing mixed CIF files (default: {DEFAULT_MIXED_DIR})",
    )
    parser.add_argument(
        "--exclusion-xlsx",
        type=Path,
        default=DEFAULT_EXCLUSION_XLSX,
        help="XLSX whose first sheet/column lists pure-data PDB IDs to exclude",
    )
    parser.add_argument(
        "--record-xlsx",
        type=Path,
        default=DEFAULT_RECORD_XLSX,
        help="XLSX whose second sheet contains mixed-data assignments in column 6",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT_XLSX,
        help=f"output XLSX path (default: {DEFAULT_OUTPUT_XLSX})",
    )
    parser.add_argument(
        "--skip-count-check",
        action="store_true",
        help="do not require the expected 2246/5/2241/77 source-file counts",
    )
    return parser.parse_args()


def normalize_pdb_id(value: Any, context: str) -> str:
    if value is None:
        return ""
    pdb_id = str(value).strip().upper()
    if not pdb_id:
        return ""
    if not PDB_ID_PATTERN.fullmatch(pdb_id):
        raise ValueError(f"{context} 中的 PDB ID 非法：{value!r}")
    return pdb_id


def normalize_dataset_code(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def load_exclusion_ids(path: Path) -> set[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        result: set[str] = set()
        for row_number, (value,) in enumerate(
            worksheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True),
            start=2,
        ):
            pdb_id = normalize_pdb_id(value, f"{path.name} 第 {row_number} 行")
            if pdb_id:
                result.add(pdb_id)
        if not result:
            raise ValueError(f"排除表未读到任何 PDB ID：{path}")
        return result
    finally:
        workbook.close()


def load_mixed_assignments(path: Path) -> dict[str, tuple[str, str]]:
    """Map each OUT PDB ID to (DATASET, DECOY_SOURCE) using sheet 2, column 6."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if len(workbook.worksheets) < 2:
            raise ValueError(f"记录表少于两张工作表：{path}")

        worksheet = workbook.worksheets[1]
        assignments: dict[str, tuple[str, str]] = {}
        source_codes: dict[str, str] = {}

        for row_number, row in enumerate(
            worksheet.iter_rows(
                min_row=2,
                min_col=1,
                max_col=6,
                values_only=True,
            ),
            start=2,
        ):
            pdb_id = normalize_pdb_id(
                row[0], f"{path.name} 第二张表第 {row_number} 行"
            )
            in_or_out = "" if row[3] is None else str(row[3]).strip().upper()
            if not pdb_id or in_or_out != "OUT":
                continue

            code = normalize_dataset_code(row[5])
            if code not in MIXED_DATASET_MAP:
                raise ValueError(
                    f"{path.name} 第二张表第 {row_number} 行："
                    f"PDB {pdb_id} 的第六列值 {row[5]!r} 没有映射规则"
                )

            if pdb_id in source_codes and source_codes[pdb_id] != code:
                raise ValueError(
                    f"PDB {pdb_id} 在记录表中存在冲突的第六列值："
                    f"{source_codes[pdb_id]!r} 和 {code!r}"
                )

            source_codes[pdb_id] = code
            assignments[pdb_id] = MIXED_DATASET_MAP[code]

        if not assignments:
            raise ValueError(f"记录表第二张表未读到任何 OUT PDB ID：{path}")
        return assignments
    finally:
        workbook.close()


def index_cif_files(directory: Path) -> dict[str, Path]:
    result: dict[str, Path] = {}
    for path in sorted(directory.iterdir(), key=lambda item: item.name.lower()):
        if not path.is_file() or path.suffix.lower() != ".cif":
            continue
        pdb_id = normalize_pdb_id(path.stem, f"文件名 {path.name}")
        if pdb_id in result:
            raise ValueError(
                f"目录 {directory} 中存在重复 PDB ID 文件："
                f"{result[pdb_id].name} 和 {path.name}"
            )
        result[pdb_id] = path
    return result


def clean_cif_value(value: Any) -> str:
    if value is None or value is False:
        return ""
    text = cif.as_string(str(value)).strip()
    return "" if text in NULL_CIF_VALUES else text


def category_column(
    category: dict[str, list[Any]], name: str, row_count: int
) -> list[Any]:
    values = category.get(name, [])
    if not values:
        return [""] * row_count
    if len(values) != row_count:
        raise ValueError(
            f"mmCIF category 列长度不一致：{name}={len(values)}, expected={row_count}"
        )
    return values


def first_cif_value(block: Any, tags: Iterable[str]) -> str:
    for tag in tags:
        values = block.find_values(tag)
        if values:
            value = clean_cif_value(values[0])
            if value:
                return value
    return ""


def normalize_sequence(value: Any) -> str:
    return re.sub(r"\s+", "", clean_cif_value(value))


def sequence_length(sequence: str) -> int:
    """Count a parenthesized modified monomer as one residue."""
    return len(re.findall(r"\([^)]*\)|[A-Za-z?]", sequence))


def extract_release_date(block: Any) -> str:
    category = block.get_mmcif_category("_pdbx_audit_revision_history.")
    if category:
        dates = category.get("revision_date", [])
        ordinals = category_column(category, "ordinal", len(dates)) if dates else []
        candidates: list[tuple[int, str]] = []
        for ordinal, date in zip(ordinals, dates):
            clean_date = clean_cif_value(date)
            if not clean_date:
                continue
            try:
                order = int(float(clean_cif_value(ordinal)))
            except (TypeError, ValueError):
                order = 10**9
            candidates.append((order, clean_date))
        if candidates:
            return min(candidates, key=lambda item: (item[0], item[1]))[1]

    return first_cif_value(block, ["_database_PDB_rev.date_original"])


def extract_methods(block: Any) -> str:
    methods: list[str] = []
    seen: set[str] = set()
    for raw in block.find_values("_exptl.method"):
        method = clean_cif_value(raw)
        if method and method not in seen:
            methods.append(method)
            seen.add(method)
    return "; ".join(methods)


def extract_rna_chain_rows(path: Path, expected_pdb_id: str) -> list[dict[str, Any]]:
    document = cif.read_file(str(path), check_level=2)
    if len(document) != 1:
        raise ValueError(f"预期 1 个 data block，实际 {len(document)} 个")
    block = document[0]

    entry_id = first_cif_value(block, ["_entry.id"]).upper()
    if not entry_id:
        raise ValueError("缺少 _entry.id")
    if entry_id != expected_pdb_id:
        raise ValueError(
            f"_entry.id={entry_id!r} 与文件名 PDB ID {expected_pdb_id!r} 不一致"
        )

    release_date = extract_release_date(block)
    if not release_date:
        raise ValueError("缺少初始发布日期")

    experiment_method = extract_methods(block)
    if not experiment_method:
        raise ValueError("缺少 _exptl.method")

    category = block.get_mmcif_category("_entity_poly.")
    entity_ids = category.get("entity_id", [])
    if not entity_ids:
        raise ValueError("缺少 _entity_poly")

    row_count = len(entity_ids)
    polymer_types = category_column(category, "type", row_count)
    strands = category_column(category, "pdbx_strand_id", row_count)
    reported_sequences = category_column(
        category, "pdbx_seq_one_letter_code", row_count
    )
    canonical_sequences = category_column(
        category, "pdbx_seq_one_letter_code_can", row_count
    )

    rows: list[dict[str, Any]] = []
    for entity_id, polymer_type, chain_text, reported, canonical in zip(
        entity_ids,
        polymer_types,
        strands,
        reported_sequences,
        canonical_sequences,
    ):
        if clean_cif_value(polymer_type).lower() != RNA_POLYMER_TYPE:
            continue

        entity_id_text = clean_cif_value(entity_id)
        sequence = normalize_sequence(canonical) or normalize_sequence(reported)
        if not sequence:
            raise ValueError(f"RNA entity {entity_id_text!r} 缺少序列")
        length = sequence_length(sequence)
        if length <= 0:
            raise ValueError(f"RNA entity {entity_id_text!r} 的序列长度为 0")

        chain_ids: list[str] = []
        for item in clean_cif_value(chain_text).split(","):
            chain_id = item.strip()
            if chain_id and chain_id not in chain_ids:
                chain_ids.append(chain_id)
        if not chain_ids:
            raise ValueError(f"RNA entity {entity_id_text!r} 缺少 chain ID")

        for chain_id in chain_ids:
            rows.append(
                {
                    "PDB_ID": expected_pdb_id,
                    "CHAIN_IDS": chain_id,
                    "SEQUENCE": sequence,
                    "LENGTH": length,
                    "RELEASE_DATE": release_date,
                    "EXPERIMENT_METHOD": experiment_method,
                }
            )

    if not rows:
        raise ValueError("未找到 type=polyribonucleotide 的 RNA entity")
    return rows


def write_error_report(path: Path, errors: list[tuple[str, str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle, delimiter="\t")
        writer.writerow(["SOURCE", "PDB_ID", "ERROR"])
        writer.writerows(errors)


def write_workbook(path: Path, rows: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "RNA_METADATA"
    worksheet.append(OUTPUT_COLUMNS)

    for row in rows:
        worksheet.append([row.get(column) for column in OUTPUT_COLUMNS])

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {
        "TARGET_ID": 18,
        "DATASET": 22,
        "PDB_ID": 12,
        "CHAIN_IDS": 14,
        "SEQUENCE": 55,
        "LENGTH": 12,
        "RELEASE_DATE": 16,
        "EXPERIMENT_METHOD": 32,
        "ROLE": 14,
        "HOMOLOGY_CLUSTER": 22,
        "GROUND_TRUTH_PATH": 48,
        "DECOY_SOURCE": 20,
        "GENERATOR_VERSION": 22,
        "PREP_PROTOCOL": 18,
        "DATABASE_SNAPSHOT": 24,
        "EXCLUSION_REASON": 22,
        "STATUS": 14,
    }
    for index, column in enumerate(OUTPUT_COLUMNS, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = widths[column]

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.showGridLines = False

    temporary = path.with_name(f".{path.stem}.tmp{path.suffix}")
    try:
        workbook.save(temporary)
        os.replace(temporary, path)
    finally:
        workbook.close()
        try:
            temporary.unlink()
        except FileNotFoundError:
            pass


def validate_paths(paths: Iterable[tuple[str, Path, str]]) -> None:
    for label, path, kind in paths:
        if kind == "dir" and not path.is_dir():
            raise FileNotFoundError(f"{label}目录不存在：{path}")
        if kind == "file" and not path.is_file():
            raise FileNotFoundError(f"{label}文件不存在：{path}")


def main() -> int:
    args = parse_args()
    pure_dir = args.pure_dir.expanduser().resolve()
    mixed_dir = args.mixed_dir.expanduser().resolve()
    exclusion_xlsx = args.exclusion_xlsx.expanduser().resolve()
    record_xlsx = args.record_xlsx.expanduser().resolve()
    output_xlsx = args.output.expanduser().resolve()
    error_report = output_xlsx.with_suffix(".errors.tsv")

    try:
        validate_paths(
            [
                ("纯 RNA CIF ", pure_dir, "dir"),
                ("mixed CIF ", mixed_dir, "dir"),
                ("排除表 ", exclusion_xlsx, "file"),
                ("记录表 ", record_xlsx, "file"),
            ]
        )
        if not output_xlsx.parent.is_dir():
            raise FileNotFoundError(f"输出目录不存在：{output_xlsx.parent}")

        exclusions = load_exclusion_ids(exclusion_xlsx)
        mixed_assignments = load_mixed_assignments(record_xlsx)
        all_pure_files = index_cif_files(pure_dir)
        mixed_files = index_cif_files(mixed_dir)

        missing_excluded = exclusions - set(all_pure_files)
        if missing_excluded:
            raise ValueError(
                "排除表中的以下 PDB ID 在纯 RNA 目录中不存在："
                + ", ".join(sorted(missing_excluded))
            )

        pure_files = {
            pdb_id: path
            for pdb_id, path in all_pure_files.items()
            if pdb_id not in exclusions
        }

        missing_assignments = set(mixed_files) - set(mixed_assignments)
        extra_assignments = set(mixed_assignments) - set(mixed_files)
        if missing_assignments:
            raise ValueError(
                "以下 mixed CIF 在记录表 OUT 行中没有第六列映射："
                + ", ".join(sorted(missing_assignments))
            )
        if extra_assignments:
            raise ValueError(
                "记录表存在 OUT PDB ID，但 mixed 目录中没有对应 CIF："
                + ", ".join(sorted(extra_assignments))
            )

        overlap = set(pure_files) & set(mixed_files)
        if overlap:
            raise ValueError(
                "排除后 pure 与 mixed 目录仍有重复 PDB ID："
                + ", ".join(sorted(overlap))
            )

        counts = {
            "pure_total": len(all_pure_files),
            "excluded": len(exclusions),
            "pure_used": len(pure_files),
            "mixed": len(mixed_files),
        }
        print(
            "文件计数："
            f"pure 总数={counts['pure_total']}，排除={counts['excluded']}，"
            f"pure 使用={counts['pure_used']}，mixed={counts['mixed']}",
            flush=True,
        )

        if not args.skip_count_check:
            expected = {
                "pure_total": EXPECTED_PURE_TOTAL,
                "excluded": EXPECTED_EXCLUDED,
                "pure_used": EXPECTED_PURE_USED,
                "mixed": EXPECTED_MIXED,
            }
            mismatches = [
                f"{key}: 实际 {counts[key]}，预期 {expected[key]}"
                for key in expected
                if counts[key] != expected[key]
            ]
            if mismatches:
                raise ValueError("文件数量校验失败；" + "；".join(mismatches))

    except Exception as error:
        print(f"[错误] 输入校验失败：{error}", file=sys.stderr, flush=True)
        return 2

    errors: list[tuple[str, str, str]] = []
    rows: list[dict[str, Any]] = []

    for source_kind, files in (("pure", pure_files), ("mixed", mixed_files)):
        # Process files individually so all bad files are reported in one run.
        for index, (pdb_id, path) in enumerate(sorted(files.items()), start=1):
            try:
                if source_kind == "pure":
                    dataset, decoy_source = "PDB", "protenix"
                    display_path = f"~/pdb_data/{path.name}"
                else:
                    dataset, decoy_source = mixed_assignments[pdb_id]
                    display_path = f"~/pdb_data_mixed/{path.name}"

                for cif_row in extract_rna_chain_rows(path, pdb_id):
                    chain_id = cif_row["CHAIN_IDS"]
                    rows.append(
                        {
                            "TARGET_ID": f"{pdb_id}_{chain_id}",
                            "DATASET": dataset,
                            "PDB_ID": pdb_id,
                            "CHAIN_IDS": chain_id,
                            "SEQUENCE": cif_row["SEQUENCE"],
                            "LENGTH": cif_row["LENGTH"],
                            "RELEASE_DATE": cif_row["RELEASE_DATE"],
                            "EXPERIMENT_METHOD": cif_row["EXPERIMENT_METHOD"],
                            "ROLE": None,
                            "HOMOLOGY_CLUSTER": None,
                            "GROUND_TRUTH_PATH": display_path,
                            "DECOY_SOURCE": decoy_source,
                            "GENERATOR_VERSION": None,
                            "PREP_PROTOCOL": "OLD",
                            "DATABASE_SNAPSHOT": None,
                            "EXCLUSION_REASON": None,
                            "STATUS": None,
                        }
                    )
            except Exception as error:
                errors.append((source_kind, pdb_id, f"{type(error).__name__}: {error}"))
                print(
                    f"[失败] {source_kind} {pdb_id}: {type(error).__name__}: {error}",
                    file=sys.stderr,
                    flush=True,
                )

            if index % 100 == 0 or index == len(files):
                print(
                    f"[{source_kind}] 已解析 {index}/{len(files)} 个 CIF；"
                    f"当前总记录 {len(rows)} 行；失败文件 {len(errors)} 个",
                    flush=True,
                )

    if errors:
        write_error_report(error_report, errors)
        print(
            f"[错误] 有 {len(errors)} 个 CIF 解析失败，未生成最终 XLSX。"
            f"错误清单：{error_report}",
            file=sys.stderr,
            flush=True,
        )
        return 1

    target_counts = Counter(row["TARGET_ID"] for row in rows)
    duplicate_targets = sorted(
        target_id for target_id, count in target_counts.items() if count > 1
    )
    if duplicate_targets:
        print(
            "[错误] 发现重复 TARGET_ID：" + ", ".join(duplicate_targets[:20]),
            file=sys.stderr,
            flush=True,
        )
        return 1

    rows.sort(key=lambda row: (row["PDB_ID"], row["CHAIN_IDS"]))
    write_workbook(output_xlsx, rows)
    try:
        error_report.unlink()
    except FileNotFoundError:
        pass

    dataset_counts = Counter(row["DATASET"] for row in rows)
    print("\n========== 生成完成 ==========", flush=True)
    print(f"输入 CIF：{len(pure_files) + len(mixed_files)}", flush=True)
    print(f"输出 RNA 链记录：{len(rows)}", flush=True)
    for dataset in sorted(dataset_counts):
        print(f"DATASET {dataset}: {dataset_counts[dataset]} 行", flush=True)
    print(f"输出文件：{output_xlsx}", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
