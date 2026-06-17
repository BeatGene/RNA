import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ============ 配置 ============
REPORT_PATH = r"C:\Users\49586\Desktop\Learning\Laboratory\Admis\graduate_first\RNA\Code\check\check_all_report.txt"
XLSX_PATH = r"C:\Users\49586\Desktop\Learning\Laboratory\Admis\graduate_first\RNA\Code\Download_PDB_RAW\rna_experimental_pdb_ids.xlsx"
# =============================


def parse_report(report_path):
    """
    从 check_all_report.txt 提取:
      - all_ids: 所有 1979 个纯 RNA 的 PDB ID
      - failed_ids: 其中 Pred 失败的 PDB ID
    """
    all_ids = set()
    failed_ids = set()

    with open(report_path, "r", encoding="utf-8") as f:
        section = None  # 'failed' / 'passed' / None
        for raw_line in f:
            line = raw_line.rstrip("\n").rstrip("\r")  # 不去掉前导空格！

            if "--- Pred 异常" in line:
                section = "failed"
                continue
            if "--- 全部通过" in line:
                section = "passed"
                continue
            # 遇到下一个 section 头
            if section and line.startswith("---"):
                section = None
                continue

            if section == "failed":
                # 格式: "  1ddy:"  (缩进 + 小写ID + 冒号)
                m = re.match(r"^\s+([0-9a-z]{4}):\s*$", line)
                if m:
                    pid = m.group(1).upper()
                    all_ids.add(pid)
                    failed_ids.add(pid)

            elif section == "passed":
                # 格式: "  9if0"  (缩进 + 小写ID)
                m = re.match(r"^\s+([0-9a-z]{4})\s*$", line)
                if m:
                    pid = m.group(1).upper()
                    all_ids.add(pid)

    return all_ids, failed_ids


def main():
    all_1979_ids, failed_ids = parse_report(REPORT_PATH)
    passed_ids = all_1979_ids - failed_ids
    print(f"报告中纯 RNA 总数: {len(all_1979_ids)}")
    print(f"  成功: {len(passed_ids)}")
    print(f"  失败: {len(failed_ids)}")

    # 读取 xlsx
    wb = load_workbook(XLSX_PATH)
    ws1 = wb.active

    # pandas 辅助获取准确行数（避免 max_row 问题）
    df = pd.read_excel(XLSX_PATH, engine="openpyxl")
    df["PDB_ID"] = df["PDB_ID"].astype(str).str.strip().str.upper()

    # 找出纯 RNA 对应的行索引
    pure_rows = df[df["PDB_ID"].isin(all_1979_ids)]
    print(f"xlsx 中匹配到的纯 RNA 行数: {len(pure_rows)}")

    # 构建 Sheet2
    if "Sheet2" in wb.sheetnames:
        del wb["Sheet2"]
    ws2 = wb.create_sheet("Sheet2")

    # 复制表头（列1-3）+ 新增列4
    headers = ["PDB_ID", "Previously_Downloaded", "Has_CIF_File", "Prep_Pred_Status"]
    for col_idx, h in enumerate(headers, 1):
        ws2.cell(row=1, column=col_idx, value=h)

    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    red_fill = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")

    pure_row_idx = 2  # Sheet2 从第 2 行开始写数据
    for _, s1_row in pure_rows.iterrows():
        pdb_id = s1_row["PDB_ID"]

        # 列1-3: 从 Sheet1 复制值（pandas 已读了）
        ws2.cell(row=pure_row_idx, column=1, value=pdb_id)
        ws2.cell(row=pure_row_idx, column=2, value=s1_row.get("Previously_Downloaded", ""))
        ws2.cell(row=pure_row_idx, column=3, value=s1_row.get("Has_CIF_File", ""))

        # 列4: Prep_Pred 状态
        cell4 = ws2.cell(row=pure_row_idx, column=4)
        if pdb_id in failed_ids:
            cell4.value = "NO"
            cell4.fill = red_fill
        else:
            cell4.value = "YES"
            cell4.fill = green_fill

        pure_row_idx += 1

    wb.save(XLSX_PATH)
    print(f"\nSheet2 已创建，共 {pure_row_idx - 2} 行")
    print(f"  🟢 YES: {len(passed_ids)}")
    print(f"  🔴 NO : {len(failed_ids)}")


if __name__ == "__main__":
    main()
