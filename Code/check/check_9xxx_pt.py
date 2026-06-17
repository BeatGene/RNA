import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ============ 配置 ============
CIF_DIR = "/remote-home/jinxianwang/tinghaoxia/RNA/Data/pdb_data/01_Pure_RNA"
RNA_DIR = "/remote-home/jinxianwang/tinghaoxia/RNA/Data/RNA"
SPLITS = ["train", "val", "test"]
OUTPUT = "pure_rna_9xxx_pt_count.xlsx"
# =============================

# 1. 抓取以 9 开头的 PDB ID（从 .cif 文件名）
all_cif = [f.replace(".cif", "") for f in os.listdir(CIF_DIR) if f.endswith(".cif")]
pdb_9xxx = sorted([cid for cid in all_cif if cid.startswith("9")])
print(f"01_Pure_RNA 中以 9 开头的 CIF 文件: {len(pdb_9xxx)} 个")

# 2. 构建 train/val/test 下所有文件夹名→路径的映射（不区分大小写）
folder_map = {}  # lowercase name → actual folder path
for split in SPLITS:
    split_dir = os.path.join(RNA_DIR, split)
    if not os.path.exists(split_dir):
        continue
    for item in os.listdir(split_dir):
        item_path = os.path.join(split_dir, item)
        if os.path.isdir(item_path):
            folder_map[item.lower()] = item_path

print(f"train/val/test 下共 {len(folder_map)} 个文件夹")

# 3. 逐 ID 检查
records = []
for pdb_id in pdb_9xxx:
    low = pdb_id.lower()
    folder_path = folder_map.get(low)

    if folder_path:
        exists = "YES"
        pt_count = len([f for f in os.listdir(folder_path) if f.endswith(".pt")])
    else:
        exists = "NO"
        pt_count = 0

    records.append({"PDB_ID": pdb_id, "Has_Folder": exists, "PT_Count": pt_count})

# 4. 写入 xlsx
df = pd.DataFrame(records)
df.to_excel(OUTPUT, index=False, engine="openpyxl")

# 5. 上色
wb = load_workbook(OUTPUT)
ws = wb.active

green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
red_fill = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")

num_rows = len(df)
for row_idx in range(2, num_rows + 2):
    cell = ws.cell(row=row_idx, column=2)
    if cell.value == "YES":
        cell.fill = green_fill
    elif cell.value == "NO":
        cell.fill = red_fill

wb.save(OUTPUT)

# 6. 汇总
yes_count = sum(1 for r in records if r["Has_Folder"] == "YES")
no_count = len(records) - yes_count
total_pt = sum(r["PT_Count"] for r in records)
print(f"\n{'='*40}")
print(f"结果已保存至: {OUTPUT}")
print(f"  有文件夹 (YES): {yes_count} 个")
print(f"  无文件夹 (NO) : {no_count} 个")
print(f"  .pt 文件总数  : {total_pt}")
