import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ============ 配置 ============
CIF_DIR = "/remote-home/jinxianwang/tinghaoxia/RNA/Data/pdb_data/01_Pure_RNA"
RNA_DIR = "/remote-home/jinxianwang/tinghaoxia/RNA/Data/RNA"
SPLITS = ["train", "val", "test"]

# 脚本所在目录
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Excel 参考文件（相对于脚本目录向上两级）
EXCEL_REF = os.path.join(SCRIPT_DIR, "..", "Download_PDB_RAW", "rna_experimental_pdb_ids.xlsx")

# 输出文件：与脚本同文件夹
OUTPUT = os.path.join(SCRIPT_DIR, "pure_rna_all_pt_count.xlsx")
# =============================

# 1. 读取 Excel 中 Sheet2 第一列的 PDB_ID 顺序
df_ref = pd.read_excel(EXCEL_REF, sheet_name="Sheet2")
pdb_order = df_ref.iloc[:, 0].dropna().astype(str).tolist()
print(f"Excel 中共 {len(pdb_order)} 个 PDB_ID (1xxx～9xxx)")

# 2. 获取所有 .cif 文件（用于判断 CIF 是否存在）
all_cif = set(f.replace(".cif", "") for f in os.listdir(CIF_DIR) if f.endswith(".cif"))
print(f"01_Pure_RNA 中 CIF 文件总数: {len(all_cif)}")

# 3. 构建 train/val/test 下所有文件夹名 → 路径的映射（不区分大小写）
folder_map = {}  # lowercase name → actual folder path
for split in SPLITS:
    split_dir = os.path.join(RNA_DIR, split)
    if not os.path.exists(split_dir):
        continue
    for item in os.listdir(split_dir):
        item_path = os.path.join(split_dir, item)
        if os.path.isdir(item_path):
            folder_map[item.lower()] = item_path

print(f"train/val/test 下共 {len(folder_map)} 个文件夹")

# 4. 按 Excel 中的顺序逐 ID 检查
records = []
for pdb_id in pdb_order:
    low = pdb_id.lower()
    folder_path = folder_map.get(low)

    has_cif = "YES" if pdb_id in all_cif else "NO"

    if folder_path:
        has_folder = "YES"
        pt_count = len([f for f in os.listdir(folder_path) if f.endswith(".pt")])
    else:
        has_folder = "NO"
        pt_count = 0

    records.append({
        "PDB_ID": pdb_id,
        "Has_CIF": has_cif,
        "Has_Folder": has_folder,
        "PT_Count": pt_count,
    })

# 5. 写入 xlsx
df_out = pd.DataFrame(records)
df_out.to_excel(OUTPUT, index=False, engine="openpyxl")

# 6. 上色
wb = load_workbook(OUTPUT)
ws = wb.active

green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
red_fill = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")

num_rows = len(df_out)
for row_idx in range(2, num_rows + 2):
    # Has_CIF 列 (第2列)
    cif_cell = ws.cell(row=row_idx, column=2)
    if cif_cell.value == "YES":
        cif_cell.fill = green_fill
    elif cif_cell.value == "NO":
        cif_cell.fill = red_fill

    # Has_Folder 列 (第3列)
    folder_cell = ws.cell(row=row_idx, column=3)
    if folder_cell.value == "YES":
        folder_cell.fill = green_fill
    elif folder_cell.value == "NO":
        folder_cell.fill = red_fill

wb.save(OUTPUT)

# 7. 汇总
cif_yes = sum(1 for r in records if r["Has_CIF"] == "YES")
cif_no = len(records) - cif_yes
folder_yes = sum(1 for r in records if r["Has_Folder"] == "YES")
folder_no = len(records) - folder_yes
total_pt = sum(r["PT_Count"] for r in records)

print(f"\n{'='*50}")
print(f"结果已保存至: {OUTPUT}")
print(f"  PDB_ID 总数       : {len(records)}")
print(f"  CIF 存在 (YES)    : {cif_yes}")
print(f"  CIF 缺失 (NO)     : {cif_no}")
print(f"  文件夹存在 (YES)  : {folder_yes}")
print(f"  文件夹缺失 (NO)   : {folder_no}")
print(f"  .pt 文件总数      : {total_pt}")
