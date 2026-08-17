import csv
import io
import tempfile
import unittest
from contextlib import redirect_stdout
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from build_mapping_targets_xlsx import (
    OUTPUT_COLUMNS,
    build_rows,
    main,
    scan_predictions,
)


SEEDS = (42, 66, 101, 2024, 8888)


def make_complete_target(root: Path, pdb_id: str) -> None:
    for seed in SEEDS:
        predictions = (
            root
            / f"pred_output_{pdb_id.lower()}_seed_{seed}"
            / pdb_id.lower()
            / f"seed_{seed}"
            / "predictions"
        )
        predictions.mkdir(parents=True)
        for sample in range(5):
            (predictions / f"{pdb_id.lower()}_sample_{sample}.cif").write_text(
                "data_test\n", encoding="utf-8"
            )


class BuildMappingTargetsXlsxTests(unittest.TestCase):
    def test_scan_requires_all_seeds_and_samples(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            make_complete_target(root, "1ABC")
            broken = root / "pred_output_2def_seed_42" / "2def" / "seed_42" / "predictions"
            broken.mkdir(parents=True)
            (broken / "2def_sample_0.cif").write_text("data_test\n", encoding="utf-8")

            scan = scan_predictions(root, expected_seeds=SEEDS, expected_samples=5)

            self.assertEqual(scan.all_target_ids, ("1ABC", "2DEF"))
            self.assertEqual(scan.complete_target_ids, ("1ABC",))
            self.assertIn("2DEF", scan.incomplete_reasons)

    def test_build_rows_uses_strict_cutoff(self):
        rows = build_rows(
            ["2DEF", "1ABC"],
            release_dates={
                "1ABC": date(2021, 9, 30),
                "2DEF": date(2021, 10, 1),
            },
            chain_counts={"1ABC": 1, "2DEF": 3},
            cutoff=date(2021, 9, 30),
        )
        self.assertEqual(rows[0][0], "1ABC")
        self.assertEqual(rows[0][3], "pre_or_on_cutoff")
        self.assertEqual(rows[1][3], "post_cutoff")
        self.assertIsNone(rows[0][5])
        self.assertIsNone(rows[0][6])

    def test_main_writes_requested_workbook(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            pred_dir = root / "predictions"
            make_complete_target(pred_dir, "1ABC")
            make_complete_target(pred_dir, "2DEF")
            incomplete = (
                pred_dir
                / "pred_output_3ghi_seed_42"
                / "3ghi"
                / "seed_42"
                / "predictions"
            )
            incomplete.mkdir(parents=True)
            (incomplete / "3ghi_sample_0.cif").write_text(
                "data_test\n", encoding="utf-8"
            )

            date_audit = root / "date_audit.tsv"
            with date_audit.open("w", encoding="utf-8", newline="") as handle:
                writer = csv.writer(handle, delimiter="\t")
                writer.writerow(["PDB_ID", "RELEASE_DATE"])
                writer.writerow(["1ABC", "2021-09-30"])
                writer.writerow(["2DEF", "2022-01-02"])
                writer.writerow(["3GHI", "2023-03-04"])

            manifest = root / "pdb_cif_manifest.csv"
            with manifest.open("w", encoding="utf-8", newline="") as handle:
                writer = csv.writer(handle)
                writer.writerow(["PDB_ID", "RNA_CHAIN_COUNT"])
                writer.writerow(["1ABC", "1"])
                writer.writerow(["2DEF", "2"])
                writer.writerow(["3GHI", "3"])

            output = root / "mapping" / "xlsx" / "targets.xlsx"
            stdout = io.StringIO()
            with redirect_stdout(stdout):
                exit_code = main(
                    [
                        "--pred-dir",
                        str(pred_dir),
                        "--date-audit",
                        str(date_audit),
                        "--manifest",
                        str(manifest),
                        "--output",
                        str(output),
                    ]
                )
            self.assertEqual(exit_code, 0)
            workbook = load_workbook(output, data_only=True)
            worksheet = workbook["Targets"]
            self.assertEqual(
                tuple(cell.value for cell in worksheet[1]),
                OUTPUT_COLUMNS,
            )
            self.assertEqual(worksheet.max_row, 4)
            self.assertEqual(worksheet["A2"].value, "1ABC")
            self.assertEqual(worksheet["B3"].value, "2DEF")
            self.assertEqual(worksheet["C2"].value.date(), date(2021, 9, 30))
            self.assertEqual(worksheet["D2"].value, "pre_or_on_cutoff")
            self.assertEqual(worksheet["D3"].value, "post_cutoff")
            self.assertEqual(worksheet["E3"].value, 2)
            self.assertIsNone(worksheet["F2"].value)
            self.assertIsNone(worksheet["G2"].value)
            self.assertEqual(worksheet["A4"].value, "3GHI")
            log = stdout.getvalue()
            self.assertIn("3GHI: missing_seeds=[66, 101, 2024, 8888]", log)
            self.assertIn("seed_42: primary_cif_count=1/5", log)
            self.assertIn("Final unique PDB_id count written: 3", log)


if __name__ == "__main__":
    unittest.main()
