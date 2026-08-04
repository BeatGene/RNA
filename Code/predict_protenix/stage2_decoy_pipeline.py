#!/usr/bin/env python3
"""Audit and incrementally complete the Protenix RNA decoy pipeline.

The script is intended to run on the new laboratory server.  It never deletes
existing prep/prediction results.  A task is skipped only after its output has
been verified, rather than merely trusting a subprocess return code.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import queue
import re
import shutil
import subprocess
import sys
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import pandas as pd


PDB_ID_RE = re.compile(r"^[A-Z0-9_]+$")
PRED_FOLDER_RE = re.compile(r"^pred_output_(.+)_seed_(\d+)$", re.IGNORECASE)
PREP_FOLDER_RE = re.compile(r"^prep_output_(.+)$", re.IGNORECASE)
PRIMARY_CIF_RE = re.compile(r"^.+_sample_(\d+)\.cif$", re.IGNORECASE)
UNRESOLVED_CIF_RE = re.compile(
    r"^.+_sample_(\d+)_wounresol\.cif$", re.IGNORECASE
)
CONFIDENCE_RE = re.compile(
    r"^.+_summary_confidence_sample_(\d+)\.json$", re.IGNORECASE
)
TRUE_VALUES = {"1", "TRUE", "T", "YES", "Y"}
MODEL_NAME = "protenix_base_default_v1.0.0"


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def normalize_pdb_id(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    pdb_id = str(value).strip().upper()
    if not pdb_id or pdb_id in {"PDB_ID", "NAN", "NONE"}:
        return ""
    if not PDB_ID_RE.fullmatch(pdb_id):
        raise ValueError(f"非法 PDB ID：{value!r}")
    return pdb_id


def is_true(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().upper() in TRUE_VALUES


def atomic_write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.tmp")
    temporary.write_text(text, encoding="utf-8")
    os.replace(temporary, path)


def append_jsonl(path: Path, event: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(event, ensure_ascii=False) + "\n")


def sha256_file(path: Path, chunk_size: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(chunk_size), b""):
            digest.update(chunk)
    return digest.hexdigest()


def parse_seeds(text: str) -> list[int]:
    seeds = []
    for item in text.split(","):
        item = item.strip()
        if item:
            seeds.append(int(item))
    if not seeds or len(seeds) != len(set(seeds)):
        raise argparse.ArgumentTypeError("seeds 必须是不重复的逗号分隔整数")
    return seeds


def locate_executable(name_or_path: str) -> str:
    candidate = Path(name_or_path).expanduser()
    if candidate.parent != Path(".") or candidate.is_absolute():
        if not candidate.is_file():
            raise FileNotFoundError(f"找不到可执行文件：{candidate}")
        return str(candidate.resolve())
    located = shutil.which(name_or_path)
    if not located:
        raise FileNotFoundError(f"PATH 中找不到命令：{name_or_path}")
    return located


@dataclass(frozen=True)
class Target:
    pdb_id: str
    cif_path: str
    legacy_1979: bool


@dataclass
class JsonInfo:
    path: str = ""
    valid: bool = False
    error: str = ""
    task_name: str = ""
    rna_entries: int = 0
    rna_chains: int = 0
    sequences: list[str] | None = None
    msa_paths: list[str] | None = None


@dataclass
class PrepInfo:
    status: str
    reason: str
    updated_json: JsonInfo
    resolved_msa_paths: list[str]
    prep_dir: str
    a3m_found: int


@dataclass
class SeedInfo:
    status: str
    reason: str
    output_dir: str
    predictions_dir: str
    primary_count: int
    valid_primary_count: int
    invalid_primary_count: int
    unresolved_count: int
    confidence_count: int
    sample_indices: str


def load_targets(manifest_path: Path) -> list[Target]:
    if not manifest_path.is_file():
        raise FileNotFoundError(f"找不到第一阶段清单：{manifest_path}")
    frame = pd.read_csv(manifest_path, dtype=str, keep_default_na=False)
    required = {"PDB_ID", "CURRENT_TARGET"}
    missing = required - set(frame.columns)
    if missing:
        raise ValueError(f"第一阶段清单缺少列：{sorted(missing)}")

    targets: list[Target] = []
    seen: set[str] = set()
    for row in frame.to_dict("records"):
        if not is_true(row["CURRENT_TARGET"]):
            continue
        pdb_id = normalize_pdb_id(row["PDB_ID"])
        if not pdb_id or pdb_id in seen:
            continue
        seen.add(pdb_id)
        targets.append(
            Target(
                pdb_id=pdb_id,
                cif_path=str(row.get("FILE_PATH", "")).strip(),
                legacy_1979=is_true(row.get("LEGACY_1979", "")),
            )
        )
    if not targets:
        raise ValueError("第一阶段清单中没有 CURRENT_TARGET=True 的记录")
    return sorted(targets, key=lambda item: item.pdb_id)


def choose_indexed_path(paths: list[Path]) -> Path | None:
    if not paths:
        return None
    return sorted(paths, key=lambda p: (len(p.parts), len(str(p)), str(p)))[0]


def index_json_files(root: Path) -> tuple[dict[str, list[Path]], dict[str, list[Path]]]:
    raw: dict[str, list[Path]] = {}
    updated: dict[str, list[Path]] = {}
    if not root.is_dir():
        return raw, updated
    for path in root.rglob("*.json"):
        stem = path.stem
        lower = stem.lower()
        if lower.endswith("-final-updated"):
            pdb_id = normalize_pdb_id(stem[: -len("-final-updated")])
            updated.setdefault(pdb_id, []).append(path.resolve())
        elif "-update-msa" not in lower and "-runtime" not in lower:
            try:
                pdb_id = normalize_pdb_id(stem)
            except ValueError:
                continue
            raw.setdefault(pdb_id, []).append(path.resolve())
    return raw, updated


def index_output_dirs(
    root: Path,
) -> tuple[dict[str, list[Path]], dict[tuple[str, int], list[Path]]]:
    prep: dict[str, list[Path]] = {}
    pred: dict[tuple[str, int], list[Path]] = {}
    if not root.is_dir():
        return prep, pred
    for current, dirnames, _ in os.walk(root):
        kept: list[str] = []
        for dirname in dirnames:
            prep_match = PREP_FOLDER_RE.fullmatch(dirname)
            pred_match = PRED_FOLDER_RE.fullmatch(dirname)
            path = (Path(current) / dirname).resolve()
            if pred_match:
                pdb_id = normalize_pdb_id(pred_match.group(1))
                pred.setdefault((pdb_id, int(pred_match.group(2))), []).append(path)
            elif prep_match:
                pdb_id = normalize_pdb_id(prep_match.group(1))
                prep.setdefault(pdb_id, []).append(path)
            else:
                kept.append(dirname)
        dirnames[:] = kept
    return prep, pred


def read_json_info(path: Path | None) -> JsonInfo:
    if path is None:
        return JsonInfo(error="文件不存在", sequences=[], msa_paths=[])
    info = JsonInfo(path=str(path), sequences=[], msa_paths=[])
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
        if not isinstance(payload, list) or not payload:
            raise ValueError("顶层必须是非空 JSON 列表")
        task = payload[0]
        if not isinstance(task, dict) or not isinstance(task.get("sequences"), list):
            raise ValueError("缺少 sequences 列表")
        info.task_name = str(task.get("name", path.stem)).strip()
        for sequence_item in task["sequences"]:
            if not isinstance(sequence_item, dict) or "rnaSequence" not in sequence_item:
                continue
            rna = sequence_item["rnaSequence"]
            sequence = str(rna.get("sequence", "")).strip().upper()
            if not sequence:
                raise ValueError("rnaSequence.sequence 为空")
            count = int(rna.get("count", 1))
            if count < 1:
                raise ValueError("rnaSequence.count 必须大于 0")
            info.rna_entries += 1
            info.rna_chains += count
            info.sequences.append(sequence)
            info.msa_paths.append(str(rna.get("unpairedMsaPath", "")).strip())
        if not info.rna_entries:
            raise ValueError("JSON 中没有 RNA 序列")
        info.valid = True
    except Exception as exc:
        info.error = f"{type(exc).__name__}: {exc}"
    return info


def a3m_files(prep_dir: Path | None) -> list[Path]:
    if prep_dir is None or not prep_dir.is_dir():
        return []
    return sorted(
        (
            path.resolve()
            for path in prep_dir.rglob("rna_msa.a3m")
            if path.is_file() and path.stat().st_size > 0
        ),
        key=str,
    )


def resolve_msa_paths(
    pdb_id: str,
    updated: JsonInfo,
    prep_dir: Path | None,
) -> tuple[list[str], int, int, int]:
    """Return runtime paths and direct-valid/rebased/missing counts."""
    found = a3m_files(prep_dir)
    by_index: dict[int, Path] = {}
    for path in found:
        try:
            index = int(path.parent.name)
        except ValueError:
            continue
        by_index.setdefault(index, path)

    resolved: list[str] = []
    direct_valid = rebased = missing = 0
    for index, text in enumerate(updated.msa_paths or []):
        source = Path(text).expanduser() if text else None
        if source and source.is_file() and source.stat().st_size > 0:
            resolved.append(str(source.resolve()))
            direct_valid += 1
            continue
        replacement = by_index.get(index)
        if replacement is None and len(found) == len(updated.msa_paths or []):
            replacement = found[index]
        if replacement is not None:
            resolved.append(str(replacement))
            rebased += 1
        else:
            resolved.append("")
            missing += 1
    return resolved, direct_valid, rebased, missing


def inspect_prep(
    pdb_id: str,
    updated_path: Path | None,
    prep_dir: Path | None,
) -> PrepInfo:
    updated = read_json_info(updated_path)
    found = a3m_files(prep_dir)
    if updated_path is None:
        return PrepInfo(
            "MISSING_UPDATED_JSON",
            "缺少 *-final-updated.json",
            updated,
            [],
            str(prep_dir or ""),
            len(found),
        )
    if not updated.valid:
        return PrepInfo(
            "INVALID_UPDATED_JSON",
            updated.error,
            updated,
            [],
            str(prep_dir or ""),
            len(found),
        )
    resolved, direct, rebased, missing = resolve_msa_paths(
        pdb_id, updated, prep_dir
    )
    if missing:
        status = "INCOMPLETE_MSA"
        reason = (
            f"RNA条目={updated.rna_entries}，有效引用={direct}，"
            f"可重定位={rebased}，缺失={missing}"
        )
    elif rebased:
        status = "COMPLETE_REBASABLE"
        reason = f"{rebased} 个旧绝对路径可重定位；不会覆盖原 JSON"
    else:
        status = "COMPLETE"
        reason = "OK"
    return PrepInfo(
        status,
        reason,
        updated,
        resolved,
        str(prep_dir or ""),
        len(found),
    )


def quick_validate_cif(path: Path) -> tuple[bool, str]:
    try:
        size = path.stat().st_size
        if size < 128:
            return False, f"文件过小({size} bytes)"
        with path.open("rb") as handle:
            prefix = handle.read(8192).lower()
        if b"data_" not in prefix:
            return False, "文件头缺少 data_"
        return True, "OK"
    except OSError as exc:
        return False, f"{type(exc).__name__}: {exc}"


def full_validate_cif(path: Path) -> tuple[bool, str]:
    valid, reason = quick_validate_cif(path)
    if not valid:
        return valid, reason
    try:
        import gemmi

        document = gemmi.cif.read_file(str(path), check_level=2)
        if len(document) != 1:
            return False, f"data block 数量={len(document)}"
        block = document[0]
        if not block.find_values("_atom_site.Cartn_x"):
            return False, "缺少 atom_site 坐标"
        return True, "OK"
    except ImportError:
        raise RuntimeError("full 校验需要安装 gemmi")
    except Exception as exc:
        return False, f"{type(exc).__name__}: {exc}"


def inspect_seed(
    output_dir: Path | None,
    expected_samples: int,
    validation: str,
) -> SeedInfo:
    empty = SeedInfo(
        "MISSING",
        "缺少输出目录",
        str(output_dir or ""),
        "",
        0,
        0,
        0,
        0,
        0,
        "",
    )
    if output_dir is None or not output_dir.is_dir():
        return empty
    prediction_dirs = sorted(
        (path for path in output_dir.rglob("predictions") if path.is_dir()),
        key=str,
    )
    if len(prediction_dirs) != 1:
        empty.status = "INVALID_LAYOUT"
        empty.reason = f"predictions 目录数量={len(prediction_dirs)}，预期 1"
        return empty

    pred_dir = prediction_dirs[0]
    primary: dict[int, list[Path]] = {}
    unresolved: set[int] = set()
    confidence: set[int] = set()
    for path in pred_dir.iterdir():
        if not path.is_file():
            continue
        unresolved_match = UNRESOLVED_CIF_RE.fullmatch(path.name)
        if unresolved_match:
            unresolved.add(int(unresolved_match.group(1)))
            continue
        primary_match = PRIMARY_CIF_RE.fullmatch(path.name)
        if primary_match:
            primary.setdefault(int(primary_match.group(1)), []).append(path)
            continue
        confidence_match = CONFIDENCE_RE.fullmatch(path.name)
        if confidence_match:
            confidence.add(int(confidence_match.group(1)))

    validator = full_validate_cif if validation == "full" else quick_validate_cif
    invalid_reasons: list[str] = []
    valid_count = 0
    for index, paths in sorted(primary.items()):
        if len(paths) != 1:
            invalid_reasons.append(f"sample_{index} 重复={len(paths)}")
            continue
        valid, reason = validator(paths[0])
        if valid:
            valid_count += 1
        else:
            invalid_reasons.append(f"{paths[0].name}: {reason}")

    expected_indices = set(range(expected_samples))
    actual_indices = set(primary)
    missing_indices = sorted(expected_indices - actual_indices)
    extra_indices = sorted(actual_indices - expected_indices)
    primary_count = sum(len(paths) for paths in primary.values())
    if (
        not missing_indices
        and not extra_indices
        and primary_count == expected_samples
        and valid_count == expected_samples
    ):
        status = "COMPLETE"
        reason = "OK"
        if len(confidence) != expected_samples:
            reason = f"结构完整；confidence JSON={len(confidence)}/{expected_samples}"
    else:
        status = "INCOMPLETE"
        pieces = [
            f"主CIF={primary_count}/{expected_samples}",
            f"有效={valid_count}/{expected_samples}",
        ]
        if missing_indices:
            pieces.append(f"缺少sample={missing_indices[:10]}")
        if extra_indices:
            pieces.append(f"额外sample={extra_indices[:10]}")
        if invalid_reasons:
            pieces.append("; ".join(invalid_reasons[:3]))
        reason = "；".join(pieces)

    return SeedInfo(
        status=status,
        reason=reason,
        output_dir=str(output_dir),
        predictions_dir=str(pred_dir),
        primary_count=primary_count,
        valid_primary_count=valid_count,
        invalid_primary_count=primary_count - valid_count,
        unresolved_count=len(unresolved),
        confidence_count=len(confidence),
        sample_indices=",".join(str(i) for i in sorted(actual_indices)),
    )


def int_to_letters(index: int) -> str:
    if index < 1:
        raise ValueError("chain index must be >= 1")
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(ord("A") + remainder) + letters
    return letters


def make_chain_mapping_rows(
    targets: list[Target],
    chain_manifest: Path,
    raw_infos: dict[str, JsonInfo],
) -> list[dict[str, Any]]:
    if not chain_manifest.is_file():
        return []
    frame = pd.read_csv(chain_manifest, dtype=str, keep_default_na=False)
    required = {"PDB_ID", "CHAIN_ID", "SEQUENCE_CANONICAL"}
    if required - set(frame.columns):
        return []
    wanted = {target.pdb_id for target in targets}
    original: dict[str, list[dict[str, str]]] = {}
    for row in frame.to_dict("records"):
        pdb_id = normalize_pdb_id(row["PDB_ID"])
        if pdb_id in wanted:
            original.setdefault(pdb_id, []).append(row)

    rows: list[dict[str, Any]] = []
    for target in targets:
        info = raw_infos.get(target.pdb_id, JsonInfo(sequences=[]))
        orig_rows = original.get(target.pdb_id, [])
        by_sequence: dict[str, list[dict[str, str]]] = {}
        for row in orig_rows:
            sequence = str(row.get("SEQUENCE_CANONICAL", "")).replace("\n", "")
            by_sequence.setdefault(sequence.upper(), []).append(row)

        predicted_index = 0
        payload: list[tuple[str, int]] = []
        if info.valid and info.path:
            try:
                doc = json.loads(Path(info.path).read_text(encoding="utf-8"))[0]
                for item in doc["sequences"]:
                    if "rnaSequence" in item:
                        rna = item["rnaSequence"]
                        payload.append(
                            (
                                str(rna["sequence"]).upper(),
                                int(rna.get("count", 1)),
                            )
                        )
            except Exception:
                payload = []

        used_original: set[str] = set()
        for sequence, count in payload:
            candidates = [
                row
                for row in by_sequence.get(sequence, [])
                if str(row["CHAIN_ID"]) not in used_original
            ]
            for copy_index in range(count):
                predicted_index += 1
                predicted_chain = int_to_letters(predicted_index)
                if copy_index < len(candidates):
                    row = candidates[copy_index]
                    original_chain = str(row["CHAIN_ID"])
                    used_original.add(original_chain)
                    status = (
                        "EXACT_UNIQUE_SEQUENCE"
                        if len(by_sequence.get(sequence, [])) == 1 and count == 1
                        else "IDENTICAL_SEQUENCE_ORDER_ASSUMED"
                    )
                    entity_id = str(row.get("ENTITY_ID", ""))
                else:
                    original_chain = ""
                    entity_id = ""
                    status = "NO_SEQUENCE_MATCH"
                rows.append(
                    {
                        "PDB_ID": target.pdb_id,
                        "PROTENIX_CHAIN_ID": predicted_chain,
                        "ORIGINAL_CHAIN_ID": original_chain,
                        "ORIGINAL_ENTITY_ID": entity_id,
                        "SEQUENCE_LENGTH": len(sequence),
                        "COPY_INDEX": copy_index + 1,
                        "MAPPING_STATUS": status,
                        "NOTE": (
                            "相同序列链不可仅凭序列唯一辨认；当前为确定性顺序映射"
                            if status == "IDENTICAL_SEQUENCE_ORDER_ASSUMED"
                            else ""
                        ),
                    }
                )
    return rows


def build_audit(
    args: argparse.Namespace,
) -> tuple[
    list[dict[str, Any]],
    list[dict[str, Any]],
    list[dict[str, Any]],
    dict[str, Any],
]:
    started = utc_now()
    targets = load_targets(args.manifest)
    raw_index, updated_index = index_json_files(args.simple_json_dir)
    prep_index, pred_index = index_output_dirs(args.complex_json_dir)

    summary_rows: list[dict[str, Any]] = []
    seed_rows: list[dict[str, Any]] = []
    raw_infos: dict[str, JsonInfo] = {}
    for target in targets:
        raw_paths = raw_index.get(target.pdb_id, [])
        updated_paths = updated_index.get(target.pdb_id, [])
        prep_paths = prep_index.get(target.pdb_id, [])
        raw_path = choose_indexed_path(raw_paths)
        updated_path = choose_indexed_path(updated_paths)
        prep_dir = choose_indexed_path(prep_paths)
        raw_info = read_json_info(raw_path)
        raw_infos[target.pdb_id] = raw_info
        prep = inspect_prep(target.pdb_id, updated_path, prep_dir)

        seed_infos: dict[int, SeedInfo] = {}
        for seed in args.seeds:
            paths = pred_index.get((target.pdb_id, seed), [])
            output_dir = choose_indexed_path(paths)
            info = inspect_seed(output_dir, args.samples, args.cif_validation)
            seed_infos[seed] = info
            seed_rows.append(
                {
                    "PDB_ID": target.pdb_id,
                    "SEED": seed,
                    "STATUS": info.status,
                    "REASON": info.reason,
                    "OUTPUT_DIR": info.output_dir,
                    "PREDICTIONS_DIR": info.predictions_dir,
                    "PRIMARY_CIF_COUNT": info.primary_count,
                    "VALID_PRIMARY_CIF_COUNT": info.valid_primary_count,
                    "INVALID_PRIMARY_CIF_COUNT": info.invalid_primary_count,
                    "UNRESOLVED_VARIANT_COUNT": info.unresolved_count,
                    "CONFIDENCE_JSON_COUNT": info.confidence_count,
                    "SAMPLE_INDICES": info.sample_indices,
                    "DUPLICATE_OUTPUT_DIR_COUNT": max(0, len(paths) - 1),
                }
            )

        complete_seeds = sum(
            info.status == "COMPLETE" for info in seed_infos.values()
        )
        if not raw_info.valid:
            overall = "NEED_JSON"
        elif prep.status not in {"COMPLETE", "COMPLETE_REBASABLE"}:
            overall = "NEED_PREP"
        elif complete_seeds != len(args.seeds):
            overall = "NEED_PRED"
        else:
            overall = "COMPLETE"

        row: dict[str, Any] = {
            "PDB_ID": target.pdb_id,
            "LEGACY_1979": target.legacy_1979,
            "CIF_PATH": target.cif_path,
            "RAW_JSON_STATUS": "VALID" if raw_info.valid else "MISSING_OR_INVALID",
            "RAW_JSON_PATH": raw_info.path,
            "RAW_JSON_REASON": raw_info.error or "OK",
            "RAW_JSON_DUPLICATES": max(0, len(raw_paths) - 1),
            "RNA_SEQUENCE_ENTRIES": raw_info.rna_entries,
            "RNA_CHAIN_COUNT": raw_info.rna_chains,
            "PREP_STATUS": prep.status,
            "PREP_REASON": prep.reason,
            "UPDATED_JSON_PATH": prep.updated_json.path,
            "PREP_OUTPUT_DIR": prep.prep_dir,
            "A3M_FILE_COUNT": prep.a3m_found,
            "PREP_OUTPUT_DUPLICATES": max(0, len(prep_paths) - 1),
            "COMPLETE_SEED_COUNT": complete_seeds,
            "VALID_DECOY_COUNT": sum(
                info.valid_primary_count for info in seed_infos.values()
            ),
            "EXPECTED_DECOY_COUNT": args.samples * len(args.seeds),
            "OVERALL_STATUS": overall,
        }
        for seed, info in seed_infos.items():
            row[f"SEED_{seed}_STATUS"] = info.status
            row[f"SEED_{seed}_CIF_COUNT"] = info.valid_primary_count
        summary_rows.append(row)

    chain_rows = make_chain_mapping_rows(
        targets, args.chain_manifest, raw_infos
    )
    finished = utc_now()
    summary = {
        "started_at_utc": started,
        "finished_at_utc": finished,
        "target_count": len(targets),
        "raw_json_valid": sum(
            row["RAW_JSON_STATUS"] == "VALID" for row in summary_rows
        ),
        "prep_complete": sum(
            row["PREP_STATUS"] in {"COMPLETE", "COMPLETE_REBASABLE"}
            for row in summary_rows
        ),
        "prep_rebasable": sum(
            row["PREP_STATUS"] == "COMPLETE_REBASABLE" for row in summary_rows
        ),
        "all_seeds_complete": sum(
            row["COMPLETE_SEED_COUNT"] == len(args.seeds)
            for row in summary_rows
        ),
        "overall_complete": sum(
            row["OVERALL_STATUS"] == "COMPLETE" for row in summary_rows
        ),
        "need_json": sum(
            row["OVERALL_STATUS"] == "NEED_JSON" for row in summary_rows
        ),
        "need_prep": sum(
            row["OVERALL_STATUS"] == "NEED_PREP" for row in summary_rows
        ),
        "need_pred": sum(
            row["OVERALL_STATUS"] == "NEED_PRED" for row in summary_rows
        ),
        "valid_decoy_count": sum(
            int(row["VALID_DECOY_COUNT"]) for row in summary_rows
        ),
        "expected_decoy_count": len(targets) * len(args.seeds) * args.samples,
        "seeds": args.seeds,
        "samples_per_seed": args.samples,
        "cif_validation": args.cif_validation,
        "all_complete": bool(summary_rows)
        and all(row["OVERALL_STATUS"] == "COMPLETE" for row in summary_rows),
    }
    return summary_rows, seed_rows, chain_rows, summary


def style_workbook(path: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    ok_fill = PatternFill("solid", fgColor="C6EFCE")
    attention_fill = PatternFill("solid", fgColor="FFEB9C")
    warn_fill = PatternFill("solid", fgColor="FFC7CE")
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
        for column in sheet.columns:
            values = [str(cell.value or "") for cell in column[:200]]
            width = min(60, max(10, max(map(len, values), default=10) + 2))
            sheet.column_dimensions[get_column_letter(column[0].column)].width = width
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                text = str(cell.value or "").upper()
                if text in {"COMPLETE", "VALID", "EXACT_UNIQUE_SEQUENCE"}:
                    cell.fill = ok_fill
                elif text == "COMPLETE_REBASABLE" or "ORDER_ASSUMED" in text:
                    cell.fill = attention_fill
                elif (
                    text.startswith("NEED_")
                    or text.startswith("MISSING")
                    or text.startswith("INVALID")
                    or text.startswith("INCOMPLETE")
                    or text == "NO_SEQUENCE_MATCH"
                ):
                    cell.fill = warn_fill
    workbook.save(path)


def write_audit_reports(
    report_dir: Path,
    summary_rows: list[dict[str, Any]],
    seed_rows: list[dict[str, Any]],
    chain_rows: list[dict[str, Any]],
    summary: dict[str, Any],
) -> None:
    report_dir.mkdir(parents=True, exist_ok=True)
    structures = pd.DataFrame(summary_rows)
    seeds = pd.DataFrame(seed_rows)
    chains = pd.DataFrame(chain_rows)
    summary_frame = pd.DataFrame(
        [{"METRIC": key, "VALUE": value} for key, value in summary.items()]
    )
    structures.to_csv(
        report_dir / "decoy_manifest.csv", index=False, encoding="utf-8-sig"
    )
    seeds.to_csv(
        report_dir / "decoy_seed_manifest.csv", index=False, encoding="utf-8-sig"
    )
    if not chains.empty:
        chains.to_csv(
            report_dir / "chain_id_mapping.csv", index=False, encoding="utf-8-sig"
        )
    atomic_write_text(
        report_dir / "summary.json",
        json.dumps(summary, ensure_ascii=False, indent=2),
    )
    workbook = report_dir / "decoy_report.xlsx"
    temporary = workbook.with_name(f".{workbook.stem}.tmp.xlsx")
    with pd.ExcelWriter(temporary, engine="openpyxl") as writer:
        structures.to_excel(writer, sheet_name="PDB总览", index=False)
        seeds.to_excel(writer, sheet_name="Seed明细", index=False)
        if not chains.empty:
            chains.to_excel(writer, sheet_name="链ID映射", index=False)
        summary_frame.to_excel(writer, sheet_name="汇总", index=False)
    style_workbook(temporary)
    os.replace(temporary, workbook)


def audit_and_write(args: argparse.Namespace) -> dict[str, Any]:
    rows, seed_rows, chain_rows, summary = build_audit(args)
    write_audit_reports(args.report_dir, rows, seed_rows, chain_rows, summary)
    print("\n审计完成")
    for key in (
        "target_count",
        "raw_json_valid",
        "prep_complete",
        "prep_rebasable",
        "all_seeds_complete",
        "overall_complete",
        "need_json",
        "need_prep",
        "need_pred",
        "valid_decoy_count",
        "expected_decoy_count",
        "all_complete",
    ):
        print(f"  {key}: {summary[key]}")
    print(f"\n易读报告：{args.report_dir / 'decoy_report.xlsx'}")
    print(f"逐PDB清单：{args.report_dir / 'decoy_manifest.csv'}")
    print(f"逐seed清单：{args.report_dir / 'decoy_seed_manifest.csv'}")
    return summary


def run_logged(
    command: list[str],
    log_path: Path,
    event_path: Path,
    event: dict[str, Any],
    env: dict[str, str] | None = None,
) -> int:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    started = utc_now()
    with log_path.open("w", encoding="utf-8") as log:
        log.write("COMMAND: " + json.dumps(command, ensure_ascii=False) + "\n")
        log.write(f"STARTED_AT_UTC: {started}\n\n")
        log.flush()
        result = subprocess.run(
            command,
            env=env,
            stdout=log,
            stderr=subprocess.STDOUT,
            text=True,
            check=False,
        )
    append_jsonl(
        event_path,
        {
            **event,
            "started_at_utc": started,
            "finished_at_utc": utc_now(),
            "return_code": result.returncode,
            "command": command,
            "log_path": str(log_path.resolve()),
        },
    )
    return result.returncode


def find_target_cif(target: Target, cif_dir: Path) -> Path | None:
    recorded = Path(target.cif_path).expanduser() if target.cif_path else None
    if recorded and recorded.is_file():
        return recorded.resolve()
    candidates = [
        path
        for pattern in (f"{target.pdb_id.lower()}.cif", f"{target.pdb_id}.cif")
        for path in cif_dir.rglob(pattern)
        if path.is_file()
    ]
    return choose_indexed_path(candidates)


def make_jsons(args: argparse.Namespace) -> None:
    executable = locate_executable(args.protenix)
    targets = load_targets(args.manifest)
    raw_index, _ = index_json_files(args.simple_json_dir)
    needed = [target for target in targets if not raw_index.get(target.pdb_id)]
    print(f"需要生成原始 JSON：{len(needed)}/{len(targets)}")
    event_path = args.report_dir / "run_events.jsonl"
    lock = threading.Lock()

    def task(target: Target) -> tuple[str, bool, str]:
        cif = find_target_cif(target, args.cif_dir)
        if cif is None:
            return target.pdb_id, False, "找不到 CIF"
        log = args.report_dir / "logs" / "json" / f"{target.pdb_id}.log"
        command = [
            executable,
            "json",
            "-i",
            str(cif),
            "-o",
            str(args.simple_json_dir),
        ]
        code = run_logged(
            command,
            log,
            event_path,
            {"stage": "json", "pdb_id": target.pdb_id},
        )
        output = args.simple_json_dir / f"{cif.stem[:20]}.json"
        valid = read_json_info(output).valid
        with lock:
            print(
                f"[JSON] {target.pdb_id}: "
                f"{'OK' if code == 0 and valid else 'FAILED'}"
            )
        return target.pdb_id, code == 0 and valid, str(log)

    args.simple_json_dir.mkdir(parents=True, exist_ok=True)
    with ThreadPoolExecutor(max_workers=args.workers) as pool:
        results = list(pool.map(task, needed))
    failures = [item for item in results if not item[1]]
    print(f"JSON 完成={len(results) - len(failures)}，失败={len(failures)}")
    audit_and_write(args)
    if failures:
        raise SystemExit(2)


def prep_command(
    executable: str,
    raw_json: Path,
    output_dir: Path,
    args: argparse.Namespace,
) -> list[str]:
    command = [
        executable,
        "prep",
        "-i",
        str(raw_json),
        "-o",
        str(output_dir),
        "-m",
        "protenix",
        "--nhmmer_n_cpu",
        str(args.nhmmer_cpus),
    ]
    options = (
        ("--seqres_database_path", args.seqres_database),
        ("--ntrna_database_path", args.ntrna_database),
        ("--rfam_database_path", args.rfam_database),
        ("--rna_central_database_path", args.rnacentral_database),
    )
    for option, path in options:
        if path:
            command.extend([option, str(path)])
    return command


def run_prep(args: argparse.Namespace) -> None:
    executable = locate_executable(args.protenix)
    targets = load_targets(args.manifest)
    raw_index, updated_index = index_json_files(args.simple_json_dir)
    prep_index, _ = index_output_dirs(args.complex_json_dir)
    needed: list[tuple[Target, Path]] = []
    for target in targets:
        raw = choose_indexed_path(raw_index.get(target.pdb_id, []))
        updated = choose_indexed_path(updated_index.get(target.pdb_id, []))
        prep_dir = choose_indexed_path(prep_index.get(target.pdb_id, []))
        status = inspect_prep(target.pdb_id, updated, prep_dir).status
        if status not in {"COMPLETE", "COMPLETE_REBASABLE"}:
            if raw is None:
                print(f"[PREP] {target.pdb_id}: 跳过，缺少原始 JSON")
            else:
                needed.append((target, raw))
    print(f"需要执行 prep：{len(needed)}/{len(targets)}")
    event_path = args.report_dir / "run_events.jsonl"
    print_lock = threading.Lock()

    def task(item: tuple[Target, Path]) -> tuple[str, bool]:
        target, raw = item
        out = args.complex_json_dir / f"prep_output_{target.pdb_id.lower()}"
        log = args.report_dir / "logs" / "prep" / f"{target.pdb_id}.log"
        command = prep_command(executable, raw, out, args)
        code = run_logged(
            command,
            log,
            event_path,
            {"stage": "prep", "pdb_id": target.pdb_id},
        )
        _, new_updated_index = index_json_files(args.simple_json_dir)
        updated = choose_indexed_path(new_updated_index.get(target.pdb_id, []))
        status = inspect_prep(target.pdb_id, updated, out).status
        ok = code == 0 and status in {"COMPLETE", "COMPLETE_REBASABLE"}
        with print_lock:
            print(f"[PREP] {target.pdb_id}: {'OK' if ok else 'FAILED'} ({status})")
        return target.pdb_id, ok

    args.complex_json_dir.mkdir(parents=True, exist_ok=True)
    with ThreadPoolExecutor(max_workers=args.workers) as pool:
        results = list(pool.map(task, needed))
    failures = [item for item in results if not item[1]]
    print(f"PREP 完成={len(results) - len(failures)}，失败={len(failures)}")
    audit_and_write(args)
    if failures:
        raise SystemExit(2)


def runtime_json(
    pdb_id: str,
    prep: PrepInfo,
    report_dir: Path,
) -> Path:
    if not prep.updated_json.valid:
        raise ValueError(f"{pdb_id}: updated JSON 不可用")
    source = Path(prep.updated_json.path)
    payload = json.loads(source.read_text(encoding="utf-8"))
    rna_index = 0
    for item in payload[0]["sequences"]:
        if "rnaSequence" not in item:
            continue
        path = prep.resolved_msa_paths[rna_index]
        if not path:
            raise ValueError(f"{pdb_id}: 第 {rna_index} 个 RNA MSA 缺失")
        item["rnaSequence"]["unpairedMsaPath"] = path
        rna_index += 1
    output = report_dir / "runtime_json" / f"{pdb_id.lower()}-runtime.json"
    atomic_write_text(output, json.dumps(payload, ensure_ascii=False, indent=4))
    return output.resolve()


def run_pred(args: argparse.Namespace) -> None:
    executable = locate_executable(args.protenix)
    targets = load_targets(args.manifest)
    _, updated_index = index_json_files(args.simple_json_dir)
    prep_index, pred_index = index_output_dirs(args.complex_json_dir)
    tasks: list[tuple[Target, int, PrepInfo]] = []
    for target in targets:
        updated = choose_indexed_path(updated_index.get(target.pdb_id, []))
        prep_dir = choose_indexed_path(prep_index.get(target.pdb_id, []))
        prep = inspect_prep(target.pdb_id, updated, prep_dir)
        if prep.status not in {"COMPLETE", "COMPLETE_REBASABLE"}:
            print(f"[PRED] {target.pdb_id}: 跳过，prep={prep.status}")
            continue
        for seed in args.seeds:
            out = choose_indexed_path(pred_index.get((target.pdb_id, seed), []))
            current = inspect_seed(out, args.samples, args.cif_validation)
            if current.status != "COMPLETE":
                tasks.append((target, seed, prep))
    print(f"需要执行 pred：{len(tasks)}/{len(targets) * len(args.seeds)} 个 seed 任务")
    if not tasks:
        audit_and_write(args)
        return

    gpu_queue: queue.Queue[str] = queue.Queue()
    for gpu in args.gpus:
        gpu_queue.put(gpu)
    print_lock = threading.Lock()
    event_path = args.report_dir / "run_events.jsonl"

    def task(item: tuple[Target, int, PrepInfo]) -> tuple[str, int, bool]:
        target, seed, prep = item
        gpu = gpu_queue.get()
        try:
            input_json = runtime_json(target.pdb_id, prep, args.report_dir)
            out = (
                args.complex_json_dir
                / f"pred_output_{target.pdb_id.lower()}_seed_{seed}"
            )
            log = (
                args.report_dir
                / "logs"
                / "pred"
                / f"{target.pdb_id}_seed_{seed}.log"
            )
            command = [
                executable,
                "pred",
                "-i",
                str(input_json),
                "-o",
                str(out),
                "-n",
                MODEL_NAME,
                "--use_msa",
                "True",
                "--use_rna_msa",
                "True",
                "--use_template",
                "False",
                "--use_default_params",
                "False",
                "--dtype",
                "bf16",
                "--sample",
                str(args.samples),
                "--step",
                "200",
                "--cycle",
                "10",
                "--enable_cache",
                "True",
                "--seeds",
                str(seed),
            ]
            env = os.environ.copy()
            env["CUDA_VISIBLE_DEVICES"] = str(gpu)
            code = run_logged(
                command,
                log,
                event_path,
                {
                    "stage": "pred",
                    "pdb_id": target.pdb_id,
                    "seed": seed,
                    "gpu": gpu,
                    "model": MODEL_NAME,
                    "samples": args.samples,
                },
                env=env,
            )
            verified = inspect_seed(out, args.samples, args.cif_validation)
            ok = code == 0 and verified.status == "COMPLETE"
            with print_lock:
                print(
                    f"[PRED GPU={gpu}] {target.pdb_id} seed={seed}: "
                    f"{'OK' if ok else 'FAILED'} "
                    f"(return={code}, verify={verified.status})"
                )
            return target.pdb_id, seed, ok
        finally:
            gpu_queue.put(gpu)

    args.complex_json_dir.mkdir(parents=True, exist_ok=True)
    with ThreadPoolExecutor(max_workers=len(args.gpus)) as pool:
        futures = [pool.submit(task, item) for item in tasks]
        results = [future.result() for future in as_completed(futures)]
    failures = [item for item in results if not item[2]]
    print(f"PRED 完成={len(results) - len(failures)}，失败={len(failures)}")
    summary = audit_and_write(args)
    if failures or not summary["all_complete"]:
        raise SystemExit(2)


def preflight(args: argparse.Namespace) -> None:
    checks: list[dict[str, Any]] = []

    def add(name: str, value: Any, ok: bool, note: str = "") -> None:
        checks.append({"CHECK": name, "VALUE": value, "OK": ok, "NOTE": note})

    try:
        executable = locate_executable(args.protenix)
        version = subprocess.run(
            [executable, "--version"],
            text=True,
            capture_output=True,
            check=False,
        )
        add(
            "protenix",
            (version.stdout or version.stderr).strip(),
            version.returncode == 0,
            executable,
        )
    except Exception as exc:
        add("protenix", "", False, str(exc))
    for binary in ("nhmmer", "hmmalign", "hmmbuild"):
        located = shutil.which(binary)
        add(binary, located or "", bool(located))
    root = Path(os.environ.get("PROTENIX_ROOT_DIR", str(Path.home()))).expanduser()
    default_files = {
        "seqres_database": root
        / "search_database"
        / "pdb_seqres_2022_09_28.fasta",
        "ntrna_database": root
        / "search_database"
        / "nt_rna_2023_02_23_clust_seq_id_90_cov_80_rep_seq.fasta",
        "rfam_database": root
        / "search_database"
        / "rfam_14_9_clust_seq_id_90_cov_80_rep_seq.fasta",
        "rnacentral_database": root
        / "search_database"
        / "rnacentral_active_seq_id_90_cov_80_linclust.fasta",
    }
    explicit_files = {
        "seqres_database": args.seqres_database,
        "ntrna_database": args.ntrna_database,
        "rfam_database": args.rfam_database,
        "rnacentral_database": args.rnacentral_database,
    }
    for name, explicit in explicit_files.items():
        path = explicit or default_files[name]
        add(
            name,
            str(path),
            path.is_file() and path.stat().st_size > 0,
            "显式路径" if explicit else f"PROTENIX_ROOT_DIR={root}",
        )
    for relative in (
        "common/components.cif",
        "common/components.cif.rdkit_mol.pkl",
        "common/obsolete_release_date.csv",
        "common/clusters-by-entity-40.txt",
        f"checkpoint/{MODEL_NAME}.pt",
    ):
        path = root / relative
        add(relative, str(path), path.is_file() and path.stat().st_size > 0)
    add("manifest", str(args.manifest), args.manifest.is_file())
    add("simple_json_dir", str(args.simple_json_dir), args.simple_json_dir.is_dir())
    add("complex_json_dir", str(args.complex_json_dir), args.complex_json_dir.is_dir())
    if args.command == "pred":
        add("CUDA_VISIBLE_GPUS", ",".join(args.gpus), bool(args.gpus))

    frame = pd.DataFrame(checks)
    args.report_dir.mkdir(parents=True, exist_ok=True)
    frame.to_csv(
        args.report_dir / "preflight.csv", index=False, encoding="utf-8-sig"
    )
    print(frame.to_string(index=False))
    if not all(bool(row["OK"]) for row in checks):
        raise SystemExit(2)


def add_common_arguments(parser: argparse.ArgumentParser) -> None:
    home = Path.home()
    parser.add_argument(
        "--manifest",
        type=Path,
        default=home / "Code/pipeline_reports/PDB_RAW/pdb_cif_manifest.csv",
        help="第一阶段 pdb_cif_manifest.csv",
    )
    parser.add_argument(
        "--chain-manifest",
        type=Path,
        default=home / "Code/pipeline_reports/PDB_RAW/rna_chain_sequences.csv",
        help="第一阶段 rna_chain_sequences.csv",
    )
    parser.add_argument(
        "--cif-dir", type=Path, default=home / "pdb_data", help="原始 CIF 根目录"
    )
    parser.add_argument(
        "--simple-json-dir",
        type=Path,
        default=home / "Json_data/Simple_json",
        help="原始及 final-updated JSON 目录",
    )
    parser.add_argument(
        "--complex-json-dir",
        type=Path,
        default=home / "Json_data/Complex_json",
        help="prep_output/pred_output 根目录",
    )
    parser.add_argument(
        "--report-dir",
        type=Path,
        default=home / "Code/pipeline_reports/DECOYS",
        help="第二阶段报告与日志目录",
    )
    parser.add_argument(
        "--seeds", type=parse_seeds, default=parse_seeds("42,43,44,45")
    )
    parser.add_argument("--samples", type=int, default=50)
    parser.add_argument(
        "--cif-validation",
        choices=("quick", "full"),
        default="quick",
        help="quick 检查文件头；full 使用 gemmi 完整解析",
    )
    parser.add_argument("--protenix", default="protenix", help="Protenix 命令或绝对路径")


def add_database_arguments(parser: argparse.ArgumentParser) -> None:
    parser.add_argument("--seqres-database", type=Path)
    parser.add_argument("--ntrna-database", type=Path)
    parser.add_argument("--rfam-database", type=Path)
    parser.add_argument("--rnacentral-database", type=Path)


def normalize_args(args: argparse.Namespace) -> None:
    for name in (
        "manifest",
        "chain_manifest",
        "cif_dir",
        "simple_json_dir",
        "complex_json_dir",
        "report_dir",
        "seqres_database",
        "ntrna_database",
        "rfam_database",
        "rnacentral_database",
    ):
        if hasattr(args, name):
            value = getattr(args, name)
            if value is not None:
                setattr(args, name, value.expanduser().resolve())
    if args.samples < 1:
        raise ValueError("--samples 必须大于 0")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Protenix 纯 RNA decoy 审计与增量补算管线"
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    audit_parser = subparsers.add_parser("audit", help="只审计，不运行 Protenix")
    add_common_arguments(audit_parser)

    json_parser = subparsers.add_parser("make-json", help="仅补缺失的原始 JSON")
    add_common_arguments(json_parser)
    json_parser.add_argument("--workers", type=int, default=4)

    prep_parser = subparsers.add_parser("prep", help="仅补缺失/损坏的 prep")
    add_common_arguments(prep_parser)
    add_database_arguments(prep_parser)
    prep_parser.add_argument("--workers", type=int, default=4)
    prep_parser.add_argument("--nhmmer-cpus", type=int, default=8)

    pred_parser = subparsers.add_parser("pred", help="仅补不完整的 seed 预测")
    add_common_arguments(pred_parser)
    pred_parser.add_argument(
        "--gpus",
        type=lambda text: [item.strip() for item in text.split(",") if item.strip()],
        default=["0"],
        help="逗号分隔 GPU 编号；每张卡同一时间一个任务",
    )

    preflight_parser = subparsers.add_parser("preflight", help="检查运行环境")
    add_common_arguments(preflight_parser)
    add_database_arguments(preflight_parser)
    return parser


def main() -> None:
    args = build_parser().parse_args()
    normalize_args(args)
    if args.command == "audit":
        audit_and_write(args)
    elif args.command == "make-json":
        make_jsons(args)
    elif args.command == "prep":
        run_prep(args)
    elif args.command == "pred":
        run_pred(args)
    elif args.command == "preflight":
        preflight(args)
    else:
        raise AssertionError(args.command)


if __name__ == "__main__":
    main()
