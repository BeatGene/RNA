#!/usr/bin/env python3
"""Generate one chain_map.tsv per RNA target for coordinate-mapping QC.

For every PDB ID listed in targets.xlsx, the script:

1. expands RNA sequence/count entries from the original Simple JSON;
2. inspects every available primary prediction CIF from the requested seeds;
3. parses the original GT CIF and joins _entity_poly, _struct_asym and
   coordinate-bearing _atom_site chain identifiers;
4. writes <OUTPUT>/<PDB_ID>/chain_map.tsv.

PASS requires a complete one-to-one exact sequence mapping between distinct
input RNA chains and distinct coordinate-bearing GT RNA chains.  Ambiguous or
non-exact cases are retained and marked ``需要人工审核``.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    import gemmi
except ImportError as exc:
    raise SystemExit(
        "Missing gemmi. Install it with: conda install -c conda-forge gemmi openpyxl"
    ) from exc

cif = gemmi.cif


PDB_ID_RE = re.compile(r"^[0-9A-Z]{4}$")
PRED_FOLDER_RE = re.compile(
    r"^pred_output_(?P<pdb_id>[0-9A-Za-z]{4})_seed_(?P<seed>[0-9]+)$",
    re.IGNORECASE,
)
PRIMARY_CIF_RE = re.compile(r"^.+_sample_(?P<sample>[0-9]+)\.cif$", re.IGNORECASE)
NULL_CIF_VALUES = {"", ".", "?"}
DEFAULT_SEEDS = (42, 66, 101, 2024, 8888)
STATUS_PASS = "PASS"
STATUS_REVIEW = "需要人工审核"
OUTPUT_COLUMNS = (
    "pred_chain_id",
    "input_sequence",
    "gt_label_asym_id",
    "gt_auth_asym_id",
    "gt_sequence",
    "identity",
    "exact_match",
    "status",
)


@dataclass(frozen=True)
class ChainRecord:
    label_asym_id: str
    auth_asym_id: str
    sequence: str


@dataclass(frozen=True)
class CandidatePath:
    seed: int
    sample: int
    path: Path


@dataclass
class CandidateInventory:
    candidates: list[CandidatePath] = field(default_factory=list)
    issues: list[str] = field(default_factory=list)


@dataclass
class TargetResult:
    pdb_id: str
    rows: list[dict[str, Any]]
    issues: list[str]
    candidate_count: int
    parsed_candidate_count: int


def normalize_pdb_id(value: object) -> str:
    pdb_id = str(value or "").strip().upper()
    if not PDB_ID_RE.fullmatch(pdb_id):
        raise ValueError(f"Invalid PDB ID: {value!r}")
    return pdb_id


def clean_cif_value(value: Any) -> str:
    text = str(value or "").strip()
    return "" if text in NULL_CIF_VALUES else text


def normalize_sequence(value: Any) -> str:
    return re.sub(r"\s+", "", clean_cif_value(value)).upper()


def sequence_length(sequence: str) -> int:
    if not sequence:
        return 0
    return len(re.findall(r"\([^)]*\)|[A-Za-z?]", sequence))


def sequence_tokens(sequence: str) -> list[str]:
    """Tokenize one-letter or parenthesized modified-residue sequences."""
    return [token.upper() for token in re.findall(r"\([^)]*\)|[A-Za-z?]", sequence)]


def global_sequence_identity(sequence_a: str, sequence_b: str) -> float:
    """Needleman-Wunsch global identity using +1/-1/-1 scoring.

    Identity is the number of exactly equal aligned residue tokens divided by
    the full alignment column count (including columns containing a gap).
    Ties between equally scoring alignments prefer more exact matches, then a
    shorter alignment, then diagonal/up/left in that order.
    """
    tokens_a = sequence_tokens(sequence_a)
    tokens_b = sequence_tokens(sequence_b)
    if tokens_a == tokens_b:
        return 1.0
    if not tokens_a and not tokens_b:
        return 1.0
    if not tokens_a or not tokens_b:
        return 0.0

    # Each DP cell stores (alignment_score, exact_matches, alignment_columns).
    # Only two rows are retained, avoiding an O(len(a)*len(b)) memory matrix.
    previous = [(-index, 0, index) for index in range(len(tokens_b) + 1)]
    for index_a, token_a in enumerate(tokens_a, start=1):
        current: list[tuple[int, int, int]] = [(-index_a, 0, index_a)]
        for index_b, token_b in enumerate(tokens_b, start=1):
            diagonal = previous[index_b - 1]
            is_match = token_a == token_b
            diagonal_candidate = (
                diagonal[0] + (1 if is_match else -1),
                diagonal[1] + int(is_match),
                diagonal[2] + 1,
                2,
            )
            up = previous[index_b]
            up_candidate = (up[0] - 1, up[1], up[2] + 1, 1)
            left = current[index_b - 1]
            left_candidate = (left[0] - 1, left[1], left[2] + 1, 0)
            best = max(
                (diagonal_candidate, up_candidate, left_candidate),
                key=lambda item: (item[0], item[1], -item[2], item[3]),
            )
            current.append(best[:3])
        previous = current
    _, matches, alignment_columns = previous[-1]
    return matches / alignment_columns if alignment_columns else 1.0


def residue_one_letter(residue_name: Any) -> str:
    normalized = clean_cif_value(residue_name).upper()
    common = {
        "A": "A",
        "C": "C",
        "G": "G",
        "U": "U",
        "ADE": "A",
        "CYT": "C",
        "GUA": "G",
        "URA": "U",
    }
    if normalized in common:
        return common[normalized]
    try:
        info = gemmi.find_tabulated_residue(normalized)
        letter = str(info.one_letter_code).strip().upper()
        if len(letter) == 1 and letter.isalpha():
            return letter
    except (AttributeError, RuntimeError, TypeError, ValueError):
        pass
    return "X"


def category_column(
    category: dict[str, list[Any]], name: str, row_count: int
) -> list[Any]:
    values = category.get(name, [])
    if not values:
        return [""] * row_count
    if len(values) != row_count:
        raise ValueError(
            f"mmCIF category column length mismatch: {name}={len(values)}, "
            f"expected={row_count}"
        )
    return values


def read_single_block(path: Path) -> Any:
    if not path.is_file():
        raise FileNotFoundError(path)
    document = cif.read_file(str(path), check_level=2)
    if len(document) != 1:
        raise ValueError(f"Expected one CIF data block, found {len(document)}: {path}")
    return document[0]


def extract_coordinate_rna_chains(path: Path) -> tuple[list[ChainRecord], list[str]]:
    """Return theoretical RNA sequences for chains that occur in _atom_site."""
    block = read_single_block(path)
    issues: list[str] = []

    entity_poly = block.get_mmcif_category("_entity_poly.")
    entity_ids = entity_poly.get("entity_id", [])
    if not entity_ids:
        raise ValueError(f"Missing _entity_poly in {path}")
    entity_count = len(entity_ids)
    entity_types = category_column(entity_poly, "type", entity_count)
    entity_reported = category_column(
        entity_poly, "pdbx_seq_one_letter_code", entity_count
    )
    entity_canonical = category_column(
        entity_poly, "pdbx_seq_one_letter_code_can", entity_count
    )
    entity_poly_seq = block.get_mmcif_category("_entity_poly_seq.")
    poly_seq_entity_ids = entity_poly_seq.get("entity_id", [])
    poly_seq_mon_ids = category_column(
        entity_poly_seq, "mon_id", len(poly_seq_entity_ids)
    )
    fallback_sequence_by_entity: dict[str, list[str]] = defaultdict(list)
    for raw_entity, raw_mon_id in zip(poly_seq_entity_ids, poly_seq_mon_ids):
        fallback_sequence_by_entity[clean_cif_value(raw_entity)].append(
            residue_one_letter(raw_mon_id)
        )
    rna_sequence_by_entity: dict[str, str] = {}
    for raw_id, raw_type, raw_reported, raw_canonical in zip(
        entity_ids,
        entity_types,
        entity_reported,
        entity_canonical,
    ):
        entity_id = clean_cif_value(raw_id)
        polymer_type = clean_cif_value(raw_type).lower()
        if "polyribonucleotide" not in polymer_type:
            continue
        sequence = normalize_sequence(raw_canonical) or normalize_sequence(raw_reported)
        if not sequence:
            sequence = "".join(fallback_sequence_by_entity.get(entity_id, []))
        if not sequence:
            issues.append(f"RNA entity {entity_id!r} has no theoretical sequence")
        rna_sequence_by_entity[entity_id] = sequence
    if not rna_sequence_by_entity:
        raise ValueError(f"No RNA entity found in _entity_poly: {path}")

    atom_site = block.get_mmcif_category("_atom_site.")
    atom_labels = atom_site.get("label_asym_id", [])
    if not atom_labels:
        raise ValueError(f"Missing _atom_site.label_asym_id: {path}")
    atom_auth = category_column(atom_site, "auth_asym_id", len(atom_labels))
    atom_entities = category_column(atom_site, "label_entity_id", len(atom_labels))

    struct_asym = block.get_mmcif_category("_struct_asym.")
    label_ids = struct_asym.get("id", [])
    struct_entities = category_column(struct_asym, "entity_id", len(label_ids))
    label_to_entity: dict[str, str] = {}
    label_order: list[str] = []
    for raw_label, raw_entity in zip(label_ids, struct_entities):
        label = clean_cif_value(raw_label)
        entity = clean_cif_value(raw_entity)
        if entity in rna_sequence_by_entity and label:
            label_to_entity[label] = entity
            label_order.append(label)
    # Protenix output CIFs may omit _struct_asym.  The coordinate records still
    # carry the same label_asym_id -> label_entity_id relationship.
    for raw_label, raw_entity in zip(atom_labels, atom_entities):
        label = clean_cif_value(raw_label)
        entity = clean_cif_value(raw_entity)
        if entity not in rna_sequence_by_entity or not label:
            continue
        if label in label_to_entity and label_to_entity[label] != entity:
            issues.append(
                f"label_asym_id={label} maps to conflicting entity IDs: "
                f"{label_to_entity[label]} vs {entity}"
            )
            continue
        if label not in label_to_entity:
            label_to_entity[label] = entity
            label_order.append(label)
    if not label_to_entity:
        raise ValueError(
            f"No RNA chain relationship found in _struct_asym or _atom_site: {path}"
        )
    auth_by_label: dict[str, list[str]] = defaultdict(list)
    atom_count_by_label: Counter[str] = Counter()
    for raw_label, raw_auth in zip(atom_labels, atom_auth):
        label = clean_cif_value(raw_label)
        if label not in label_to_entity:
            continue
        auth = clean_cif_value(raw_auth)
        atom_count_by_label[label] += 1
        if auth and auth not in auth_by_label[label]:
            auth_by_label[label].append(auth)

    chains: list[ChainRecord] = []
    for label in label_order:
        if atom_count_by_label[label] == 0:
            issues.append(
                f"GT/pred RNA label_asym_id={label} is defined but has no _atom_site coordinates"
            )
            continue
        auth_ids = auth_by_label.get(label, [])
        if len(auth_ids) != 1:
            issues.append(
                f"label_asym_id={label} has {len(auth_ids)} coordinate auth_asym_id values: "
                f"{auth_ids}"
            )
        chains.append(
            ChainRecord(
                label_asym_id=label,
                auth_asym_id="|".join(auth_ids),
                sequence=rna_sequence_by_entity[label_to_entity[label]],
            )
        )
    if not chains:
        raise ValueError(f"No coordinate-bearing RNA chain found: {path}")
    return chains, issues


def read_targets_xlsx(path: Path) -> list[str]:
    if not path.is_file():
        raise FileNotFoundError(f"targets.xlsx not found: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if "Targets" not in workbook.sheetnames:
            raise ValueError(f"Workbook has no Targets sheet: {path}")
        worksheet = workbook["Targets"]
        workbook_rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    rows = iter(workbook_rows)
    try:
        header = next(rows)
    except StopIteration as exc:
        raise ValueError(f"Targets sheet is empty: {path}") from exc
    normalized_header = {
        str(value or "").strip().upper(): index for index, value in enumerate(header)
    }
    if "PDB_ID" not in normalized_header:
        raise ValueError(f"Targets sheet has no PDB_id column: {path}")
    pdb_index = normalized_header["PDB_ID"]
    target_ids: list[str] = []
    seen: set[str] = set()
    for row_number, row in enumerate(rows, start=2):
        if pdb_index >= len(row) or not str(row[pdb_index] or "").strip():
            continue
        pdb_id = normalize_pdb_id(row[pdb_index])
        if pdb_id in seen:
            raise ValueError(f"Duplicate PDB_id {pdb_id} in targets.xlsx row {row_number}")
        seen.add(pdb_id)
        target_ids.append(pdb_id)
    if not target_ids:
        raise ValueError(f"No target PDB IDs found in {path}")
    return target_ids


def index_files_by_stem(directory: Path, suffix: str) -> dict[str, Path]:
    if not directory.is_dir():
        raise NotADirectoryError(directory)
    result: dict[str, Path] = {}
    for path in directory.iterdir():
        if not path.is_file() or path.suffix.lower() != suffix.lower():
            continue
        try:
            pdb_id = normalize_pdb_id(path.stem)
        except ValueError:
            continue
        if pdb_id in result:
            raise ValueError(f"Duplicate {suffix} files for {pdb_id}: {result[pdb_id]}, {path}")
        result[pdb_id] = path
    return result


def read_input_rna_chains(path: Path) -> tuple[list[str], list[str]]:
    issues: list[str] = []
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, list) or not payload:
        raise ValueError(f"JSON top level must be a non-empty list: {path}")
    if len(payload) != 1:
        issues.append(f"Simple JSON task count={len(payload)} (expected 1); first task used")
    task = payload[0]
    if not isinstance(task, dict) or not isinstance(task.get("sequences"), list):
        raise ValueError(f"Simple JSON first task has no sequences list: {path}")
    chains: list[str] = []
    for item_index, item in enumerate(task["sequences"]):
        if not isinstance(item, dict) or "rnaSequence" not in item:
            continue
        rna = item["rnaSequence"]
        if not isinstance(rna, dict):
            raise ValueError(f"rnaSequence item {item_index} is not an object: {path}")
        sequence = normalize_sequence(rna.get("sequence", ""))
        if not sequence:
            raise ValueError(f"rnaSequence item {item_index} has empty sequence: {path}")
        count = int(rna.get("count", 1))
        if count < 1:
            raise ValueError(f"rnaSequence item {item_index} has invalid count={count}: {path}")
        chains.extend([sequence] * count)
    if not chains:
        raise ValueError(f"No RNA sequence found in {path}")
    return chains, issues


def index_prediction_seed_dirs(pred_dir: Path) -> dict[str, dict[int, Path]]:
    if not pred_dir.is_dir():
        raise NotADirectoryError(pred_dir)
    result: dict[str, dict[int, Path]] = defaultdict(dict)
    for path in pred_dir.iterdir():
        if not path.is_dir():
            continue
        match = PRED_FOLDER_RE.fullmatch(path.name)
        if not match:
            continue
        pdb_id = normalize_pdb_id(match.group("pdb_id"))
        seed = int(match.group("seed"))
        if seed in result[pdb_id]:
            raise ValueError(
                f"Duplicate normalized prediction folder for {pdb_id} seed {seed}: "
                f"{result[pdb_id][seed]}, {path}"
            )
        result[pdb_id][seed] = path
    return result


def inventory_candidates(
    pdb_id: str,
    seed_dirs: dict[int, Path],
    *,
    expected_seeds: tuple[int, ...],
    expected_samples: int,
) -> CandidateInventory:
    inventory = CandidateInventory()
    missing_seeds = [seed for seed in expected_seeds if seed not in seed_dirs]
    if missing_seeds:
        inventory.issues.append(f"missing prediction seeds={missing_seeds}")
    unexpected_seeds = sorted(set(seed_dirs) - set(expected_seeds))
    if unexpected_seeds:
        inventory.issues.append(f"unexpected prediction seeds ignored={unexpected_seeds}")

    for seed in expected_seeds:
        seed_dir = seed_dirs.get(seed)
        if seed_dir is None:
            continue
        prediction_dirs = sorted(
            path for path in seed_dir.rglob("predictions") if path.is_dir()
        )
        if len(prediction_dirs) != 1:
            inventory.issues.append(
                f"seed_{seed}: predictions directory count={len(prediction_dirs)} (expected 1)"
            )
            continue
        sample_paths: dict[int, list[Path]] = defaultdict(list)
        for path in prediction_dirs[0].iterdir():
            if not path.is_file():
                continue
            match = PRIMARY_CIF_RE.fullmatch(path.name)
            if match:
                sample_paths[int(match.group("sample"))].append(path)
        expected_indices = set(range(expected_samples))
        actual_indices = set(sample_paths)
        missing = sorted(expected_indices - actual_indices)
        extra = sorted(actual_indices - expected_indices)
        if missing:
            inventory.issues.append(f"seed_{seed}: missing samples={missing}")
        if extra:
            inventory.issues.append(f"seed_{seed}: extra samples ignored={extra}")
        for sample in range(expected_samples):
            paths = sample_paths.get(sample, [])
            if len(paths) != 1:
                if len(paths) > 1:
                    inventory.issues.append(
                        f"seed_{seed} sample_{sample}: duplicate primary CIF count={len(paths)}"
                    )
                continue
            inventory.candidates.append(
                CandidatePath(seed=seed, sample=sample, path=paths[0])
            )
    inventory.candidates.sort(key=lambda item: (item.seed, item.sample, str(item.path)))
    return inventory


def excel_chain_id(index: int) -> str:
    if index < 1:
        raise ValueError("chain index must be >= 1")
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(ord("A") + remainder) + letters
    return letters


def choose_prediction_chains(
    input_sequences: list[str],
    inventory: CandidateInventory,
) -> tuple[list[ChainRecord], list[str], int, bool]:
    issues = list(inventory.issues)
    signatures: list[tuple[tuple[str, str], ...]] = []
    parsed_chains: list[list[ChainRecord]] = []
    for candidate in inventory.candidates:
        try:
            chains, cif_issues = extract_coordinate_rna_chains(candidate.path)
            issues.extend(
                f"seed_{candidate.seed} sample_{candidate.sample}: {issue}"
                for issue in cif_issues
            )
            signatures.append(
                tuple((chain.label_asym_id, chain.sequence) for chain in chains)
            )
            parsed_chains.append(chains)
        except Exception as exc:
            issues.append(
                f"seed_{candidate.seed} sample_{candidate.sample}: "
                f"prediction CIF parse failed: {type(exc).__name__}: {exc}"
            )

    if parsed_chains:
        selected = parsed_chains[0]
        signatures_consistent = len(set(signatures)) == 1
        if not signatures_consistent:
            issues.append(
                f"prediction RNA chain signatures are inconsistent across "
                f"{len(signatures)} parsed candidates"
            )
    else:
        signatures_consistent = False
        selected = [
            ChainRecord(excel_chain_id(index), "", sequence)
            for index, sequence in enumerate(input_sequences, start=1)
        ]
        issues.append("no prediction CIF could be parsed; predicted chain IDs are fallbacks")

    return selected, issues, len(parsed_chains), signatures_consistent


def make_chain_rows(
    input_sequences: list[str],
    predicted_chains: list[ChainRecord],
    gt_chains: list[ChainRecord],
    *,
    prediction_validation_ok: bool,
    inherited_issues: list[str],
) -> tuple[list[dict[str, Any]], list[str]]:
    issues = list(inherited_issues)
    row_count = len(input_sequences) if input_sequences else len(predicted_chains)
    if len(input_sequences) != len(predicted_chains):
        issues.append(
            f"input RNA chain count={len(input_sequences)} but prediction chain count="
            f"{len(predicted_chains)} (logged for QC; not used as a PASS condition)"
        )

    row_inputs: list[str] = []
    row_pred_ids: list[str] = []
    for index in range(row_count):
        row_inputs.append(input_sequences[index] if index < len(input_sequences) else "")
        if index < len(predicted_chains):
            row_pred_ids.append(predicted_chains[index].label_asym_id)
        else:
            row_pred_ids.append(excel_chain_id(index + 1))

    assigned_gt: dict[int, int] = {}
    used_gt: set[int] = set()
    input_groups: dict[str, list[int]] = defaultdict(list)
    gt_groups: dict[str, list[int]] = defaultdict(list)
    for index, sequence in enumerate(row_inputs):
        if sequence:
            input_groups[sequence].append(index)
    for index, chain in enumerate(gt_chains):
        gt_groups[chain.sequence].append(index)

    for sequence, input_indices in input_groups.items():
        gt_indices = gt_groups.get(sequence, [])
        available = [index for index in gt_indices if index not in used_gt]
        for input_index, gt_index in zip(input_indices, available):
            assigned_gt[input_index] = gt_index
            used_gt.add(gt_index)
        if len(input_indices) != len(gt_indices):
            issues.append(
                f"sequence multiplicity mismatch for length {sequence_length(sequence)}: "
                f"input={len(input_indices)}, GT={len(gt_indices)}"
            )
        elif len(input_indices) > 1:
            issues.append(
                f"identical sequence occurs {len(input_indices)} times; deterministic order "
                f"mapping requires manual review"
            )

    # After all exact matches are consumed, a single remaining input chain and
    # a single remaining GT chain still form a unique chain-level pairing.  It
    # is retained for review rather than discarded, because the sequence
    # mismatch is precisely what the exact_match/status columns should expose.
    unmatched_inputs = [
        index
        for index, sequence in enumerate(row_inputs)
        if sequence and index not in assigned_gt
    ]
    unmatched_gt = [index for index in range(len(gt_chains)) if index not in used_gt]
    if len(unmatched_inputs) == 1 and len(unmatched_gt) == 1:
        input_index = unmatched_inputs[0]
        gt_index = unmatched_gt[0]
        assigned_gt[input_index] = gt_index
        used_gt.add(gt_index)
        issues.append(
            f"one-to-one residual chain mapping retained despite non-exact sequence: "
            f"pred={row_pred_ids[input_index]}, GT={gt_chains[gt_index].label_asym_id}"
        )

    rows: list[dict[str, Any]] = []
    input_sequences_unique = len(set(input_sequences)) == len(input_sequences)
    gt_sequences = [chain.sequence for chain in gt_chains]
    gt_sequences_unique = len(set(gt_sequences)) == len(gt_sequences)
    all_inputs_uniquely_assigned = (
        bool(input_sequences)
        and len(input_sequences) == len(gt_chains)
        and len(row_inputs) == len(input_sequences)
        and set(assigned_gt) == set(range(len(input_sequences)))
        and set(assigned_gt.values()) == set(range(len(gt_chains)))
        and all(
            row_inputs[index] == gt_chains[assigned_gt[index]].sequence
            for index in range(len(input_sequences))
        )
    )
    all_gt_coordinate_chain_ids_unique = all(
        bool(chain.auth_asym_id)
        and len(set(chain.auth_asym_id.split("|"))) == 1
        for chain in gt_chains
    )
    target_pass = (
        input_sequences_unique
        and gt_sequences_unique
        and all_inputs_uniquely_assigned
        and all_gt_coordinate_chain_ids_unique
        and prediction_validation_ok
    )

    for index, (pred_chain_id, input_sequence) in enumerate(
        zip(row_pred_ids, row_inputs)
    ):
        gt_index = assigned_gt.get(index)
        gt_chain = gt_chains[gt_index] if gt_index is not None else None
        gt_sequence = gt_chain.sequence if gt_chain else ""
        identity = (
            global_sequence_identity(input_sequence, gt_sequence)
            if input_sequence and gt_sequence
            else None
        )
        exact_match = bool(input_sequence and gt_sequence and input_sequence == gt_sequence)
        rows.append(
            {
                "pred_chain_id": pred_chain_id,
                "input_sequence": input_sequence,
                "gt_label_asym_id": gt_chain.label_asym_id if gt_chain else "",
                "gt_auth_asym_id": gt_chain.auth_asym_id if gt_chain else "",
                "gt_sequence": gt_sequence,
                "identity": f"{identity:.6f}" if identity is not None else "",
                "exact_match": str(exact_match),
                "status": STATUS_PASS if target_pass else STATUS_REVIEW,
            }
        )
    return rows, issues


def process_target(
    pdb_id: str,
    *,
    json_path: Path | None,
    gt_path: Path | None,
    seed_dirs: dict[int, Path],
    expected_seeds: tuple[int, ...],
    expected_samples: int,
) -> TargetResult:
    issues: list[str] = []
    input_sequences: list[str] = []
    if json_path is None:
        issues.append("Simple JSON file is missing")
    else:
        try:
            input_sequences, json_issues = read_input_rna_chains(json_path)
            issues.extend(json_issues)
        except Exception as exc:
            issues.append(f"Simple JSON parse failed: {type(exc).__name__}: {exc}")

    inventory = inventory_candidates(
        pdb_id,
        seed_dirs,
        expected_seeds=expected_seeds,
        expected_samples=expected_samples,
    )
    (
        predicted_chains,
        prediction_issues,
        parsed_candidate_count,
        signatures_consistent,
    ) = choose_prediction_chains(input_sequences, inventory)
    issues.extend(prediction_issues)

    gt_chains: list[ChainRecord] = []
    if gt_path is None:
        issues.append("GT CIF file is missing")
    else:
        try:
            gt_chains, gt_issues = extract_coordinate_rna_chains(gt_path)
            issues.extend(f"GT: {issue}" for issue in gt_issues)
        except Exception as exc:
            issues.append(f"GT CIF parse failed: {type(exc).__name__}: {exc}")

    prediction_validation_ok = (
        len(inventory.candidates) > 0
        and parsed_candidate_count == len(inventory.candidates)
        and signatures_consistent
    )
    rows, issues = make_chain_rows(
        input_sequences,
        predicted_chains,
        gt_chains,
        prediction_validation_ok=prediction_validation_ok,
        inherited_issues=issues,
    )
    if not rows:
        issues.append("No chain row could be constructed; header-only TSV written")
    return TargetResult(
        pdb_id=pdb_id,
        rows=rows,
        issues=list(dict.fromkeys(issues)),
        candidate_count=len(inventory.candidates),
        parsed_candidate_count=parsed_candidate_count,
    )


def write_chain_map(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.tmp")
    try:
        with temporary.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(
                handle,
                fieldnames=OUTPUT_COLUMNS,
                delimiter="\t",
                lineterminator="\n",
                extrasaction="raise",
            )
            writer.writeheader()
            writer.writerows(rows)
        os.replace(temporary, path)
    finally:
        if temporary.exists():
            temporary.unlink()


def parse_seed_list(value: str) -> tuple[int, ...]:
    try:
        seeds = tuple(int(piece.strip()) for piece in value.split(",") if piece.strip())
    except ValueError as exc:
        raise argparse.ArgumentTypeError("Seeds must be comma-separated integers") from exc
    if not seeds or len(seeds) != len(set(seeds)):
        raise argparse.ArgumentTypeError("Seeds must be a non-empty unique list")
    return seeds


def build_parser() -> argparse.ArgumentParser:
    home = Path.home()
    parser = argparse.ArgumentParser(
        description="Generate per-PDB chain_map.tsv files for RNA coordinate mapping QC."
    )
    parser.add_argument(
        "--targets-xlsx",
        type=Path,
        default=home / "Json_data" / "mapping" / "xlsx" / "targets.xlsx",
    )
    parser.add_argument(
        "--simple-json-dir",
        type=Path,
        default=home / "Json_data" / "Simple_json",
    )
    parser.add_argument("--gt-cif-dir", type=Path, default=home / "pdb_data")
    parser.add_argument(
        "--pred-dir",
        type=Path,
        default=home / "Json_data" / "Foldbench_predictions",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=home / "Json_data" / "mapping" / "tsv",
    )
    parser.add_argument(
        "--seeds",
        type=parse_seed_list,
        default=DEFAULT_SEEDS,
        help="Comma-separated seeds (default: 42,66,101,2024,8888).",
    )
    parser.add_argument("--samples-per-seed", type=int, default=5)
    parser.add_argument("--workers", type=int, default=8)
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Replace existing chain_map.tsv files.",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    targets_xlsx = args.targets_xlsx.expanduser()
    simple_json_dir = args.simple_json_dir.expanduser()
    gt_cif_dir = args.gt_cif_dir.expanduser()
    pred_dir = args.pred_dir.expanduser()
    output_dir = args.output_dir.expanduser()
    if args.samples_per_seed < 1:
        raise ValueError("--samples-per-seed must be positive")
    if args.workers < 1:
        raise ValueError("--workers must be positive")

    target_ids = read_targets_xlsx(targets_xlsx)
    json_index = index_files_by_stem(simple_json_dir, ".json")
    gt_index = index_files_by_stem(gt_cif_dir, ".cif")
    prediction_index = index_prediction_seed_dirs(pred_dir)
    existing = [
        output_dir / pdb_id / "chain_map.tsv"
        for pdb_id in target_ids
        if (output_dir / pdb_id / "chain_map.tsv").exists()
    ]
    if existing and not args.overwrite:
        raise FileExistsError(
            f"{len(existing)} chain_map.tsv files already exist (first: {existing[0]}). "
            f"Use --overwrite only when replacement is intended."
        )

    print(f"Targets loaded from workbook: {len(target_ids)}")
    print(f"Workers: {args.workers}")
    results: dict[str, TargetResult] = {}
    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        future_to_pdb = {
            executor.submit(
                process_target,
                pdb_id,
                json_path=json_index.get(pdb_id),
                gt_path=gt_index.get(pdb_id),
                seed_dirs=prediction_index.get(pdb_id, {}),
                expected_seeds=args.seeds,
                expected_samples=args.samples_per_seed,
            ): pdb_id
            for pdb_id in target_ids
        }
        completed = 0
        for future in as_completed(future_to_pdb):
            pdb_id = future_to_pdb[future]
            try:
                results[pdb_id] = future.result()
            except Exception as exc:
                results[pdb_id] = TargetResult(
                    pdb_id=pdb_id,
                    rows=[],
                    issues=[f"Unhandled target error: {type(exc).__name__}: {exc}"],
                    candidate_count=0,
                    parsed_candidate_count=0,
                )
            completed += 1
            if completed % 50 == 0 or completed == len(target_ids):
                print(f"Processed targets: {completed}/{len(target_ids)}", flush=True)

    for pdb_id in target_ids:
        write_chain_map(
            output_dir / pdb_id / "chain_map.tsv",
            results[pdb_id].rows,
        )

    file_count = len(results)
    row_count = sum(len(result.rows) for result in results.values())
    pass_rows = sum(
        row["status"] == STATUS_PASS
        for result in results.values()
        for row in result.rows
    )
    review_rows = row_count - pass_rows
    pass_targets = sum(
        bool(result.rows)
        and all(row["status"] == STATUS_PASS for row in result.rows)
        for result in results.values()
    )
    review_targets = file_count - pass_targets
    issue_targets = [pdb_id for pdb_id in target_ids if results[pdb_id].issues]
    print("\nPer-target issues requiring attention:")
    if issue_targets:
        for pdb_id in issue_targets:
            result = results[pdb_id]
            print(
                f"  {pdb_id} (candidate CIFs {result.parsed_candidate_count}/"
                f"{result.candidate_count} parsed):"
            )
            for issue in result.issues:
                print(f"    - {issue}")
    else:
        print("  None")
    print("\nFinal summary:")
    print(f"  PDB directories written: {file_count}")
    print(f"  chain_map.tsv files written: {file_count}")
    print(f"  PASS targets: {pass_targets}")
    print(f"  需要人工审核 targets: {review_targets}")
    print(f"  chain rows written: {row_count}")
    print(f"  PASS rows: {pass_rows}")
    print(f"  需要人工审核 rows: {review_rows}")
    print(f"  targets with logged issues: {len(issue_targets)}")
    print(f"  output root: {output_dir}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"ERROR: {type(exc).__name__}: {exc}", file=sys.stderr)
        raise SystemExit(1)
