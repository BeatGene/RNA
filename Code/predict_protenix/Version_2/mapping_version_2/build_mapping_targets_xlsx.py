#!/usr/bin/env python3
"""Build the target overview workbook for RNA coordinate-mapping QC.

The target IDs are discovered from Protenix prediction directory names.
Release dates and RNA chain counts are joined from the existing pipeline audit
tables.  The output workbook intentionally leaves ``mapping_status`` and
``review_status`` empty for later QC stages.
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


PRED_FOLDER_RE = re.compile(
    r"^pred_output_(?P<pdb_id>[0-9A-Za-z]{4})_seed_(?P<seed>[0-9]+)$",
    re.IGNORECASE,
)
PRIMARY_CIF_RE = re.compile(r"^.+_sample_(?P<sample>[0-9]+)\.cif$", re.IGNORECASE)
PDB_ID_RE = re.compile(r"^[0-9A-Z]{4}$")
DEFAULT_SEEDS = (42, 66, 101, 2024, 8888)
DEFAULT_CUTOFF = date(2021, 9, 30)
OUTPUT_COLUMNS = (
    "target_id",
    "PDB_id",
    "release_date",
    "time_group",
    "chain_count",
    "mapping_status",
    "review_status",
)


@dataclass(frozen=True)
class PredictionScan:
    all_target_ids: tuple[str, ...]
    complete_target_ids: tuple[str, ...]
    incomplete_reasons: dict[str, tuple[str, ...]]


def normalize_pdb_id(value: object) -> str:
    pdb_id = str(value or "").strip().upper()
    if not PDB_ID_RE.fullmatch(pdb_id):
        raise ValueError(f"Invalid PDB ID: {value!r}")
    return pdb_id


def parse_iso_date(value: object, *, context: str) -> date:
    text = str(value or "").strip()
    try:
        return date.fromisoformat(text[:10])
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Invalid RELEASE_DATE in {context}: {value!r}") from exc


def parse_positive_int(value: object, *, context: str) -> int:
    text = str(value or "").strip()
    try:
        parsed_float = float(text)
        parsed_int = int(parsed_float)
    except (TypeError, ValueError, OverflowError) as exc:
        raise ValueError(f"Invalid positive integer in {context}: {value!r}") from exc
    if parsed_float != parsed_int or parsed_int < 1:
        raise ValueError(f"Invalid positive integer in {context}: {value!r}")
    return parsed_int


def detect_delimiter(path: Path) -> str:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(65536)
    if not sample.strip():
        raise ValueError(f"Metadata file is empty: {path}")
    try:
        return csv.Sniffer().sniff(sample, delimiters=",\t").delimiter
    except csv.Error:
        return "\t" if "\t" in sample.partition("\n")[0] else ","


def read_metadata_index(
    path: Path,
    *,
    value_column: str,
    value_parser,
) -> dict[str, object]:
    if not path.is_file():
        raise FileNotFoundError(f"Metadata file not found: {path}")
    delimiter = detect_delimiter(path)
    result: dict[str, object] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=delimiter)
        if reader.fieldnames is None:
            raise ValueError(f"Metadata file has no header: {path}")
        original_by_normalized = {
            name.strip().upper(): name for name in reader.fieldnames if name is not None
        }
        required = {"PDB_ID", value_column.upper()}
        missing = sorted(required - set(original_by_normalized))
        if missing:
            raise ValueError(f"Missing columns in {path}: {', '.join(missing)}")
        pdb_column = original_by_normalized["PDB_ID"]
        actual_value_column = original_by_normalized[value_column.upper()]
        for row_number, row in enumerate(reader, start=2):
            raw_pdb_id = row.get(pdb_column, "")
            if not str(raw_pdb_id or "").strip():
                continue
            pdb_id = normalize_pdb_id(raw_pdb_id)
            if pdb_id in result:
                raise ValueError(f"Duplicate PDB_ID {pdb_id} in {path} at row {row_number}")
            context = f"{path}:{row_number} ({pdb_id})"
            result[pdb_id] = value_parser(row.get(actual_value_column, ""), context=context)
    return result


def find_single_predictions_dir(seed_dir: Path) -> tuple[Path | None, str | None]:
    prediction_dirs = sorted(
        path for path in seed_dir.rglob("predictions") if path.is_dir()
    )
    if len(prediction_dirs) != 1:
        return None, f"predictions_dir_count={len(prediction_dirs)} (expected 1)"
    return prediction_dirs[0], None


def inspect_sample_indices(
    seed_dir: Path,
    *,
    expected_samples: int,
) -> tuple[bool, str]:
    predictions_dir, layout_error = find_single_predictions_dir(seed_dir)
    if layout_error is not None or predictions_dir is None:
        return False, layout_error or "predictions directory not found"

    sample_to_count: dict[int, int] = defaultdict(int)
    for path in predictions_dir.iterdir():
        if not path.is_file():
            continue
        match = PRIMARY_CIF_RE.fullmatch(path.name)
        if match:
            sample_to_count[int(match.group("sample"))] += 1

    expected = set(range(expected_samples))
    actual = set(sample_to_count)
    duplicates = sorted(index for index, count in sample_to_count.items() if count != 1)
    missing = sorted(expected - actual)
    extra = sorted(actual - expected)
    if missing or extra or duplicates:
        pieces = [f"primary_cif_count={sum(sample_to_count.values())}/{expected_samples}"]
        if missing:
            pieces.append(f"missing_samples={missing}")
        if extra:
            pieces.append(f"extra_samples={extra}")
        if duplicates:
            pieces.append(f"duplicate_samples={duplicates}")
        return False, "; ".join(pieces)
    return True, "OK"


def scan_predictions(
    pred_dir: Path,
    *,
    expected_seeds: Iterable[int],
    expected_samples: int,
) -> PredictionScan:
    if not pred_dir.is_dir():
        raise NotADirectoryError(f"Prediction directory not found: {pred_dir}")

    target_seed_dirs: dict[str, dict[int, Path]] = defaultdict(dict)
    for path in sorted(pred_dir.iterdir(), key=lambda item: item.name.lower()):
        if not path.is_dir():
            continue
        match = PRED_FOLDER_RE.fullmatch(path.name)
        if not match:
            continue
        pdb_id = normalize_pdb_id(match.group("pdb_id"))
        seed = int(match.group("seed"))
        if seed in target_seed_dirs[pdb_id]:
            previous = target_seed_dirs[pdb_id][seed]
            raise ValueError(
                f"Duplicate normalized prediction folder for {pdb_id} seed {seed}: "
                f"{previous} and {path}"
            )
        target_seed_dirs[pdb_id][seed] = path

    if not target_seed_dirs:
        raise ValueError(f"No pred_output_<PDB>_seed_<SEED> directories found in {pred_dir}")

    required_seeds = tuple(expected_seeds)
    required_seed_set = set(required_seeds)
    complete: list[str] = []
    incomplete: dict[str, tuple[str, ...]] = {}
    for pdb_id in sorted(target_seed_dirs):
        seed_dirs = target_seed_dirs[pdb_id]
        reasons: list[str] = []
        missing_seeds = sorted(required_seed_set - set(seed_dirs))
        if missing_seeds:
            reasons.append(f"missing_seeds={missing_seeds}")
        for seed in required_seeds:
            seed_dir = seed_dirs.get(seed)
            if seed_dir is None:
                continue
            valid, detail = inspect_sample_indices(
                seed_dir,
                expected_samples=expected_samples,
            )
            if not valid:
                reasons.append(f"seed_{seed}: {detail}")
        if reasons:
            incomplete[pdb_id] = tuple(reasons)
        else:
            complete.append(pdb_id)

    return PredictionScan(
        all_target_ids=tuple(sorted(target_seed_dirs)),
        complete_target_ids=tuple(complete),
        incomplete_reasons=incomplete,
    )


def build_rows(
    target_ids: Iterable[str],
    *,
    release_dates: dict[str, date],
    chain_counts: dict[str, int],
    cutoff: date,
) -> list[tuple[object, ...]]:
    normalized_targets = sorted({normalize_pdb_id(value) for value in target_ids})
    rows: list[tuple[object, ...]] = []
    for pdb_id in normalized_targets:
        release_date = release_dates.get(pdb_id)
        if release_date is None:
            time_group = None
        else:
            time_group = "post_cutoff" if release_date > cutoff else "pre_or_on_cutoff"
        rows.append(
            (
                pdb_id,
                pdb_id,
                release_date,
                time_group,
                chain_counts.get(pdb_id),
                None,
                None,
            )
        )
    return rows


def write_workbook(rows: list[tuple[object, ...]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Targets"
    worksheet.append(OUTPUT_COLUMNS)
    for row in rows:
        worksheet.append(row)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for cell in worksheet["C"][1:]:
        cell.number_format = "yyyy-mm-dd"
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.column_dimensions["A"].width = 14
    worksheet.column_dimensions["B"].width = 12
    worksheet.column_dimensions["C"].width = 15
    worksheet.column_dimensions["D"].width = 20
    worksheet.column_dimensions["E"].width = 14
    worksheet.column_dimensions["F"].width = 20
    worksheet.column_dimensions["G"].width = 20

    if rows:
        table = Table(displayName="MappingTargets", ref=worksheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    temporary_path = output_path.with_name(f".{output_path.name}.tmp.xlsx")
    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, output_path)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()


def parse_seed_list(value: str) -> tuple[int, ...]:
    try:
        seeds = tuple(int(piece.strip()) for piece in value.split(",") if piece.strip())
    except ValueError as exc:
        raise argparse.ArgumentTypeError("Seeds must be comma-separated integers") from exc
    if not seeds or len(seeds) != len(set(seeds)):
        raise argparse.ArgumentTypeError("Seeds must be a non-empty unique list")
    return seeds


def parse_cutoff(value: str) -> date:
    try:
        return date.fromisoformat(value)
    except ValueError as exc:
        raise argparse.ArgumentTypeError("Cutoff must use YYYY-MM-DD") from exc


def default_date_audit(home: Path) -> Path:
    report_dir = (
        home
        / "Code"
        / "pipeline_reports"
        / "DATA_SPLIT_2241_CHAINMASK_20260807T114307Z_EXECUTE"
    )
    csv_path = report_dir / "date_audit.csv"
    tsv_path = report_dir / "date_audit.tsv"
    if csv_path.exists() or not tsv_path.exists():
        return csv_path
    return tsv_path


def build_parser() -> argparse.ArgumentParser:
    home = Path.home()
    parser = argparse.ArgumentParser(
        description="Generate targets.xlsx for RNA mapping QC from prediction folder names."
    )
    parser.add_argument(
        "--pred-dir",
        type=Path,
        default=home / "Json_data" / "Foldbench_predictions",
    )
    parser.add_argument("--date-audit", type=Path, default=default_date_audit(home))
    parser.add_argument(
        "--manifest",
        type=Path,
        default=home / "Code" / "pipeline_reports" / "PDB_RAW" / "pdb_cif_manifest.csv",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=home / "Json_data" / "mapping" / "xlsx" / "targets.xlsx",
    )
    parser.add_argument("--cutoff", type=parse_cutoff, default=DEFAULT_CUTOFF)
    parser.add_argument(
        "--seeds",
        type=parse_seed_list,
        default=DEFAULT_SEEDS,
        help="Comma-separated seed list (default: 42,66,101,2024,8888).",
    )
    parser.add_argument("--samples-per-seed", type=int, default=5)
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Replace an existing workbook. Without this flag, existing output is preserved.",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    pred_dir = args.pred_dir.expanduser()
    date_audit = args.date_audit.expanduser()
    manifest = args.manifest.expanduser()
    output = args.output.expanduser()
    if args.samples_per_seed < 1:
        raise ValueError("--samples-per-seed must be positive")
    if output.exists() and not args.overwrite:
        raise FileExistsError(
            f"Output already exists: {output}. Use --overwrite only if replacement is intended."
        )

    scan = scan_predictions(
        pred_dir,
        expected_seeds=args.seeds,
        expected_samples=args.samples_per_seed,
    )
    print(f"Matching target IDs found in folder names: {len(scan.all_target_ids)}")
    print(f"Targets with all required seeds and samples: {len(scan.complete_target_ids)}")
    print(f"Targets with completeness warnings (still included): {len(scan.incomplete_reasons)}")

    release_dates = read_metadata_index(
        date_audit,
        value_column="RELEASE_DATE",
        value_parser=parse_iso_date,
    )
    chain_counts = read_metadata_index(
        manifest,
        value_column="RNA_CHAIN_COUNT",
        value_parser=parse_positive_int,
    )
    rows = build_rows(
        scan.all_target_ids,
        release_dates=release_dates,
        chain_counts=chain_counts,
        cutoff=args.cutoff,
    )
    write_workbook(rows, output)

    pre_count = sum(row[3] == "pre_or_on_cutoff" for row in rows)
    post_count = sum(row[3] == "post_cutoff" for row in rows)
    missing_date_ids = sorted(
        pdb_id for pdb_id in scan.all_target_ids if pdb_id not in release_dates
    )
    missing_chain_ids = sorted(
        pdb_id for pdb_id in scan.all_target_ids if pdb_id not in chain_counts
    )
    print(f"Time groups: pre_or_on_cutoff={pre_count}, post_cutoff={post_count}")
    print(f"Wrote workbook to: {output}")
    print("\nPrediction-completeness warnings (targets were not excluded):")
    if scan.incomplete_reasons:
        for pdb_id in sorted(scan.incomplete_reasons):
            print(f"  {pdb_id}: {' | '.join(scan.incomplete_reasons[pdb_id])}")
    else:
        print("  None")
    print("\nMetadata warnings (cells were left blank; targets were not excluded):")
    if missing_date_ids:
        print(f"  Missing RELEASE_DATE ({len(missing_date_ids)}): {', '.join(missing_date_ids)}")
    if missing_chain_ids:
        print(
            f"  Missing RNA_CHAIN_COUNT ({len(missing_chain_ids)}): "
            f"{', '.join(missing_chain_ids)}"
        )
    if not missing_date_ids and not missing_chain_ids:
        print("  None")
    print(f"\nFinal unique PDB_id count written: {len(rows)}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"ERROR: {type(exc).__name__}: {exc}", file=sys.stderr)
        raise SystemExit(1)
