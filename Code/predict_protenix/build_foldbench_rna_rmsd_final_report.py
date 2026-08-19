#!/usr/bin/env python3
"""Build the frozen-exclusion final RNA RMSD report.

Inputs are the candidate/rank-1 CSV files produced by
``summarize_foldbench_rna_rmsd.py`` plus a user-approved frozen PDB exclusion
list.  Excluded targets remain excluded even if a later rescue result appears.
"""

from __future__ import annotations

import argparse
import json
import math
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable, Optional

import numpy as np
import pandas as pd


THRESHOLDS = (2.0, 5.0, 10.0, 20.0)
LENGTH_ORDER = ("1-20", "21-50", "51-100", "101-200", ">200")
CHAIN_ORDER = ("1", "2", "3", "4", ">=5")

FULL_OPTIONS = (
    "--fault-tolerant --min-pep-length 4 --min-nuc-length 4 "
    "--lddt --rigid-scores --tm-score --dockq"
)
RIGID_OPTIONS = (
    "--fault-tolerant --min-pep-length 4 --min-nuc-length 4 --rigid-scores"
)


def finite_series(series: pd.Series) -> pd.Series:
    values = pd.to_numeric(series, errors="coerce")
    return values[np.isfinite(values) & values.ge(0)]


def read_exclusion_list(path: Path) -> tuple[str, ...]:
    if not path.is_file():
        raise FileNotFoundError(f"Frozen exclusion list not found: {path}")
    result = []
    for raw in path.read_text(encoding="utf-8").splitlines():
        item = raw.strip().upper()
        if not item or item.startswith("#"):
            continue
        if not re.fullmatch(r"[0-9A-Z]{4}", item):
            raise ValueError(f"Invalid PDB ID in {path}: {raw!r}")
        result.append(item)
    if len(result) != len(set(result)):
        raise ValueError(f"Duplicate PDB IDs in {path}")
    return tuple(result)


def prepare_rank1(rank1_path: Path, frozen_excluded: set[str]) -> pd.DataFrame:
    frame = pd.read_csv(rank1_path)
    required = {
        "pdb_id",
        "seed",
        "sample",
        "ranking_score",
        "rmsd",
        "eval_status",
        "time_group",
        "rna_total_length",
        "length_group",
        "chain_count_group",
    }
    missing = sorted(required - set(frame.columns))
    if missing:
        raise ValueError(f"{rank1_path} is missing columns: {', '.join(missing)}")
    frame["pdb_id"] = frame["pdb_id"].astype(str).str.upper()
    if frame["pdb_id"].duplicated().any():
        raise ValueError("rank1_targets.csv contains duplicate PDB IDs")
    unknown = frozen_excluded - set(frame["pdb_id"])
    if unknown:
        raise ValueError(f"Exclusion list contains unknown targets: {sorted(unknown)}")
    frame["rmsd"] = pd.to_numeric(frame["rmsd"], errors="coerce")
    metric_valid = (
        frame["eval_status"].eq("SUCCESS")
        & np.isfinite(frame["rmsd"])
        & frame["rmsd"].ge(0)
    )
    frame["frozen_excluded"] = frame["pdb_id"].isin(frozen_excluded)
    frame["included_in_final_rmsd"] = metric_valid & ~frame["frozen_excluded"]
    frame["length_group"] = frame["length_group"].astype("string")
    frame["chain_count_group"] = frame["chain_count_group"].astype("string")
    return frame


def prepare_candidates(path: Path) -> pd.DataFrame:
    frame = pd.read_csv(path)
    required = {"pdb_id", "seed", "sample", "rmsd", "eval_status"}
    missing = sorted(required - set(frame.columns))
    if missing:
        raise ValueError(f"{path} is missing columns: {', '.join(missing)}")
    frame["pdb_id"] = frame["pdb_id"].astype(str).str.upper()
    frame["rmsd"] = pd.to_numeric(frame["rmsd"], errors="coerce")
    frame["valid_rmsd"] = (
        frame["eval_status"].eq("SUCCESS")
        & np.isfinite(frame["rmsd"])
        & frame["rmsd"].ge(0)
    )
    return frame


def summarize_group(frame: pd.DataFrame, *, section: str, group: str) -> dict[str, object]:
    total = len(frame)
    included = frame[frame["included_in_final_rmsd"]]
    values = finite_series(included["rmsd"])
    q25 = values.quantile(0.25) if len(values) else np.nan
    q75 = values.quantile(0.75) if len(values) else np.nan
    result: dict[str, object] = {
        "section": section,
        "group": group,
        "n_total": total,
        "n_valid": len(values),
        "n_excluded_or_missing": total - len(values),
        "coverage_percent": 100 * len(values) / total if total else np.nan,
        "median_rmsd_A": values.median() if len(values) else np.nan,
        "q25_rmsd_A": q25,
        "q75_rmsd_A": q75,
        "iqr_rmsd_A": q75 - q25 if len(values) else np.nan,
        "mean_rmsd_A": values.mean() if len(values) else np.nan,
        "p90_rmsd_A": values.quantile(0.90) if len(values) else np.nan,
        "max_rmsd_A": values.max() if len(values) else np.nan,
    }
    for threshold in THRESHOLDS:
        key = f"rmsd_le_{int(threshold)}A_percent_of_valid"
        result[key] = 100 * (values <= threshold).mean() if len(values) else np.nan
    return result


def build_main_table(rank1: pd.DataFrame) -> pd.DataFrame:
    rows = [summarize_group(rank1, section="Overall", group="All targets")]
    rows.append(
        summarize_group(
            rank1[rank1["time_group"].eq("post_cutoff")],
            section="Time",
            group="Post-cutoff (>2021-09-30)",
        )
    )
    for label in LENGTH_ORDER:
        rows.append(
            summarize_group(
                rank1[rank1["length_group"].eq(label)],
                section="RNA length",
                group=f"{label} nt",
            )
        )
    for label in CHAIN_ORDER:
        rows.append(
            summarize_group(
                rank1[rank1["chain_count_group"].eq(label)],
                section="RNA chain count",
                group=label,
            )
        )
    return pd.DataFrame(rows)


def ecdf(values: pd.Series) -> tuple[np.ndarray, np.ndarray]:
    x = np.sort(finite_series(values).to_numpy(dtype=float))
    y = np.arange(1, len(x) + 1, dtype=float) / len(x) if len(x) else np.array([])
    return x, y


def setup_matplotlib():
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    plt.rcParams.update(
        {
            "font.family": "DejaVu Sans",
            "font.size": 10,
            "axes.titlesize": 13,
            "axes.labelsize": 11,
            "legend.fontsize": 9,
            "figure.dpi": 120,
            "savefig.dpi": 300,
            "pdf.fonttype": 42,
            "ps.fonttype": 42,
        }
    )
    return plt


def save_figure(fig, base: Path) -> None:
    base.parent.mkdir(parents=True, exist_ok=True)
    for suffix in (".png", ".pdf", ".svg"):
        fig.savefig(base.with_suffix(suffix), bbox_inches="tight")


def plot_ecdf(rank1: pd.DataFrame, figures: Path) -> dict[str, object]:
    plt = setup_matplotlib()
    definitions = [
        ("All", rank1, "#1f77b4"),
        ("Pre/on cutoff", rank1[rank1["time_group"].eq("pre_or_on_cutoff")], "#2ca02c"),
        ("Post-cutoff", rank1[rank1["time_group"].eq("post_cutoff")], "#d62728"),
    ]
    pooled = rank1.loc[rank1["included_in_final_rmsd"], "rmsd"]
    cap = max(20.0, math.ceil(float(pooled.quantile(0.99)) / 5.0) * 5.0)
    fig, ax = plt.subplots(figsize=(9.5, 6.3), constrained_layout=True)
    curve_stats = {}
    for label, subset, color in definitions:
        valid = subset[subset["included_in_final_rmsd"]]
        x, y = ecdf(valid["rmsd"])
        excluded = len(subset) - len(x)
        curve_stats[label] = {
            "n_total": len(subset),
            "n_valid": len(x),
            "n_excluded": excluded,
        }
        ax.step(
            x,
            y,
            where="post",
            linewidth=2.4,
            color=color,
            label=f"{label}: valid {len(x)}/{len(subset)}, excluded {excluded}",
        )
    for threshold in THRESHOLDS:
        ax.axvline(threshold, color="#777777", linestyle="--", linewidth=0.9, alpha=0.75)
        ax.text(
            threshold,
            0.025,
            f"{int(threshold)} Å",
            rotation=90,
            va="bottom",
            ha="right",
            color="#555555",
            fontsize=8,
        )
    ax.set_xlim(0, cap)
    ax.set_ylim(0, 1.01)
    ax.set_xlabel("OpenStructure rigid C3′ RMSD threshold (Å)")
    ax.set_ylabel("Fraction of valid strict rank-1 targets with RMSD ≤ threshold")
    ax.set_title("Protenix RNA folding accuracy: strict rank-1 RMSD ECDF")
    ax.grid(True, alpha=0.22)
    ax.legend(loc="lower right", frameon=True)
    ax.text(
        0.0,
        -0.17,
        f"Linear x-axis is limited to pooled P99 ({pooled.quantile(0.99):.2f} Å; display cap {cap:.0f} Å). "
        "All extrema are retained in the tables and the full-range log figure.",
        transform=ax.transAxes,
        fontsize=8.5,
        va="top",
    )
    save_figure(fig, figures / "Figure1_RMSD_ECDF")
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(9.5, 6.3), constrained_layout=True)
    for label, subset, color in definitions:
        valid = subset[subset["included_in_final_rmsd"]]
        x, y = ecdf(valid["rmsd"])
        positive = x > 0
        ax.step(x[positive], y[positive], where="post", linewidth=2.4, color=color, label=label)
    for threshold in THRESHOLDS:
        ax.axvline(threshold, color="#777777", linestyle="--", linewidth=0.8, alpha=0.65)
    ax.set_xscale("log")
    ax.set_ylim(0, 1.01)
    ax.set_xlabel("OpenStructure rigid C3′ RMSD threshold (Å, log scale)")
    ax.set_ylabel("Fraction of valid strict rank-1 targets")
    ax.set_title("Strict rank-1 RMSD ECDF — full range")
    ax.grid(True, which="both", alpha=0.22)
    ax.legend(loc="lower right")
    save_figure(fig, figures / "FigureS1_RMSD_ECDF_full_log")
    plt.close(fig)
    return {"linear_cap_A": cap, "curves": curve_stats}


def _stratified_panel(ax, rank1: pd.DataFrame, column: str, order: Iterable[str], title: str, xlabel: str) -> None:
    rng = np.random.default_rng(20260820)
    arrays = []
    tick_labels = []
    colors = ["#4C78A8", "#72B7B2", "#F2CF5B", "#F58518", "#E45756"]
    for index, label in enumerate(order, start=1):
        subset = rank1[rank1[column].eq(label)]
        values = finite_series(subset.loc[subset["included_in_final_rmsd"], "rmsd"])
        arrays.append(values.to_numpy())
        excluded = len(subset) - len(values)
        tick_labels.append(f"{label}\nn={len(values)}, excl={excluded}")
        if len(values):
            jitter = rng.uniform(-0.18, 0.18, size=len(values))
            ax.scatter(
                np.full(len(values), index) + jitter,
                values,
                s=8,
                alpha=0.16,
                color=colors[(index - 1) % len(colors)],
                edgecolors="none",
                rasterized=True,
            )
    box = ax.boxplot(
        arrays,
        positions=np.arange(1, len(arrays) + 1),
        widths=0.52,
        showfliers=False,
        patch_artist=True,
        medianprops={"color": "black", "linewidth": 1.5},
    )
    for patch, color in zip(box["boxes"], colors):
        patch.set_facecolor(color)
        patch.set_alpha(0.58)
    ax.set_yscale("log")
    ax.set_xticks(np.arange(1, len(arrays) + 1), tick_labels)
    ax.set_xlabel(xlabel)
    ax.set_ylabel("Strict rank-1 C3′ RMSD (Å, log scale)")
    ax.set_title(title)
    ax.grid(True, axis="y", which="both", alpha=0.22)


def plot_stratified(rank1: pd.DataFrame, figures: Path) -> None:
    plt = setup_matplotlib()
    fig, axes = plt.subplots(1, 2, figsize=(15, 6.5), constrained_layout=True)
    _stratified_panel(
        axes[0], rank1, "length_group", LENGTH_ORDER,
        "RMSD stratified by total input RNA length", "Total RNA length (nt)",
    )
    _stratified_panel(
        axes[1], rank1, "chain_count_group", CHAIN_ORDER,
        "RMSD stratified by GT RNA chain count", "RNA chain count",
    )
    fig.suptitle("Protenix RNA folding performance by target complexity", fontsize=15)
    fig.text(
        0.5,
        -0.025,
        "Dots show every valid target, including extrema. Boxes show median and IQR; boxplot fliers are hidden only to avoid duplicate points. "
        "All extrema and excluded counts are retained in the tables.",
        ha="center",
        fontsize=8.5,
    )
    save_figure(fig, figures / "Figure2_RMSD_stratified_boxplots")
    plt.close(fig)


def format_main_table_for_display(table: pd.DataFrame) -> pd.DataFrame:
    result = pd.DataFrame()
    result["Section"] = table["section"]
    result["Group"] = table["group"]
    result["Valid/total"] = table.apply(lambda r: f"{int(r.n_valid)}/{int(r.n_total)}", axis=1)
    result["Coverage"] = table["coverage_percent"].map(lambda x: f"{x:.1f}%")
    result["Median Å"] = table["median_rmsd_A"].map(lambda x: f"{x:.2f}")
    result["IQR Å"] = table.apply(lambda r: f"{r.q25_rmsd_A:.2f}–{r.q75_rmsd_A:.2f}", axis=1)
    result["Mean Å"] = table["mean_rmsd_A"].map(lambda x: f"{x:.2f}")
    result["P90 Å"] = table["p90_rmsd_A"].map(lambda x: f"{x:.2f}")
    for threshold in (2, 5, 10):
        source = f"rmsd_le_{threshold}A_percent_of_valid"
        result[f"≤{threshold} Å"] = table[source].map(lambda x: f"{x:.1f}%")
    return result


def plot_main_table(table: pd.DataFrame, figures: Path) -> None:
    plt = setup_matplotlib()
    display = format_main_table_for_display(table)
    fig_height = 2.2 + 0.48 * len(display)
    fig, ax = plt.subplots(figsize=(16, fig_height))
    ax.axis("off")
    tbl = ax.table(
        cellText=display.values,
        colLabels=display.columns,
        cellLoc="center",
        colLoc="center",
        loc="center",
    )
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(8.5)
    tbl.scale(1.0, 1.45)
    section_colors = {
        "Overall": "#DCEAF7",
        "Time": "#FADBD8",
        "RNA length": "#E8F5E9",
        "RNA chain count": "#FFF3CD",
    }
    for (row, col), cell in tbl.get_celld().items():
        if row == 0:
            cell.set_facecolor("#1F4E78")
            cell.set_text_props(color="white", weight="bold")
        else:
            section = display.iloc[row - 1]["Section"]
            cell.set_facecolor(section_colors.get(section, "white"))
            if col in (0, 1):
                cell.set_text_props(ha="left")
    ax.set_title("Table 1. Strict rank-1 RNA RMSD summary", fontsize=14, pad=15)
    fig.text(
        0.5,
        0.015,
        "Threshold fractions use valid strict rank-1 targets within each row as denominator. Coverage reports valid/total after the frozen exclusion list.",
        ha="center",
        fontsize=8.5,
    )
    save_figure(fig, figures / "Table1_main_RMSD_results")
    plt.close(fig)


def read_json_status(path: Path) -> str:
    if not path.is_file():
        return "NOT_AVAILABLE"
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return "INVALID_JSON"
    return str(payload.get("status", "UNKNOWN"))


def evaluation_status(json_path: Path, error_path: Path) -> str:
    status = read_json_status(json_path)
    if status == "NOT_AVAILABLE" and error_path.is_file():
        return "FAILED"
    return status


def classify_stage(text: str) -> tuple[str, str]:
    lower = text.lower()
    if "gdt" in lower and "window size" in lower:
        return "RIGID_GDT", "Too few mapped positions for GDT before RMSD serialization"
    if "arrays used as indices must be of integer (or boolean) type" in lower:
        return "CHAIN_MAPPING", "OpenStructure 2.8 NumPy integer/boolean index error"
    if "computing rmsd" in lower and "error! no assignable chain" in lower:
        return "DOWNSTREAM_TM_SCORE", "USalign reported no assignable chain after RMSD stage was reached"
    if "need at least one array to concatenate" in lower:
        return "ALL_ATOM_LDDT", "OpenStructure 2.8 lDDT empty-array error"
    if "timeout after" in lower:
        return "TIMEOUT", "Per-candidate evaluation timeout"
    if "computing chain mapping" in lower:
        return "CHAIN_MAPPING", "Unclassified failure during/after chain mapping"
    return "UNKNOWN", "See stderr excerpt"


def build_command(ost: str, model: str, reference: str, output: str, options: str) -> str:
    return f"{ost} compare-structures -m {model} -r {reference} -o {output} {options}"


def build_supplement(
    root: Path,
    rank1: pd.DataFrame,
    candidates: pd.DataFrame,
    excluded_ids: tuple[str, ...],
    audit_path: Path,
) -> pd.DataFrame:
    audit = pd.read_csv(audit_path, sep="\t") if audit_path.is_file() else pd.DataFrame()
    if not audit.empty:
        audit["pdb_id"] = audit["pdb_id"].astype(str).str.upper()
        audit = audit.set_index("pdb_id", drop=False)
    rows = []
    for pdb_id in excluded_ids:
        selected = rank1.loc[rank1["pdb_id"].eq(pdb_id)].iloc[0]
        seed, sample = int(selected["seed"]), int(selected["sample"])
        standard_json = root / "details" / pdb_id / f"seed_{seed}" / f"sample_{sample}.json"
        standard_err = root / "errors" / pdb_id / f"seed_{seed}" / f"sample_{sample}.stderr.txt"
        rescue_json = root / "rigid_only_rescue" / "details" / pdb_id / f"seed_{seed}" / f"sample_{sample}.json"
        rescue_err = root / "rigid_only_rescue" / "errors" / pdb_id / f"seed_{seed}" / f"sample_{sample}.stderr.txt"
        error_text = "\n".join(
            path.read_text(encoding="utf-8", errors="replace")
            for path in (rescue_err, standard_err)
            if path.is_file()
        )
        stage, detail = classify_stage(error_text)
        target_candidates = candidates[candidates["pdb_id"].eq(pdb_id)]
        lower_valid = target_candidates[
            target_candidates["valid_rmsd"]
            & ~(
                target_candidates["seed"].eq(seed)
                & target_candidates["sample"].eq(sample)
            )
        ]
        model = str(selected.get("prediction_path", ""))
        reference = str(selected.get("reference_path", ""))
        ost = "/storage9920/home/tinghao.xia/miniconda3/envs/foldbench/bin/ost"
        audit_reason = ""
        if not audit.empty and pdb_id in audit.index:
            audit_reason = str(audit.loc[pdb_id].get("exclusion_reason", ""))
        rows.append(
            {
                "pdb_id": pdb_id,
                "time_group": selected.get("time_group", ""),
                "rna_total_length": selected.get("rna_total_length", np.nan),
                "chain_count": selected.get("chain_count", np.nan),
                "rank1_seed": seed,
                "rank1_sample": sample,
                "rank1_ranking_score": selected.get("ranking_score", np.nan),
                "failure_stage": stage,
                "failure_detail": detail,
                "audit_exclusion_reason": audit_reason,
                "standard_result": evaluation_status(standard_json, standard_err),
                "standard_command": build_command(ost, model, reference, str(standard_json), FULL_OPTIONS),
                "rescue_attempted": rescue_json.is_file() or rescue_err.is_file(),
                "rescue_result": evaluation_status(rescue_json, rescue_err)
                if rescue_json.is_file() or rescue_err.is_file()
                else "NOT_ATTEMPTED",
                "rescue_command": build_command(ost, model, reference, str(rescue_json), RIGID_OPTIONS),
                "lower_rank_valid_candidate_available": len(lower_valid) > 0,
                "lower_rank_valid_candidate_count": len(lower_valid),
                "error_log_excerpt": error_text[-3000:].replace("\r", " ").replace("\n", " "),
            }
        )
    return pd.DataFrame(rows)


def autosize_workbook(path: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = load_workbook(path)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for column_cells in sheet.iter_cols(min_row=1, max_row=min(sheet.max_row, 250)):
            header = str(column_cells[0].value or "")
            width = max(len(str(cell.value or "")) for cell in column_cells) + 2
            cap = 80 if "command" in header or "excerpt" in header else 38
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width, 10), cap)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    workbook.save(path)


def write_outputs(
    output: Path,
    *,
    root: Path,
    rank1: pd.DataFrame,
    main_table: pd.DataFrame,
    supplement: pd.DataFrame,
    excluded_ids: tuple[str, ...],
    ecdf_meta: dict[str, object],
) -> None:
    tables = output / "tables"
    tables.mkdir(parents=True, exist_ok=True)
    (output / "frozen_excluded_pdb_ids.txt").write_text(
        "".join(f"{pdb_id}\n" for pdb_id in excluded_ids), encoding="utf-8"
    )
    main_table.to_csv(tables / "Table1_main_RMSD_results.tsv", sep="\t", index=False)
    supplement.to_csv(tables / "TableS1_excluded_PDB.tsv", sep="\t", index=False)
    rank1.loc[rank1["included_in_final_rmsd"]].to_csv(
        tables / "strict_rank1_valid_targets.csv", index=False
    )
    definitions = pd.DataFrame(
        {
            "item": [
                "Primary metric",
                "Rank-1 selection",
                "Frozen exclusion rule",
                "Threshold denominator",
                "Cutoff",
                "Standard command",
                "Rescue command",
            ],
            "definition": [
                "OpenStructure 2.8 rigid-score RMSD over mapped nucleic-acid C3' atoms after one global fit.",
                "Maximum Protenix ranking_score before consulting any ground-truth metric.",
                f"Targets listed in exclude_strict_rank1_rmsd_pdb.txt remain excluded (n={len(excluded_ids)}).",
                "Percentages at 2/5/10/20 Å use valid, non-excluded strict rank-1 targets within each row.",
                "post_cutoff means release date after 2021-09-30.",
                FULL_OPTIONS,
                RIGID_OPTIONS,
            ],
        }
    )
    workbook = output / "RNA_RMSD_final_report.xlsx"
    with pd.ExcelWriter(workbook, engine="openpyxl") as writer:
        definitions.to_excel(writer, sheet_name="README", index=False)
        main_table.to_excel(writer, sheet_name="Main_results", index=False)
        supplement.to_excel(writer, sheet_name="Excluded_PDB", index=False)
        rank1.loc[rank1["included_in_final_rmsd"]].to_excel(
            writer, sheet_name="Rank1_valid", index=False
        )
    autosize_workbook(workbook)

    overall = main_table.iloc[0]
    post = main_table[
        main_table["group"].eq("Post-cutoff (>2021-09-30)")
    ].iloc[0]
    summary = [
        "Final Protenix RNA strict rank-1 C3' RMSD report",
        "",
        f"Total targets: {len(rank1)}",
        f"Frozen excluded targets: {len(excluded_ids)}",
        f"Valid targets: {int(overall['n_valid'])}",
        f"Coverage: {overall['coverage_percent']:.2f}%",
        "",
        "Overall valid targets:",
        f"  median: {overall['median_rmsd_A']:.3f} A",
        f"  IQR: {overall['q25_rmsd_A']:.3f}-{overall['q75_rmsd_A']:.3f} A",
        f"  mean: {overall['mean_rmsd_A']:.3f} A",
        f"  P90: {overall['p90_rmsd_A']:.3f} A",
        f"  <=2 A: {overall['rmsd_le_2A_percent_of_valid']:.2f}%",
        f"  <=5 A: {overall['rmsd_le_5A_percent_of_valid']:.2f}%",
        f"  <=10 A: {overall['rmsd_le_10A_percent_of_valid']:.2f}%",
        "",
        "Post-cutoff valid targets:",
        f"  valid/total: {int(post['n_valid'])}/{int(post['n_total'])}",
        f"  median: {post['median_rmsd_A']:.3f} A",
        f"  IQR: {post['q25_rmsd_A']:.3f}-{post['q75_rmsd_A']:.3f} A",
        f"  mean: {post['mean_rmsd_A']:.3f} A",
        f"  P90: {post['p90_rmsd_A']:.3f} A",
        f"  <=10 A: {post['rmsd_le_10A_percent_of_valid']:.2f}%",
    ]
    (output / "final_report_summary.txt").write_text("\n".join(summary) + "\n", encoding="utf-8")

    captions = [
        "Figure 1. Empirical cumulative distribution of strict rank-1 OpenStructure rigid C3' RMSD for all, pre/on-cutoff, and post-cutoff targets. Vertical reference lines mark 2, 5, 10, and 20 A. Valid and frozen-excluded target counts are shown in the legend. The linear panel is capped at pooled P99; the supplementary log panel preserves the full range.",
        "",
        "Figure 2. Strict rank-1 C3' RMSD stratified by total input RNA length and GT RNA chain count. Dots show all valid targets, boxes show median and interquartile range, and the y-axis is logarithmic. Excluded counts are shown under each stratum; all numerical extrema remain in Table 1.",
        "",
        "Table 1. Main strict rank-1 RMSD results. Threshold fractions use valid targets within each stratum; coverage reports the valid fraction before exclusions are removed from metric summaries.",
        "",
        "Table S1. Frozen excluded PDB targets with failure stage, exact standard/rescue command, rescue status, and availability of a lower-ranked candidate with valid RMSD.",
    ]
    (output / "figure_and_table_captions.txt").write_text("\n".join(captions) + "\n", encoding="utf-8")
    manifest = {
        "generated_at_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "input_root": str(root),
        "output_root": str(output),
        "target_count": len(rank1),
        "frozen_excluded_count": len(excluded_ids),
        "frozen_excluded_pdb_ids": list(excluded_ids),
        "ecdf": ecdf_meta,
        "thresholds_A": list(THRESHOLDS),
    }
    (output / "report_manifest.json").write_text(
        json.dumps(manifest, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )


def run(args: argparse.Namespace) -> int:
    root = Path(args.rmsd_root).expanduser().resolve()
    reports = root / "reports"
    output = Path(args.output_dir).expanduser().resolve() if args.output_dir else root / "final_report"
    output.mkdir(parents=True, exist_ok=True)
    figures = output / "figures"
    excluded_path = (
        Path(args.exclusion_list).expanduser().resolve()
        if args.exclusion_list
        else reports / "exclude_strict_rank1_rmsd_pdb.txt"
    )
    excluded_ids = read_exclusion_list(excluded_path)
    frozen_excluded = set(excluded_ids)
    rank1 = prepare_rank1(reports / "rank1_targets.csv", frozen_excluded)
    expected_targets = getattr(args, "expected_targets", None)
    expected_excluded = getattr(args, "expected_excluded", None)
    expected_valid = getattr(args, "expected_valid", None)
    if expected_targets is not None and len(rank1) != expected_targets:
        raise ValueError(
            f"Expected {expected_targets} total targets, found {len(rank1)}"
        )
    if expected_excluded is not None and len(excluded_ids) != expected_excluded:
        raise ValueError(
            f"Expected {expected_excluded} frozen exclusions, found {len(excluded_ids)}"
        )
    valid_count = int(rank1["included_in_final_rmsd"].sum())
    if expected_valid is not None and valid_count != expected_valid:
        raise ValueError(
            f"Expected {expected_valid} valid final targets, found {valid_count}"
        )
    candidates = prepare_candidates(reports / "all_candidates.csv")
    main_table = build_main_table(rank1)
    supplement = build_supplement(
        root,
        rank1,
        candidates,
        excluded_ids,
        reports / "pdb_rmsd_audit.tsv",
    )
    ecdf_meta = plot_ecdf(rank1, figures)
    plot_stratified(rank1, figures)
    plot_main_table(main_table, figures)
    write_outputs(
        output,
        root=root,
        rank1=rank1,
        main_table=main_table,
        supplement=supplement,
        excluded_ids=excluded_ids,
        ecdf_meta=ecdf_meta,
    )
    print((output / "final_report_summary.txt").read_text(encoding="utf-8"))
    print(f"Final report written to: {output}")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--rmsd-root", default="~/Json_data/Foldbench_evaluation/rmsd")
    parser.add_argument("--exclusion-list", default=None)
    parser.add_argument("--output-dir", default=None)
    parser.add_argument("--expected-targets", type=int, default=None)
    parser.add_argument("--expected-excluded", type=int, default=None)
    parser.add_argument("--expected-valid", type=int, default=None)
    return parser


def main(argv: Optional[list[str]] = None) -> int:
    return run(build_parser().parse_args(argv))


if __name__ == "__main__":
    raise SystemExit(main())
