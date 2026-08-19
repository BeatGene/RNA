#!/usr/bin/env python3
"""Summarize candidate-level FoldBench RNA metrics and strict rank-1 results."""

from __future__ import annotations

import argparse
import json
import math
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd


LENGTH_BIN_LABELS = ("1-20", "21-50", "51-100", "101-200", ">200")
METRIC_COLUMNS = ("rmsd", "lddt", "tm_score", "oligo_gdtts", "oligo_gdtha")


def finite_number(value: object) -> float:
    try:
        result = float(value)
    except (TypeError, ValueError, OverflowError):
        return np.nan
    return result if math.isfinite(result) else np.nan


def result_path(output_root: Path, pdb_id: str, seed: int, sample: int) -> Path:
    return output_root / "details" / pdb_id / f"seed_{seed}" / f"sample_{sample}.json"


def rescue_result_path(output_root: Path, pdb_id: str, seed: int, sample: int) -> Path:
    return (
        output_root
        / "rigid_only_rescue"
        / "details"
        / pdb_id
        / f"seed_{seed}"
        / f"sample_{sample}.json"
    )


def read_ost_result(path: Path) -> dict[str, object]:
    result: dict[str, object] = {
        "eval_status": "missing_output",
        "ost_status": None,
        "eval_issue": "output JSON does not exist",
        **{name: np.nan for name in METRIC_COLUMNS},
    }
    if not path.is_file():
        return result
    try:
        with path.open("r", encoding="utf-8") as handle:
            payload = json.load(handle)
    except Exception as exc:
        result["eval_status"] = "invalid_output"
        result["eval_issue"] = f"{type(exc).__name__}: {exc}"
        return result
    if not isinstance(payload, dict):
        result["eval_status"] = "invalid_output"
        result["eval_issue"] = "JSON root is not an object"
        return result
    result["ost_status"] = payload.get("status")
    for metric in METRIC_COLUMNS:
        result[metric] = finite_number(payload.get(metric))
    if str(payload.get("status", "")).upper() != "SUCCESS":
        result["eval_status"] = "ost_failure"
        result["eval_issue"] = f"OST status={payload.get('status')!r}"
    elif not np.isfinite(result["rmsd"]) or float(result["rmsd"]) < 0:
        result["eval_status"] = "invalid_output"
        result["eval_issue"] = f"invalid RMSD={payload.get('rmsd')!r}"
    else:
        result["eval_status"] = "SUCCESS"
        result["eval_issue"] = ""
    return result


def load_candidate_results(output_root: Path, manifest_path: Path) -> pd.DataFrame:
    manifest = pd.read_csv(
        manifest_path,
        dtype={"pdb_id": "string", "seed": "int64", "sample": "int64"},
    )
    required = {
        "pdb_id",
        "seed",
        "sample",
        "prediction_path",
        "reference_path",
        "confidence_path",
        "ranking_score",
        "discovery_issue",
    }
    missing = sorted(required - set(manifest.columns))
    if missing:
        raise ValueError(f"Manifest is missing columns: {', '.join(missing)}")
    manifest["pdb_id"] = manifest["pdb_id"].str.upper()
    duplicate = manifest.duplicated(["pdb_id", "seed", "sample"], keep=False)
    if duplicate.any():
        keys = manifest.loc[duplicate, ["pdb_id", "seed", "sample"]]
        raise ValueError(f"Duplicate candidate keys in manifest:\n{keys.to_string(index=False)}")
    records = []
    for row in manifest.itertuples(index=False):
        path = result_path(output_root, row.pdb_id, int(row.seed), int(row.sample))
        rescue_path = rescue_result_path(
            output_root, row.pdb_id, int(row.seed), int(row.sample)
        )
        standard_result = read_ost_result(path)
        rescue_result = read_ost_result(rescue_path)
        if standard_result["eval_status"] == "SUCCESS":
            selected_result = standard_result
            selected_path = path
            protocol = "foldbench_full"
        elif rescue_result["eval_status"] == "SUCCESS":
            selected_result = rescue_result
            selected_path = rescue_path
            protocol = "rigid_only_rescue"
        else:
            selected_result = standard_result
            selected_path = path
            protocol = "unavailable"
        records.append(
            {
                "pdb_id": row.pdb_id,
                "seed": int(row.seed),
                "sample": int(row.sample),
                "ranking_score": finite_number(row.ranking_score),
                "prediction_path": row.prediction_path,
                "reference_path": row.reference_path,
                "confidence_path": row.confidence_path,
                "discovery_issue": row.discovery_issue
                if isinstance(row.discovery_issue, str)
                else "",
                "evaluation_protocol": protocol,
                "evaluation_json": str(selected_path),
                "standard_evaluation_json": str(path),
                "rigid_rescue_json": str(rescue_path),
                "standard_eval_status": standard_result["eval_status"],
                "standard_eval_issue": standard_result["eval_issue"],
                **selected_result,
            }
        )
    return pd.DataFrame.from_records(records).sort_values(
        ["pdb_id", "seed", "sample"], ignore_index=True
    )


def select_strict_rank1(all_candidates: pd.DataFrame) -> pd.DataFrame:
    """Select highest ranking_score per target without metric-based fallback."""
    target_rows = []
    for pdb_id, group in all_candidates.groupby("pdb_id", sort=True):
        ranked = group[np.isfinite(group["ranking_score"])].sort_values(
            ["ranking_score", "seed", "sample"],
            ascending=[False, True, True],
            kind="mergesort",
        )
        if ranked.empty:
            selected = group.sort_values(["seed", "sample"], kind="mergesort").iloc[0].copy()
            selected["rank1_selection_status"] = "NO_FINITE_RANKING_SCORE"
        else:
            selected = ranked.iloc[0].copy()
            selected["rank1_selection_status"] = "SUCCESS"
        valid_rmsd = pd.to_numeric(group["rmsd"], errors="coerce").dropna()
        selected["candidate_count"] = len(group)
        selected["evaluated_candidate_count"] = int((group["eval_status"] == "SUCCESS").sum())
        selected["oracle_min_rmsd"] = valid_rmsd.min() if not valid_rmsd.empty else np.nan
        selected["candidate_median_rmsd"] = (
            valid_rmsd.median() if not valid_rmsd.empty else np.nan
        )
        selected["candidate_mean_rmsd"] = valid_rmsd.mean() if not valid_rmsd.empty else np.nan
        top_score = finite_number(selected["ranking_score"])
        selected["ranking_score_tie_count"] = (
            int(np.isclose(group["ranking_score"], top_score, rtol=0, atol=1e-12).sum())
            if np.isfinite(top_score)
            else 0
        )
        target_rows.append(selected)
    return pd.DataFrame(target_rows).reset_index(drop=True)


def load_target_metadata(path: Path) -> pd.DataFrame:
    frame = pd.read_excel(path, sheet_name="Targets", dtype={"PDB_id": "string"})
    if "PDB_id" not in frame:
        raise ValueError(f"Targets sheet in {path} has no PDB_id column")
    frame["pdb_id"] = frame["PDB_id"].str.upper()
    if frame["pdb_id"].duplicated().any():
        raise ValueError(f"Duplicate PDB_id values in {path}")
    keep = [
        column
        for column in (
            "pdb_id",
            "release_date",
            "time_group",
            "chain_count",
            "mapping_status",
            "review_status",
        )
        if column in frame
    ]
    return frame[keep]


def index_simple_jsons(root: Path) -> dict[str, Path]:
    result = {}
    if not root.is_dir():
        return result
    for path in root.iterdir():
        if path.is_file() and path.suffix.lower() == ".json":
            pdb_id = path.stem.upper()
            if len(pdb_id) == 4:
                if pdb_id in result:
                    raise ValueError(f"Duplicate case-insensitive Simple_json for {pdb_id}")
                result[pdb_id] = path
    return result


def rna_input_size(path: Path) -> tuple[float, float, str]:
    try:
        with path.open("r", encoding="utf-8") as handle:
            payload = json.load(handle)
        entries = payload if isinstance(payload, list) else [payload]
        total_length = 0
        chain_count = 0
        for entry in entries:
            for sequence_item in entry.get("sequences", []):
                if "rnaSequence" not in sequence_item:
                    continue
                item = sequence_item["rnaSequence"]
                sequence = str(item["sequence"]).strip().upper()
                count = int(item.get("count", 1))
                if count < 1:
                    raise ValueError(f"invalid rnaSequence count={count}")
                total_length += len(sequence) * count
                chain_count += count
        if chain_count == 0:
            raise ValueError("no rnaSequence entries")
        return float(total_length), float(chain_count), ""
    except Exception as exc:
        return np.nan, np.nan, f"{type(exc).__name__}: {exc}"


def add_target_metadata(
    rank1: pd.DataFrame,
    *,
    targets_xlsx: Path,
    simple_json_root: Path,
) -> pd.DataFrame:
    metadata = load_target_metadata(targets_xlsx)
    merged = rank1.merge(metadata, on="pdb_id", how="left", validate="one_to_one")
    json_index = index_simple_jsons(simple_json_root)
    lengths, input_chains, issues = [], [], []
    for pdb_id in merged["pdb_id"]:
        path = json_index.get(pdb_id)
        if path is None:
            lengths.append(np.nan)
            input_chains.append(np.nan)
            issues.append("Simple_json is missing")
        else:
            length, count, issue = rna_input_size(path)
            lengths.append(length)
            input_chains.append(count)
            issues.append(issue)
    merged["rna_total_length"] = lengths
    merged["input_rna_chain_count"] = input_chains
    merged["length_metadata_issue"] = issues
    merged["length_group"] = pd.cut(
        merged["rna_total_length"],
        bins=[0, 20, 50, 100, 200, np.inf],
        labels=LENGTH_BIN_LABELS,
        include_lowest=True,
        right=True,
    )
    chain_numeric = pd.to_numeric(merged.get("chain_count"), errors="coerce")
    merged["chain_count_group"] = chain_numeric.map(
        lambda value: (
            "unknown"
            if not np.isfinite(value)
            else (">=5" if value >= 5 else str(int(value)))
        )
    )
    return merged


def describe_rmsd(frame: pd.DataFrame, *, label: str) -> dict[str, object]:
    values = pd.to_numeric(frame["rmsd"], errors="coerce")
    finite = values[np.isfinite(values)]
    result: dict[str, object] = {
        "group": label,
        "n_total": len(frame),
        "n_with_rmsd": len(finite),
        "n_missing_rmsd": len(frame) - len(finite),
        "rmsd_coverage_percent": 100 * len(finite) / len(frame) if len(frame) else np.nan,
    }
    statistics = {
        "mean": finite.mean() if len(finite) else np.nan,
        "std": finite.std(ddof=1) if len(finite) > 1 else np.nan,
        "min": finite.min() if len(finite) else np.nan,
        "p10": finite.quantile(0.10) if len(finite) else np.nan,
        "p25": finite.quantile(0.25) if len(finite) else np.nan,
        "median": finite.median() if len(finite) else np.nan,
        "p75": finite.quantile(0.75) if len(finite) else np.nan,
        "p90": finite.quantile(0.90) if len(finite) else np.nan,
        "p95": finite.quantile(0.95) if len(finite) else np.nan,
        "p99": finite.quantile(0.99) if len(finite) else np.nan,
        "max": finite.max() if len(finite) else np.nan,
    }
    result.update({f"rmsd_{key}_angstrom": value for key, value in statistics.items()})
    for threshold in (2, 4, 5, 10):
        result[f"rmsd_le_{threshold}A_percent"] = (
            100 * (finite <= threshold).mean() if len(finite) else np.nan
        )
    return result


def grouped_statistics(frame: pd.DataFrame, column: str) -> pd.DataFrame:
    rows = []
    grouped = frame.groupby(column, observed=False, dropna=False, sort=False)
    for value, group in grouped:
        label = "missing" if pd.isna(value) else str(value)
        rows.append(describe_rmsd(group, label=label))
    return pd.DataFrame(rows)


def build_statistics(
    all_candidates: pd.DataFrame,
    rank1: pd.DataFrame,
) -> dict[str, pd.DataFrame]:
    valid_candidates = all_candidates[all_candidates["eval_status"] == "SUCCESS"]
    oracle = rank1.copy()
    oracle["rmsd"] = oracle["oracle_min_rmsd"]
    overall = pd.DataFrame(
        [
            describe_rmsd(rank1, label="rank1_by_ranking_score"),
            describe_rmsd(valid_candidates, label="all_successful_candidates"),
            describe_rmsd(oracle, label="oracle_best_of_available_candidates"),
        ]
    )
    return {
        "Overall": overall,
        "By_time": grouped_statistics(rank1, "time_group"),
        "By_length": grouped_statistics(rank1, "length_group"),
        "By_chain_count": grouped_statistics(rank1, "chain_count_group"),
    }


def classify_candidate_failure(output_root: Path, row: object) -> str:
    if getattr(row, "eval_status") == "SUCCESS" and np.isfinite(
        finite_number(getattr(row, "rmsd"))
    ):
        return ""
    pdb_id = str(getattr(row, "pdb_id"))
    seed = int(getattr(row, "seed"))
    sample = int(getattr(row, "sample"))
    paths = (
        output_root
        / "rigid_only_rescue"
        / "errors"
        / pdb_id
        / f"seed_{seed}"
        / f"sample_{sample}.stderr.txt",
        output_root
        / "errors"
        / pdb_id
        / f"seed_{seed}"
        / f"sample_{sample}.stderr.txt",
    )
    text = "\n".join(
        path.read_text(encoding="utf-8", errors="replace")
        for path in paths
        if path.is_file()
    )
    lower = text.lower()
    if "gdt" in lower and "window size" in lower:
        return "GDT_TOO_FEW_MAPPED_POSITIONS"
    if "arrays used as indices must be of integer (or boolean) type" in lower:
        return "CHAIN_MAPPING_INDEX_ERROR"
    if "need at least one array to concatenate" in lower:
        return "LDDT_EMPTY_ARRAY"
    if "computing rmsd" in lower and "error! no assignable chain" in lower:
        return "DOWNSTREAM_TM_NO_ASSIGNABLE_CHAIN_AFTER_RMSD"
    if "timeout after" in lower:
        return "TIMEOUT"
    if "ost status is 'failure'" in lower:
        return "OST_FAILURE_OTHER"
    return "NO_ERROR_LOG" if not text else "OTHER"


def build_pdb_rmsd_audit(
    all_candidates: pd.DataFrame,
    rank1: pd.DataFrame,
) -> pd.DataFrame:
    candidate_summary = (
        all_candidates.groupby("pdb_id", as_index=False)
        .agg(
            candidate_count=("sample", "size"),
            valid_candidate_count=("valid_rmsd", "sum"),
        )
    )
    candidate_summary["failed_candidate_count"] = (
        candidate_summary["candidate_count"]
        - candidate_summary["valid_candidate_count"]
    )
    reasons = (
        all_candidates.loc[~all_candidates["valid_rmsd"]]
        .groupby("pdb_id")["rmsd_failure_reason"]
        .agg(lambda values: ";".join(sorted(set(values))))
        .rename("candidate_failure_reasons")
        .reset_index()
    )
    rank_columns = [
        "pdb_id",
        "seed",
        "sample",
        "ranking_score",
        "rmsd",
        "rank1_valid_rmsd",
        "eval_status",
        "evaluation_protocol",
    ]
    rank_audit = rank1[rank_columns].rename(
        columns={
            "seed": "rank1_seed",
            "sample": "rank1_sample",
            "rmsd": "rank1_rmsd",
            "eval_status": "rank1_eval_status",
            "evaluation_protocol": "rank1_evaluation_protocol",
        }
    )
    audit = candidate_summary.merge(
        rank_audit, on="pdb_id", how="outer", validate="one_to_one"
    ).merge(reasons, on="pdb_id", how="left", validate="one_to_one")
    audit["exclude_from_strict_rank1_rmsd"] = ~audit["rank1_valid_rmsd"]
    audit["exclusion_reason"] = np.select(
        [
            audit["valid_candidate_count"].eq(0),
            ~audit["rank1_valid_rmsd"],
        ],
        [
            "NO_CANDIDATE_HAS_VALID_RMSD",
            "STRICT_RANK1_RMSD_FAILED_BUT_LOWER_CANDIDATE_AVAILABLE",
        ],
        default="",
    )
    return audit.sort_values("pdb_id", ignore_index=True)


def write_pdb_audit_outputs(
    report_dir: Path,
    *,
    all_candidates: pd.DataFrame,
    audit: pd.DataFrame,
) -> None:
    audit.to_csv(report_dir / "pdb_rmsd_audit.tsv", sep="\t", index=False)
    excluded = audit[audit["exclude_from_strict_rank1_rmsd"]]
    no_candidate = audit[audit["valid_candidate_count"].eq(0)]
    partial = audit[
        audit["valid_candidate_count"].gt(0)
        & audit["failed_candidate_count"].gt(0)
    ]
    (report_dir / "exclude_strict_rank1_rmsd_pdb.txt").write_text(
        "".join(f"{item}\n" for item in excluded["pdb_id"]), encoding="utf-8"
    )
    (report_dir / "no_valid_candidate_rmsd_pdb.txt").write_text(
        "".join(f"{item}\n" for item in no_candidate["pdb_id"]), encoding="utf-8"
    )
    failed = all_candidates.loc[~all_candidates["valid_rmsd"]]
    failed.to_csv(
        report_dir / "candidate_rmsd_failures_audited.tsv", sep="\t", index=False
    )
    candidate_reasons = failed["rmsd_failure_reason"].value_counts()
    lines = [
        f"Total PDB targets: {len(audit)}",
        f"PDBs with valid strict rank-1 RMSD: {int(audit['rank1_valid_rmsd'].sum())}",
        f"PDBs excluded from strict rank-1 RMSD: {len(excluded)}",
        f"PDBs with no valid RMSD among any candidates: {len(no_candidate)}",
        f"PDBs with partial candidate failures: {len(partial)}",
        "",
        "Failed candidate counts by reason:",
    ]
    lines.extend(f"  {reason}: {count}" for reason, count in candidate_reasons.items())
    lines.extend(["", "Excluded PDB IDs:"])
    lines.extend(f"  {item}" for item in excluded["pdb_id"])
    (report_dir / "pdb_rmsd_audit_summary.txt").write_text(
        "\n".join(lines) + "\n", encoding="utf-8"
    )


def ecdf(values: pd.Series) -> tuple[np.ndarray, np.ndarray]:
    array = np.sort(pd.to_numeric(values, errors="coerce").dropna().to_numpy())
    return array, np.arange(1, len(array) + 1) / len(array) if len(array) else np.array([])


def write_plots(rank1: pd.DataFrame, report_dir: Path) -> None:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    report_dir.mkdir(parents=True, exist_ok=True)
    values = pd.to_numeric(rank1["rmsd"], errors="coerce").dropna()
    if values.empty:
        return
    p99 = float(values.quantile(0.99))
    x_cap = max(p99, 1.0)

    fig, axes = plt.subplots(2, 2, figsize=(13, 9), constrained_layout=True)
    clipped = values[values <= x_cap]
    axes[0, 0].hist(clipped, bins=45, color="#4C78A8", edgecolor="white")
    axes[0, 0].axvline(values.median(), color="#E45756", linestyle="--", label="median")
    axes[0, 0].axvline(values.mean(), color="#F2CF5B", linestyle=":", label="mean")
    axes[0, 0].set(
        title=f"Rank-1 RMSD distribution (x <= P99={p99:.2f} Å)",
        xlabel="RMSD (Å)",
        ylabel="Targets",
    )
    axes[0, 0].legend()

    x, y = ecdf(values)
    axes[0, 1].plot(x, y, color="#4C78A8", linewidth=2)
    axes[0, 1].set_xlim(0, x_cap)
    axes[0, 1].set_ylim(0, 1.01)
    axes[0, 1].grid(alpha=0.25)
    axes[0, 1].set(
        title="Rank-1 RMSD empirical CDF",
        xlabel="RMSD (Å)",
        ylabel="Fraction of targets",
    )

    palette = {"pre_or_on_cutoff": "#4C78A8", "post_cutoff": "#E45756"}
    for label, group in rank1.groupby("time_group", dropna=False):
        group_x, group_y = ecdf(group["rmsd"])
        if not len(group_x):
            continue
        text = "missing" if pd.isna(label) else str(label)
        axes[1, 0].plot(
            group_x,
            group_y,
            linewidth=2,
            label=f"{text} (n={len(group_x)})",
            color=palette.get(text),
        )
    axes[1, 0].set_xlim(0, x_cap)
    axes[1, 0].set_ylim(0, 1.01)
    axes[1, 0].grid(alpha=0.25)
    axes[1, 0].set(
        title="RMSD ECDF by release-time group",
        xlabel="RMSD (Å)",
        ylabel="Fraction of targets",
    )
    axes[1, 0].legend(fontsize=8)

    box_values, labels = [], []
    for label in LENGTH_BIN_LABELS:
        group = pd.to_numeric(
            rank1.loc[rank1["length_group"].astype("string") == label, "rmsd"],
            errors="coerce",
        ).dropna()
        if len(group):
            box_values.append(group.to_numpy())
            labels.append(f"{label}\n(n={len(group)})")
    if box_values:
        axes[1, 1].boxplot(box_values, tick_labels=labels, showfliers=False)
        axes[1, 1].set_ylim(bottom=0)
    axes[1, 1].set(
        title="Rank-1 RMSD by total input RNA length\n(outliers hidden; retained in tables)",
        xlabel="Total nucleotides",
        ylabel="RMSD (Å)",
    )
    axes[1, 1].tick_params(axis="x", labelsize=8)

    fig.suptitle(
        f"Protenix RNA FoldBench-style RMSD — strict rank-1 (n={len(values)})",
        fontsize=15,
    )
    fig.savefig(report_dir / "rmsd_overview.png", dpi=200)
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(8, 5), constrained_layout=True)
    positive = values[values > 0]
    x, y = ecdf(positive)
    ax.plot(x, y, linewidth=2, color="#4C78A8")
    ax.set_xscale("log")
    ax.set_ylim(0, 1.01)
    ax.grid(alpha=0.25, which="both")
    ax.set(
        title="Rank-1 RMSD empirical CDF — full range",
        xlabel="RMSD (Å, logarithmic scale)",
        ylabel="Fraction of targets",
    )
    fig.savefig(report_dir / "rmsd_ecdf_full_log.png", dpi=200)
    plt.close(fig)


def autosize_workbook(path: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill

    workbook = load_workbook(path)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        for column_cells in sheet.iter_cols(min_row=1, max_row=min(sheet.max_row, 250)):
            width = max(len(str(cell.value or "")) for cell in column_cells) + 2
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width, 10), 45)
    workbook.save(path)


def write_text_summary(
    path: Path,
    *,
    all_candidates: pd.DataFrame,
    rank1: pd.DataFrame,
    overall: pd.DataFrame,
) -> None:
    primary = overall.loc[overall["group"] == "rank1_by_ranking_score"].iloc[0]
    lines = [
        "FoldBench-style RNA RMSD final summary",
        "",
        f"Candidate rows: {len(all_candidates)}",
        f"Successful candidate evaluations: {(all_candidates['eval_status'] == 'SUCCESS').sum()}",
        f"  standard FoldBench full protocol: {(all_candidates['evaluation_protocol'] == 'foldbench_full').sum()}",
        f"  rigid-only metric-stage rescue: {(all_candidates['evaluation_protocol'] == 'rigid_only_rescue').sum()}",
        f"Failed or missing candidate evaluations: {(all_candidates['eval_status'] != 'SUCCESS').sum()}",
        f"Targets: {len(rank1)}",
        f"Rank-1 targets with RMSD: {int(primary['n_with_rmsd'])}",
        f"Rank-1 targets missing RMSD: {int(primary['n_missing_rmsd'])}",
        "",
        "Primary strict rank-1 RMSD statistics:",
        f"  mean:   {primary['rmsd_mean_angstrom']:.3f} Å",
        f"  median: {primary['rmsd_median_angstrom']:.3f} Å",
        f"  P25:    {primary['rmsd_p25_angstrom']:.3f} Å",
        f"  P75:    {primary['rmsd_p75_angstrom']:.3f} Å",
        f"  P90:    {primary['rmsd_p90_angstrom']:.3f} Å",
        f"  P95:    {primary['rmsd_p95_angstrom']:.3f} Å",
        f"  P99:    {primary['rmsd_p99_angstrom']:.3f} Å",
        "",
        "Cumulative target fractions:",
        f"  RMSD <= 2 Å:  {primary['rmsd_le_2A_percent']:.2f}%",
        f"  RMSD <= 4 Å:  {primary['rmsd_le_4A_percent']:.2f}%",
        f"  RMSD <= 5 Å:  {primary['rmsd_le_5A_percent']:.2f}%",
        f"  RMSD <= 10 Å: {primary['rmsd_le_10A_percent']:.2f}%",
        "",
        "Rank-1 is selected strictly by maximum ranking_score before consulting RMSD.",
        "RMSD is OpenStructure 2.8 rigid-score RMSD using mapped nucleic-acid C3' atoms.",
    ]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run_summary(args: argparse.Namespace) -> int:
    output_root = Path(args.output_root).expanduser().resolve()
    manifest_path = (
        Path(args.manifest).expanduser().resolve()
        if args.manifest
        else output_root / "manifest" / "candidates.csv"
    )
    report_dir = (
        Path(args.report_dir).expanduser().resolve()
        if args.report_dir
        else output_root / "reports"
    )
    report_dir.mkdir(parents=True, exist_ok=True)

    all_candidates = load_candidate_results(output_root, manifest_path)
    rank1 = select_strict_rank1(all_candidates)
    rank1 = add_target_metadata(
        rank1,
        targets_xlsx=Path(args.targets_xlsx).expanduser().resolve(),
        simple_json_root=Path(args.simple_json_root).expanduser().resolve(),
    )
    all_candidates["valid_rmsd"] = (
        all_candidates["eval_status"].eq("SUCCESS")
        & np.isfinite(pd.to_numeric(all_candidates["rmsd"], errors="coerce"))
        & pd.to_numeric(all_candidates["rmsd"], errors="coerce").ge(0)
    )
    all_candidates["rmsd_failure_reason"] = [
        classify_candidate_failure(output_root, row)
        for row in all_candidates.itertuples(index=False)
    ]
    rank1["rank1_valid_rmsd"] = (
        rank1["eval_status"].eq("SUCCESS")
        & np.isfinite(pd.to_numeric(rank1["rmsd"], errors="coerce"))
        & pd.to_numeric(rank1["rmsd"], errors="coerce").ge(0)
    )
    pdb_audit = build_pdb_rmsd_audit(all_candidates, rank1)
    tables = build_statistics(all_candidates, rank1)
    failures = all_candidates[all_candidates["eval_status"] != "SUCCESS"].copy()

    all_candidates.to_csv(report_dir / "all_candidates.csv", index=False)
    rank1.to_csv(report_dir / "rank1_targets.csv", index=False)
    failures.to_csv(report_dir / "failed_candidates.tsv", sep="\t", index=False)
    write_pdb_audit_outputs(
        report_dir, all_candidates=all_candidates, audit=pdb_audit
    )
    for name, frame in tables.items():
        frame.to_csv(report_dir / f"{name.lower()}.csv", index=False)

    readme = pd.DataFrame(
        {
            "item": [
                "Primary result",
                "Rank-1 rule",
                "RMSD definition",
                "Missing metrics",
                "Length definition",
                "Time cutoff",
                "Rigid-only rescue",
            ],
            "description": [
                "Overall sheet row rank1_by_ranking_score and rank1_targets sheet.",
                "Maximum Protenix ranking_score per target; never falls back based on RMSD.",
                "OpenStructure 2.8 --rigid-scores; mapped nucleic-acid C3' coordinates after one global Kabsch fit.",
                "Retained as missing and reported in denominators; not replaced by another candidate.",
                "Sum of len(rnaSequence.sequence) * count in the target Simple_json.",
                "pre_or_on_cutoff versus post_cutoff using 2021-09-30 split already stored in targets.xlsx.",
                "Used only for audited metric-stage failures: all-atom lDDT empty-array before RMSD, or downstream USalign TM-score failure after RMSD was reached. Missing metrics remain missing and evaluation_protocol is rigid_only_rescue.",
            ],
        }
    )
    workbook_path = report_dir / "rmsd_report.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        readme.to_excel(writer, sheet_name="README", index=False)
        for name, frame in tables.items():
            frame.to_excel(writer, sheet_name=name, index=False)
        rank1.to_excel(writer, sheet_name="Rank1_targets", index=False)
        all_candidates.to_excel(writer, sheet_name="All_candidates", index=False)
        failures.to_excel(writer, sheet_name="Failed_candidates", index=False)
    autosize_workbook(workbook_path)
    write_plots(rank1, report_dir)
    write_text_summary(
        report_dir / "run_summary.txt",
        all_candidates=all_candidates,
        rank1=rank1,
        overall=tables["Overall"],
    )
    console_summary = (report_dir / "run_summary.txt").read_text(encoding="utf-8")
    # Keep console output portable to non-UTF-8 terminals; the report retains Å.
    print(console_summary.replace("Å", "A"))
    print(f"Reports written to: {report_dir}")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Summarize FoldBench RNA RMSD outputs.")
    parser.add_argument(
        "--output-root", default="~/Json_data/Foldbench_evaluation/rmsd"
    )
    parser.add_argument("--manifest", default=None)
    parser.add_argument("--report-dir", default=None)
    parser.add_argument(
        "--targets-xlsx", default="~/Json_data/mapping/xlsx/targets.xlsx"
    )
    parser.add_argument("--simple-json-root", default="~/Json_data/Simple_json")
    return parser


def main(argv: Optional[list[str]] = None) -> int:
    return run_summary(build_parser().parse_args(argv))


if __name__ == "__main__":
    raise SystemExit(main())
