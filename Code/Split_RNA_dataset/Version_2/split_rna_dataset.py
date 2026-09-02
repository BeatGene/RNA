#!/usr/bin/env python3
"""Chronologically split pure-RNA PDB entries and mask redundant test chains.

The unit of assignment and materialization is a complete PDB entry.  This script
never copies or deletes an mmCIF file.  In execute mode it only creates empty
directories such as ``~/Data/train/157d``.

By default the command is a dry run.  Every run creates a timestamped report
directory below ``~/Code/pipeline_reports``.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import os
import platform
import re
import shutil
import subprocess
import sys
import traceback
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Sequence

try:
    from gemmi import cif
    from openpyxl import load_workbook
except ImportError as exc:
    missing = getattr(exc, "name", str(exc))
    raise SystemExit(
        f"Missing dependency {missing!r}. Install with:\n"
        "python3 -m pip install --user gemmi openpyxl"
    ) from exc


SCRIPT_DIR = Path(__file__).resolve().parent
PIPELINE_VERSION = "2.0-chain-evaluation-mask"
DEFAULT_CIF_DIR = Path("~/pdb_data").expanduser()
DEFAULT_DATA_DIR = Path("~/Data").expanduser()
DEFAULT_REPORT_ROOT = Path("~/Code/pipeline_reports").expanduser()
DEFAULT_EXCLUSION_XLSX = (
    SCRIPT_DIR.parent
    / "Download_PDB_RAW"
    / "Second_PDB_ID_xlsx_and_InOut"
    / "alignment_pdb_ids_not_in_experimental_pure_rna.xlsx"
)

EXPECTED_SOURCE_CIFS = 2246
EXPECTED_EXCLUSIONS = 5
EXPECTED_INCLUDED_PDBS = 2241
SHORT_SEQUENCE_THRESHOLD = 15
PDB_ID_RE = re.compile(r"^[A-Z0-9]{4}$")
RNA_POLYMER_TYPE = "polyribonucleotide"
NULL_VALUES = {"", ".", "?"}
REPORT_COLUMNS: dict[str, list[str]] = {
    "source_inventory.tsv": [
        "PDB_ID",
        "CIF_PATH",
        "FILE_SIZE_BYTES",
        "SHA256",
        "EXCLUDED",
    ],
    "date_audit.tsv": [
        "PDB_ID",
        "RELEASE_DATE",
        "SELECTED_SOURCE",
        "REVISION_ORDINAL",
        "REVISION_DATE",
        "LEGACY_ORIGINAL_DATE",
        "SOURCES_AGREE",
    ],
    "entity_metadata.tsv": [
        "PDB_ID",
        "ENTITY_ID",
        "CHAIN_IDS",
        "SEQUENCE",
        "SEARCH_SEQUENCE",
        "LENGTH",
        "AMBIGUOUS_BASES",
        "RELEASE_DATE",
        "EXPERIMENT_METHOD",
        "RESOLUTION_ANGSTROM",
    ],
    "split_before_dedup.tsv": [
        "PDB_ID",
        "RELEASE_DATE",
        "INITIAL_SPLIT",
        "RNA_ENTITY_COUNT",
        "RNA_CHAIN_COUNT",
        "RESOLUTION_ANGSTROM",
    ],
    "reference_redundancy_hits.tsv": [
        "QUERY_PDB_ID",
        "QUERY_ENTITY_ID",
        "TARGET_PDB_ID",
        "TARGET_ENTITY_ID",
        "IDENTITY",
        "QUERY_COVERAGE",
        "TARGET_COVERAGE",
        "ALIGNMENT_LENGTH",
        "EVALUE",
        "BITS",
        "ALIGNMENT_SOURCE",
    ],
    "test_internal_hits.tsv": [
        "QUERY_PDB_ID",
        "QUERY_ENTITY_ID",
        "TARGET_PDB_ID",
        "TARGET_ENTITY_ID",
        "IDENTITY",
        "QUERY_COVERAGE",
        "TARGET_COVERAGE",
        "ALIGNMENT_LENGTH",
        "EVALUE",
        "BITS",
        "ALIGNMENT_SOURCE",
    ],
    "test_chain_evaluation.tsv": [
        "PDB_ID",
        "CHAIN_ID",
        "ENTITY_ID",
        "RELEASE_DATE",
        "SEQUENCE_LENGTH",
        "EVALUATE",
        "CHAIN_STATUS",
        "REASON",
        "MATCH_PDB_ID",
        "MATCH_ENTITY_ID",
        "IDENTITY",
        "QUERY_COVERAGE",
        "TARGET_COVERAGE",
        "ALIGNMENT_SOURCE",
        "INTERNAL_CLUSTER_ID",
        "REPRESENTATIVE_PDB_ID",
        "REPRESENTATIVE_ENTITY_ID",
    ],
    "test_pdb_evaluation.tsv": [
        "PDB_ID",
        "RELEASE_DATE",
        "PDB_STATUS",
        "TOTAL_RNA_CHAINS",
        "EVALUATE_CHAIN_COUNT",
        "MASKED_CHAIN_COUNT",
        "EVALUATE_CHAIN_IDS",
        "MASKED_CHAIN_IDS",
        "TARGET_DIRECTORY",
    ],
    "final_manifest.tsv": [
        "PDB_ID",
        "RELEASE_DATE",
        "INITIAL_SPLIT",
        "FINAL_SPLIT",
        "FINAL_STATUS",
        "EXCLUSION_REASON",
        "EVALUATE_CHAIN_IDS",
        "MASKED_CHAIN_IDS",
        "TARGET_DIRECTORY",
    ],
    "mkdir_actions.tsv": ["ACTION", "SPLIT", "PDB_ID", "DIRECTORY"],
}


@dataclass(frozen=True)
class DateAudit:
    release_date: date
    selected_source: str
    revision_ordinal: str
    revision_date: str
    legacy_original_date: str
    sources_agree: bool | None


@dataclass(frozen=True)
class Entity:
    pdb_id: str
    entity_id: str
    chain_ids: tuple[str, ...]
    sequence: str
    search_sequence: str
    release_date: date
    experiment_method: str
    resolution: float | None


@dataclass(frozen=True)
class Entry:
    pdb_id: str
    cif_path: Path
    sha256: str
    release: DateAudit
    entities: tuple[Entity, ...]
    resolution: float | None


@dataclass(frozen=True)
class Hit:
    query_pdb_id: str
    query_entity_id: str
    target_pdb_id: str
    target_entity_id: str
    identity: float
    query_coverage: float
    target_coverage: float
    alignment_length: int
    evalue: str
    bits: str
    alignment_source: str = "MMSEQS2"


class RunLogger:
    def __init__(self, path: Path) -> None:
        self.path = path

    def log(self, message: str) -> None:
        stamp = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
        line = f"[{stamp}] {message}"
        print(line, flush=True)
        with self.path.open("a", encoding="utf-8") as handle:
            handle.write(line + "\n")


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Split 2241 pure-RNA PDB entries by release date, create a chain-level "
            "test evaluation mask with MMseqs2, and optionally create empty PDB directories."
        )
    )
    parser.add_argument("--cif-dir", type=Path, default=DEFAULT_CIF_DIR)
    parser.add_argument("--exclusion-xlsx", type=Path, default=DEFAULT_EXCLUSION_XLSX)
    parser.add_argument("--data-dir", type=Path, default=DEFAULT_DATA_DIR)
    parser.add_argument("--report-root", type=Path, default=DEFAULT_REPORT_ROOT)
    parser.add_argument("--train-end", type=date.fromisoformat, default=date(2023, 12, 31))
    parser.add_argument("--val-end", type=date.fromisoformat, default=date(2024, 12, 31))
    parser.add_argument("--min-seq-id", type=float, default=0.80)
    parser.add_argument("--min-query-cov", type=float, default=0.80)
    parser.add_argument("--min-target-cov", type=float, default=0.80)
    parser.add_argument("--threads", type=int, default=max(1, min(16, os.cpu_count() or 1)))
    parser.add_argument("--mmseqs", default="mmseqs", help="MMseqs2 executable")
    parser.add_argument(
        "--execute",
        action="store_true",
        help="create empty PDB directories; without this flag the run is read-only",
    )
    parser.add_argument(
        "--skip-count-check",
        action="store_true",
        help="development/testing only: do not enforce the 2246/5/2241 counts",
    )
    args = parser.parse_args(argv)
    if args.train_end >= args.val_end:
        parser.error("--train-end must be earlier than --val-end")
    for name in ("min_seq_id", "min_query_cov", "min_target_cov"):
        value = getattr(args, name)
        if not 0 < value <= 1:
            parser.error(f"--{name.replace('_', '-')} must be in (0, 1]")
    if args.threads < 1:
        parser.error("--threads must be positive")
    return args


def clean_cif_value(value: Any) -> str:
    if value is None or value is False:
        return ""
    text = cif.as_string(str(value)).strip()
    return "" if text in NULL_VALUES else text


def category_column(category: dict[str, list[Any]], name: str, size: int) -> list[Any]:
    values = category.get(name, [])
    if not values:
        return [""] * size
    if len(values) != size:
        raise ValueError(f"mmCIF column length mismatch: {name}={len(values)}, expected={size}")
    return values


def first_value(block: Any, tags: Iterable[str]) -> str:
    for tag in tags:
        values = block.find_values(tag)
        if values:
            value = clean_cif_value(values[0])
            if value:
                return value
    return ""


def normalize_pdb_id(value: Any, context: str) -> str:
    pdb_id = "" if value is None else str(value).strip().upper()
    if not PDB_ID_RE.fullmatch(pdb_id):
        raise ValueError(f"Invalid PDB ID in {context}: {value!r}")
    return pdb_id


def parse_iso_date(value: str, context: str) -> date:
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", value):
        raise ValueError(f"Invalid release-date format in {context}: {value!r}")
    try:
        parsed = date.fromisoformat(value)
    except ValueError as exc:
        raise ValueError(f"Invalid release date in {context}: {value!r}") from exc
    if parsed > date.today():
        raise ValueError(f"Future release date in {context}: {value!r}")
    return parsed


def extract_release_date(block: Any, pdb_id: str) -> DateAudit:
    category = block.get_mmcif_category("_pdbx_audit_revision_history.")
    revision_candidates: list[tuple[int, str, str]] = []
    if category:
        raw_dates = category.get("revision_date", [])
        raw_ordinals = category_column(category, "ordinal", len(raw_dates))
        for raw_ordinal, raw_date in zip(raw_ordinals, raw_dates):
            date_text = clean_cif_value(raw_date)
            if not date_text:
                continue
            ordinal_text = clean_cif_value(raw_ordinal)
            try:
                ordinal_sort = int(float(ordinal_text))
            except (TypeError, ValueError):
                ordinal_sort = 10**9
            revision_candidates.append((ordinal_sort, ordinal_text, date_text))

    revision_ordinal = ""
    revision_date = ""
    if revision_candidates:
        _, revision_ordinal, revision_date = min(
            revision_candidates, key=lambda item: (item[0], item[2])
        )
        parse_iso_date(revision_date, f"{pdb_id} revision history")

    legacy_date = first_value(block, ["_database_PDB_rev.date_original"])
    if legacy_date:
        parse_iso_date(legacy_date, f"{pdb_id} legacy original date")

    if revision_date and legacy_date and revision_date != legacy_date:
        raise ValueError(
            f"{pdb_id} has conflicting initial release dates: "
            f"revision_history={revision_date}, date_original={legacy_date}"
        )
    selected = revision_date or legacy_date
    if not selected:
        raise ValueError(f"{pdb_id} has no initial release date")
    return DateAudit(
        release_date=parse_iso_date(selected, pdb_id),
        selected_source=(
            "_pdbx_audit_revision_history.revision_date"
            if revision_date
            else "_database_PDB_rev.date_original"
        ),
        revision_ordinal=revision_ordinal,
        revision_date=revision_date,
        legacy_original_date=legacy_date,
        sources_agree=(revision_date == legacy_date if revision_date and legacy_date else None),
    )


def normalize_sequence(value: Any) -> str:
    return re.sub(r"\s+", "", clean_cif_value(value)).upper()


def make_search_sequence(sequence: str) -> str:
    # MMseqs2 nucleotide mode is most portable with DNA alphabet.  U/T are
    # chemically equivalent for sequence matching here; all other canonical
    # ambiguity symbols are represented as N.
    return "".join(base if base in "ACGT" else "N" for base in sequence.replace("U", "T"))


def extract_methods(block: Any) -> str:
    values: list[str] = []
    for raw in block.find_values("_exptl.method"):
        method = clean_cif_value(raw)
        if method and method not in values:
            values.append(method)
    if not values:
        raise ValueError("missing _exptl.method")
    return "; ".join(values)


def extract_resolution(block: Any) -> float | None:
    candidates: list[float] = []
    for tag in (
        "_refine.ls_d_res_high",
        "_em_3d_reconstruction.resolution",
        "_reflns.d_resolution_high",
    ):
        for raw in block.find_values(tag):
            value = clean_cif_value(raw)
            if not value:
                continue
            try:
                number = float(value)
            except ValueError:
                continue
            if math.isfinite(number) and number > 0:
                candidates.append(number)
    return min(candidates) if candidates else None


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def parse_entry(path: Path, expected_pdb_id: str) -> Entry:
    document = cif.read_file(str(path), check_level=2)
    if len(document) != 1:
        raise ValueError(f"expected one data block, found {len(document)}")
    block = document[0]
    entry_id = first_value(block, ["_entry.id"]).upper()
    if entry_id != expected_pdb_id:
        raise ValueError(f"_entry.id={entry_id!r}, filename ID={expected_pdb_id!r}")

    release = extract_release_date(block, expected_pdb_id)
    method = extract_methods(block)
    resolution = extract_resolution(block)
    category = block.get_mmcif_category("_entity_poly.")
    entity_ids = category.get("entity_id", [])
    if not entity_ids:
        raise ValueError("missing _entity_poly")
    size = len(entity_ids)
    types = category_column(category, "type", size)
    strands = category_column(category, "pdbx_strand_id", size)
    reported = category_column(category, "pdbx_seq_one_letter_code", size)
    canonical = category_column(category, "pdbx_seq_one_letter_code_can", size)

    polymer_types = {clean_cif_value(value).lower() for value in types if clean_cif_value(value)}
    if polymer_types != {RNA_POLYMER_TYPE}:
        raise ValueError(f"not pure RNA polymer content: {sorted(polymer_types)}")

    entities: list[Entity] = []
    seen_entity_ids: set[str] = set()
    for raw_id, raw_type, raw_strands, raw_reported, raw_canonical in zip(
        entity_ids, types, strands, reported, canonical
    ):
        if clean_cif_value(raw_type).lower() != RNA_POLYMER_TYPE:
            continue
        entity_id = clean_cif_value(raw_id)
        if not entity_id or entity_id in seen_entity_ids:
            raise ValueError(f"missing or duplicate RNA entity ID: {entity_id!r}")
        seen_entity_ids.add(entity_id)
        sequence = normalize_sequence(raw_canonical) or normalize_sequence(raw_reported)
        if not sequence:
            raise ValueError(f"RNA entity {entity_id} has no sequence")
        if not re.fullmatch(r"[A-Z]+", sequence):
            raise ValueError(f"RNA entity {entity_id} has unsupported canonical sequence: {sequence!r}")
        chain_ids = tuple(
            dict.fromkeys(
                item.strip()
                for item in clean_cif_value(raw_strands).split(",")
                if item.strip()
            )
        )
        if not chain_ids:
            raise ValueError(f"RNA entity {entity_id} has no chain IDs")
        entities.append(
            Entity(
                pdb_id=expected_pdb_id,
                entity_id=entity_id,
                chain_ids=chain_ids,
                sequence=sequence,
                search_sequence=make_search_sequence(sequence),
                release_date=release.release_date,
                experiment_method=method,
                resolution=resolution,
            )
        )
    if not entities:
        raise ValueError("no RNA entities")
    all_chain_ids = [chain_id for entity in entities for chain_id in entity.chain_ids]
    duplicate_chain_ids = sorted(
        chain_id for chain_id, count in Counter(all_chain_ids).items() if count > 1
    )
    if duplicate_chain_ids:
        raise ValueError(
            "RNA chain IDs occur in more than one entity: " + ", ".join(duplicate_chain_ids)
        )
    return Entry(
        pdb_id=expected_pdb_id,
        cif_path=path,
        sha256=sha256_file(path),
        release=release,
        entities=tuple(entities),
        resolution=resolution,
    )


def load_exclusion_ids(path: Path) -> set[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        ids = {
            normalize_pdb_id(row[0], f"{path.name} first column")
            for row in worksheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True)
            if row and row[0] is not None and str(row[0]).strip()
        }
    finally:
        workbook.close()
    if not ids:
        raise ValueError(f"No exclusion IDs found in {path}")
    return ids


def index_cifs(cif_dir: Path) -> dict[str, Path]:
    result: dict[str, Path] = {}
    for path in sorted(cif_dir.iterdir(), key=lambda item: item.name.lower()):
        if not path.is_file() or path.suffix.lower() != ".cif":
            continue
        pdb_id = normalize_pdb_id(path.stem, f"filename {path.name}")
        if pdb_id in result:
            raise ValueError(f"Duplicate CIF for {pdb_id}: {result[pdb_id]} and {path}")
        result[pdb_id] = path
    return result


def initial_split(release_date: date, train_end: date, val_end: date) -> str:
    if release_date <= train_end:
        return "train"
    if release_date <= val_end:
        return "val"
    return "test"


def write_tsv(path: Path, columns: Sequence[str], rows: Iterable[dict[str, Any]]) -> None:
    temporary = path.with_name(f".{path.name}.tmp")
    with temporary.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, delimiter="\t", extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: "" if row.get(key) is None else row.get(key) for key in columns})
    os.replace(temporary, path)


def write_json(path: Path, value: Any) -> None:
    temporary = path.with_name(f".{path.name}.tmp")
    with temporary.open("w", encoding="utf-8") as handle:
        json.dump(value, handle, ensure_ascii=False, indent=2, sort_keys=True)
        handle.write("\n")
    os.replace(temporary, path)


def make_report_dir(root: Path, execute: bool) -> Path:
    root.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    mode = "EXECUTE" if execute else "DRYRUN"
    candidate = root / f"DATA_SPLIT_2241_CHAINMASK_{timestamp}_{mode}"
    counter = 2
    while candidate.exists():
        candidate = root / f"DATA_SPLIT_2241_CHAINMASK_{timestamp}_{mode}_{counter}"
        counter += 1
    candidate.mkdir()
    return candidate


def fasta_id(entity: Entity) -> str:
    safe_entity = re.sub(r"[^A-Za-z0-9_.-]", "_", entity.entity_id)
    return f"{entity.pdb_id}__ENTITY_{safe_entity}"


def write_fasta(path: Path, entities: Iterable[Entity]) -> dict[str, Entity]:
    mapping: dict[str, Entity] = {}
    with path.open("w", encoding="ascii", newline="\n") as handle:
        for entity in entities:
            record_id = fasta_id(entity)
            if record_id in mapping:
                raise ValueError(f"Duplicate FASTA record ID: {record_id}")
            mapping[record_id] = entity
            handle.write(f">{record_id}\n{entity.search_sequence}\n")
    return mapping


def run_mmseqs(
    *,
    executable: str,
    query_fasta: Path,
    target_fasta: Path,
    output_tsv: Path,
    temporary_dir: Path,
    min_seq_id: float,
    min_cov: float,
    threads: int,
    logger: RunLogger,
) -> None:
    command = [
        executable,
        "easy-search",
        str(query_fasta),
        str(target_fasta),
        str(output_tsv),
        str(temporary_dir),
        "--search-type",
        "3",
        "--min-seq-id",
        str(min_seq_id),
        "-c",
        str(min_cov),
        "--cov-mode",
        "0",
        "--max-seqs",
        "10000",
        "--threads",
        str(threads),
        "--format-output",
        "query,target,fident,qcov,tcov,alnlen,evalue,bits",
    ]
    logger.log("Running: " + " ".join(command))
    with logger.path.open("a", encoding="utf-8") as log_handle:
        completed = subprocess.run(
            command,
            stdout=log_handle,
            stderr=subprocess.STDOUT,
            text=True,
            check=False,
        )
    if completed.returncode != 0:
        raise RuntimeError(f"MMseqs2 failed with exit code {completed.returncode}")


def parse_fraction(value: str) -> float:
    number = float(value)
    return number / 100.0 if number > 1.0 else number


def load_hits(
    path: Path,
    query_map: dict[str, Entity],
    target_map: dict[str, Entity],
    min_seq_id: float,
    min_query_cov: float,
    min_target_cov: float,
    *,
    exclude_same_pdb: bool,
) -> list[Hit]:
    hits: list[Hit] = []
    with path.open("r", encoding="utf-8") as handle:
        reader = csv.reader(handle, delimiter="\t")
        for line_number, row in enumerate(reader, start=1):
            if len(row) != 8:
                raise ValueError(f"Malformed MMseqs2 row {line_number} in {path}: {row!r}")
            query = query_map.get(row[0])
            target = target_map.get(row[1])
            if query is None or target is None:
                raise ValueError(f"Unknown MMseqs2 record on row {line_number}: {row[:2]!r}")
            if exclude_same_pdb and query.pdb_id == target.pdb_id:
                continue
            identity = parse_fraction(row[2])
            query_cov = parse_fraction(row[3])
            target_cov = parse_fraction(row[4])
            if (
                identity + 1e-12 < min_seq_id
                or query_cov + 1e-12 < min_query_cov
                or target_cov + 1e-12 < min_target_cov
            ):
                continue
            hits.append(
                Hit(
                    query_pdb_id=query.pdb_id,
                    query_entity_id=query.entity_id,
                    target_pdb_id=target.pdb_id,
                    target_entity_id=target.entity_id,
                    identity=identity,
                    query_coverage=query_cov,
                    target_coverage=target_cov,
                    alignment_length=int(row[5]),
                    evalue=row[6],
                    bits=row[7],
                    alignment_source="MMSEQS2",
                )
            )
    return sorted(
        hits,
        key=lambda hit: (
            hit.query_pdb_id,
            hit.target_pdb_id,
            hit.query_entity_id,
            hit.target_entity_id,
            -hit.identity,
        ),
    )


def hit_row(hit: Hit) -> dict[str, Any]:
    return {
        "QUERY_PDB_ID": hit.query_pdb_id,
        "QUERY_ENTITY_ID": hit.query_entity_id,
        "TARGET_PDB_ID": hit.target_pdb_id,
        "TARGET_ENTITY_ID": hit.target_entity_id,
        "IDENTITY": f"{hit.identity:.6f}",
        "QUERY_COVERAGE": f"{hit.query_coverage:.6f}",
        "TARGET_COVERAGE": f"{hit.target_coverage:.6f}",
        "ALIGNMENT_LENGTH": hit.alignment_length,
        "EVALUE": hit.evalue,
        "BITS": hit.bits,
        "ALIGNMENT_SOURCE": hit.alignment_source,
    }


def levenshtein_distance(left: str, right: str, maximum: int) -> int:
    """Return global edit distance, stopping when it is provably above maximum."""
    if abs(len(left) - len(right)) > maximum:
        return maximum + 1
    if len(left) > len(right):
        left, right = right, left
    previous = list(range(len(left) + 1))
    for row_index, right_base in enumerate(right, start=1):
        current = [row_index]
        row_minimum = row_index
        for column_index, left_base in enumerate(left, start=1):
            value = min(
                current[-1] + 1,
                previous[column_index] + 1,
                previous[column_index - 1] + (left_base != right_base),
            )
            current.append(value)
            row_minimum = min(row_minimum, value)
        if row_minimum > maximum:
            return maximum + 1
        previous = current
    return previous[-1]


def short_sequence_hits(
    query_entities: Sequence[Entity],
    target_entities: Sequence[Entity],
    min_seq_id: float,
    min_query_cov: float,
    min_target_cov: float,
    *,
    exclude_same_pdb: bool,
) -> list[Hit]:
    """Cover pairs that nucleotide seed search may miss because a sequence is tiny.

    A global edit alignment covers both sequences completely.  Its identity is
    defined as ``1 - edit_distance / max(lengths)``.  Only pairs involving a
    sequence shorter than SHORT_SEQUENCE_THRESHOLD are evaluated here.
    """
    if min_query_cov > 1 or min_target_cov > 1:
        return []
    hits: list[Hit] = []
    for query in query_entities:
        for target in target_entities:
            if exclude_same_pdb and query.pdb_id == target.pdb_id:
                continue
            if min(len(query.search_sequence), len(target.search_sequence)) >= SHORT_SEQUENCE_THRESHOLD:
                continue
            longest = max(len(query.search_sequence), len(target.search_sequence))
            if longest == 0:
                continue
            maximum_edits = math.floor((1.0 - min_seq_id + 1e-12) * longest)
            distance = levenshtein_distance(
                query.search_sequence, target.search_sequence, maximum_edits
            )
            if distance > maximum_edits:
                continue
            identity = 1.0 - distance / longest
            if identity + 1e-12 < min_seq_id:
                continue
            hits.append(
                Hit(
                    query_pdb_id=query.pdb_id,
                    query_entity_id=query.entity_id,
                    target_pdb_id=target.pdb_id,
                    target_entity_id=target.entity_id,
                    identity=identity,
                    query_coverage=1.0,
                    target_coverage=1.0,
                    alignment_length=longest,
                    evalue="",
                    bits="",
                    alignment_source="SHORT_GLOBAL_FALLBACK",
                )
            )
    return hits


def merge_hits(*hit_groups: Sequence[Hit]) -> list[Hit]:
    """Merge identical directed entity pairs, preferring MMseqs2 then best identity."""
    merged: dict[tuple[str, str, str, str], Hit] = {}
    for hit in (hit for group in hit_groups for hit in group):
        key = (
            hit.query_pdb_id,
            hit.query_entity_id,
            hit.target_pdb_id,
            hit.target_entity_id,
        )
        current = merged.get(key)
        candidate_key = (
            hit.identity,
            hit.query_coverage,
            hit.target_coverage,
            hit.alignment_source == "MMSEQS2",
        )
        if current is None or candidate_key > (
            current.identity,
            current.query_coverage,
            current.target_coverage,
            current.alignment_source == "MMSEQS2",
        ):
            merged[key] = hit
    return sorted(
        merged.values(),
        key=lambda hit: (
            hit.query_pdb_id,
            hit.target_pdb_id,
            hit.query_entity_id,
            hit.target_entity_id,
        ),
    )


def hit_query_key(hit: Hit) -> tuple[str, str]:
    return hit.query_pdb_id, hit.query_entity_id


def hit_target_key(hit: Hit) -> tuple[str, str]:
    return hit.target_pdb_id, hit.target_entity_id


def entity_key(entity: Entity) -> tuple[str, str]:
    return entity.pdb_id, entity.entity_id


def connected_components(
    nodes: set[tuple[str, str]], hits: Sequence[Hit]
) -> list[set[tuple[str, str]]]:
    adjacency: dict[tuple[str, str], set[tuple[str, str]]] = {
        node: set() for node in nodes
    }
    for hit in hits:
        query = hit_query_key(hit)
        target = hit_target_key(hit)
        if query in nodes and target in nodes and query != target:
            adjacency[query].add(target)
            adjacency[target].add(query)
    result: list[set[tuple[str, str]]] = []
    unseen = set(nodes)
    while unseen:
        start = min(unseen)
        stack = [start]
        component: set[tuple[str, str]] = set()
        while stack:
            node = stack.pop()
            if node in component:
                continue
            component.add(node)
            unseen.discard(node)
            stack.extend(adjacency[node] - component)
        result.append(component)
    return sorted(result, key=lambda component: min(component))


def representative_key(
    key: tuple[str, str], entries: dict[str, Entry]
) -> tuple[float, date, str, str]:
    pdb_id, entity_id = key
    entry = entries[pdb_id]
    return (
        entry.resolution if entry.resolution is not None else math.inf,
        entry.release.release_date,
        pdb_id,
        entity_id,
    )


def preflight_layout(data_dir: Path, expected: dict[str, set[str]]) -> None:
    for split in ("train", "val", "test"):
        split_dir = data_dir / split
        if not split_dir.exists():
            continue
        if not split_dir.is_dir():
            raise ValueError(f"Expected directory but found another file type: {split_dir}")
        for child in split_dir.iterdir():
            child_id = child.name.upper()
            if child.is_dir() and PDB_ID_RE.fullmatch(child_id):
                if child_id not in expected[split]:
                    raise ValueError(
                        f"Stale or wrong-split PDB directory already exists: {child}"
                    )


def materialize_empty_directories(
    data_dir: Path,
    expected: dict[str, set[str]],
    execute: bool,
) -> list[dict[str, Any]]:
    preflight_layout(data_dir, expected)
    actions: list[dict[str, Any]] = []
    for split in ("train", "val", "test"):
        split_dir = data_dir / split
        if execute:
            split_dir.mkdir(parents=True, exist_ok=True)
        for pdb_id in sorted(expected[split]):
            directory = split_dir / pdb_id.lower()
            existed = directory.is_dir()
            action = "EXISTS" if existed else ("CREATED" if execute else "WOULD_CREATE")
            if execute and not existed:
                directory.mkdir()
            actions.append(
                {
                    "ACTION": action,
                    "SPLIT": split,
                    "PDB_ID": pdb_id,
                    "DIRECTORY": str(directory),
                }
            )
    return actions


def mmseqs_version(executable: str) -> str:
    resolved = shutil.which(executable)
    if not resolved:
        raise FileNotFoundError(
            f"MMseqs2 executable {executable!r} was not found in PATH. "
            "Install MMseqs2 on the server before running this pipeline."
        )
    completed = subprocess.run(
        [resolved, "version"], capture_output=True, text=True, check=False
    )
    text = (completed.stdout or completed.stderr).strip()
    if completed.returncode != 0:
        raise RuntimeError(f"Could not query MMseqs2 version: {text}")
    return text


def run_pipeline(args: argparse.Namespace, report_dir: Path, logger: RunLogger) -> dict[str, Any]:
    cif_dir = args.cif_dir.expanduser().resolve()
    exclusion_xlsx = args.exclusion_xlsx.expanduser().resolve()
    data_dir = args.data_dir.expanduser().resolve()
    if not cif_dir.is_dir():
        raise FileNotFoundError(f"CIF directory not found: {cif_dir}")
    if not exclusion_xlsx.is_file():
        raise FileNotFoundError(f"Exclusion XLSX not found: {exclusion_xlsx}")

    version = mmseqs_version(args.mmseqs)
    logger.log(f"MMseqs2 version: {version}")
    exclusions = load_exclusion_ids(exclusion_xlsx)
    cif_files = index_cifs(cif_dir)
    missing_exclusions = exclusions - set(cif_files)
    if missing_exclusions:
        raise ValueError(f"Exclusion IDs missing from CIF directory: {sorted(missing_exclusions)}")
    included_files = {pdb_id: path for pdb_id, path in cif_files.items() if pdb_id not in exclusions}
    counts = {
        "source_cifs": len(cif_files),
        "exclusion_ids": len(exclusions),
        "included_pdbs": len(included_files),
    }
    logger.log(f"Input counts: {counts}")
    if not args.skip_count_check:
        expected_counts = {
            "source_cifs": EXPECTED_SOURCE_CIFS,
            "exclusion_ids": EXPECTED_EXCLUSIONS,
            "included_pdbs": EXPECTED_INCLUDED_PDBS,
        }
        if counts != expected_counts:
            raise ValueError(f"Input-count audit failed: actual={counts}, expected={expected_counts}")

    inventory_rows: list[dict[str, Any]] = []
    entries: dict[str, Entry] = {}
    errors: list[dict[str, str]] = []
    for index, (pdb_id, path) in enumerate(sorted(cif_files.items()), start=1):
        excluded = pdb_id in exclusions
        try:
            digest = sha256_file(path) if excluded else ""
            if not excluded:
                entry = parse_entry(path, pdb_id)
                entries[pdb_id] = entry
                digest = entry.sha256
            inventory_rows.append(
                {
                    "PDB_ID": pdb_id,
                    "CIF_PATH": str(path),
                    "FILE_SIZE_BYTES": path.stat().st_size,
                    "SHA256": digest,
                    "EXCLUDED": excluded,
                }
            )
        except Exception as exc:
            errors.append({"PDB_ID": pdb_id, "CIF_PATH": str(path), "ERROR": f"{type(exc).__name__}: {exc}"})
        if index % 100 == 0 or index == len(cif_files):
            logger.log(f"Parsed {index}/{len(cif_files)} CIF files; errors={len(errors)}")
    write_tsv(report_dir / "source_inventory.tsv", REPORT_COLUMNS["source_inventory.tsv"], inventory_rows)
    if errors:
        write_tsv(report_dir / "parse_errors.tsv", ["PDB_ID", "CIF_PATH", "ERROR"], errors)
        raise ValueError(f"{len(errors)} CIF files failed validation; see parse_errors.tsv")
    if len(entries) != EXPECTED_INCLUDED_PDBS and not args.skip_count_check:
        raise ValueError(f"Parsed {len(entries)} included PDBs, expected {EXPECTED_INCLUDED_PDBS}")

    date_rows: list[dict[str, Any]] = []
    entity_rows: list[dict[str, Any]] = []
    split_rows: list[dict[str, Any]] = []
    assignments: dict[str, str] = {}
    for pdb_id, entry in sorted(entries.items()):
        release = entry.release
        date_rows.append(
            {
                "PDB_ID": pdb_id,
                "RELEASE_DATE": release.release_date.isoformat(),
                "SELECTED_SOURCE": release.selected_source,
                "REVISION_ORDINAL": release.revision_ordinal,
                "REVISION_DATE": release.revision_date,
                "LEGACY_ORIGINAL_DATE": release.legacy_original_date,
                "SOURCES_AGREE": release.sources_agree,
            }
        )
        for entity in entry.entities:
            entity_rows.append(
                {
                    "PDB_ID": pdb_id,
                    "ENTITY_ID": entity.entity_id,
                    "CHAIN_IDS": ",".join(entity.chain_ids),
                    "SEQUENCE": entity.sequence,
                    "SEARCH_SEQUENCE": entity.search_sequence,
                    "LENGTH": len(entity.search_sequence),
                    "AMBIGUOUS_BASES": entity.search_sequence.count("N"),
                    "RELEASE_DATE": entity.release_date.isoformat(),
                    "EXPERIMENT_METHOD": entity.experiment_method,
                    "RESOLUTION_ANGSTROM": entity.resolution,
                }
            )
        split = initial_split(release.release_date, args.train_end, args.val_end)
        assignments[pdb_id] = split
        split_rows.append(
            {
                "PDB_ID": pdb_id,
                "RELEASE_DATE": release.release_date.isoformat(),
                "INITIAL_SPLIT": split,
                "RNA_ENTITY_COUNT": len(entry.entities),
                "RNA_CHAIN_COUNT": sum(len(entity.chain_ids) for entity in entry.entities),
                "RESOLUTION_ANGSTROM": entry.resolution,
            }
        )
    write_tsv(report_dir / "date_audit.tsv", REPORT_COLUMNS["date_audit.tsv"], date_rows)
    write_tsv(report_dir / "entity_metadata.tsv", REPORT_COLUMNS["entity_metadata.tsv"], entity_rows)
    write_tsv(report_dir / "split_before_dedup.tsv", REPORT_COLUMNS["split_before_dedup.tsv"], split_rows)
    before_counts = Counter(assignments.values())
    logger.log(f"Chronological split before de-redundancy: {dict(before_counts)}")

    work_dir = report_dir / "mmseqs_work"
    work_dir.mkdir()
    reference_entities = [
        entity
        for pdb_id, entry in sorted(entries.items())
        if assignments[pdb_id] in {"train", "val"}
        for entity in entry.entities
    ]
    test_entities = [
        entity
        for pdb_id, entry in sorted(entries.items())
        if assignments[pdb_id] == "test"
        for entity in entry.entities
    ]
    reference_map = write_fasta(work_dir / "train_val_entities.fasta", reference_entities)
    test_map = write_fasta(work_dir / "test_entities.fasta", test_entities)
    raw_reference_hits = work_dir / "test_vs_train_val.raw.tsv"
    run_mmseqs(
        executable=args.mmseqs,
        query_fasta=work_dir / "test_entities.fasta",
        target_fasta=work_dir / "train_val_entities.fasta",
        output_tsv=raw_reference_hits,
        temporary_dir=work_dir / "tmp_reference",
        min_seq_id=args.min_seq_id,
        min_cov=min(args.min_query_cov, args.min_target_cov),
        threads=args.threads,
        logger=logger,
    )
    mmseqs_reference_hits = load_hits(
        raw_reference_hits,
        test_map,
        reference_map,
        args.min_seq_id,
        args.min_query_cov,
        args.min_target_cov,
        exclude_same_pdb=False,
    )
    fallback_reference_hits = short_sequence_hits(
        test_entities,
        reference_entities,
        args.min_seq_id,
        args.min_query_cov,
        args.min_target_cov,
        exclude_same_pdb=False,
    )
    reference_hits = merge_hits(mmseqs_reference_hits, fallback_reference_hits)
    logger.log(
        "Reference hits: "
        f"MMseqs2={len(mmseqs_reference_hits)}, "
        f"short-fallback={len(fallback_reference_hits)}, "
        f"merged={len(reference_hits)}"
    )
    write_tsv(
        report_dir / "reference_redundancy_hits.tsv",
        REPORT_COLUMNS["reference_redundancy_hits.tsv"],
        (hit_row(hit) for hit in reference_hits),
    )
    test_entity_by_key = {entity_key(entity): entity for entity in test_entities}
    reference_masked_entities = {hit_query_key(hit) for hit in reference_hits}
    reference_affected_pdbs = {pdb_id for pdb_id, _ in reference_masked_entities}
    logger.log(
        "Test entities homologous to train/val: "
        f"entities={len(reference_masked_entities)}, "
        f"affected_PDBs={len(reference_affected_pdbs)}"
    )

    remaining_entities = [
        entity
        for entity in test_entities
        if entity_key(entity) not in reference_masked_entities
    ]
    remaining_fasta = work_dir / "remaining_test_entities.fasta"
    remaining_map = write_fasta(remaining_fasta, remaining_entities)
    raw_internal_hits = work_dir / "remaining_test_internal.raw.tsv"
    if remaining_entities:
        run_mmseqs(
            executable=args.mmseqs,
            query_fasta=remaining_fasta,
            target_fasta=remaining_fasta,
            output_tsv=raw_internal_hits,
            temporary_dir=work_dir / "tmp_internal",
            min_seq_id=args.min_seq_id,
            min_cov=min(args.min_query_cov, args.min_target_cov),
            threads=args.threads,
            logger=logger,
        )
        mmseqs_internal_hits = load_hits(
            raw_internal_hits,
            remaining_map,
            remaining_map,
            args.min_seq_id,
            args.min_query_cov,
            args.min_target_cov,
            exclude_same_pdb=True,
        )
        fallback_internal_hits = short_sequence_hits(
            remaining_entities,
            remaining_entities,
            args.min_seq_id,
            args.min_query_cov,
            args.min_target_cov,
            exclude_same_pdb=True,
        )
        internal_hits = merge_hits(mmseqs_internal_hits, fallback_internal_hits)
        logger.log(
            "Internal-test hits: "
            f"MMseqs2={len(mmseqs_internal_hits)}, "
            f"short-fallback={len(fallback_internal_hits)}, "
            f"merged={len(internal_hits)}"
        )
    else:
        raw_internal_hits.write_text("", encoding="utf-8")
        internal_hits = []
    write_tsv(
        report_dir / "test_internal_hits.tsv",
        REPORT_COLUMNS["test_internal_hits.tsv"],
        (hit_row(hit) for hit in internal_hits),
    )

    remaining_entity_keys = {entity_key(entity) for entity in remaining_entities}
    components = connected_components(remaining_entity_keys, internal_hits)
    representative_for: dict[tuple[str, str], tuple[str, str]] = {}
    cluster_for: dict[tuple[str, str], str] = {}
    representative_entities: set[tuple[str, str]] = set()
    for index, component in enumerate(components, start=1):
        representative = min(
            component, key=lambda key: representative_key(key, entries)
        )
        representative_entities.add(representative)
        cluster_id = f"TEST_ENTITY_CLUSTER_{index:04d}"
        for key in component:
            representative_for[key] = representative
            cluster_for[key] = cluster_id

    internal_masked_entities = remaining_entity_keys - representative_entities
    internal_affected_pdbs = {pdb_id for pdb_id, _ in internal_masked_entities}
    logger.log(
        "Internal-test entity clustering: "
        f"input_entities={len(remaining_entity_keys)}, "
        f"clusters={len(components)}, "
        f"masked_entities={len(internal_masked_entities)}, "
        f"affected_PDBs={len(internal_affected_pdbs)}"
    )

    best_reference_match: dict[tuple[str, str], Hit] = {}
    for hit in sorted(
        reference_hits,
        key=lambda item: (
            -item.identity,
            -item.query_coverage,
            -item.target_coverage,
            item.target_pdb_id,
            item.target_entity_id,
        ),
    ):
        best_reference_match.setdefault(hit_query_key(hit), hit)

    chain_rows: list[dict[str, Any]] = []
    pdb_rows: list[dict[str, Any]] = []
    mask_json: dict[str, Any] = {
        "schema_version": 1,
        "policy": {
            "unit": "RNA entity propagated to all member chains",
            "evaluate_status": "EVALUATE",
            "masked_statuses": [
                "MASK_REFERENCE_HOMOLOG",
                "MASK_INTERNAL_REDUNDANT",
            ],
            "pdb_directory_rule": "create test directory when at least one RNA chain is EVALUATE",
        },
        "pdbs": {},
    }
    pdb_status_by_id: dict[str, str] = {}
    evaluate_chains_by_pdb: dict[str, list[str]] = {}
    masked_chains_by_pdb: dict[str, list[str]] = {}

    for pdb_id in sorted(pdb for pdb, split in assignments.items() if split == "test"):
        entry = entries[pdb_id]
        evaluate_chain_ids: list[str] = []
        masked_chain_ids: list[str] = []
        json_chains: dict[str, Any] = {}
        for entity in entry.entities:
            key = entity_key(entity)
            match: Hit | None = None
            if key in reference_masked_entities:
                evaluate = False
                chain_status = "MASK_REFERENCE_HOMOLOG"
                reason = "Entity homologous to train/val at configured thresholds"
                match = best_reference_match[key]
                cluster_id = ""
                representative = None
            elif key in internal_masked_entities:
                evaluate = False
                chain_status = "MASK_INTERNAL_REDUNDANT"
                reason = "Non-representative entity in a test homology component"
                cluster_id = cluster_for[key]
                representative = representative_for[key]
            else:
                evaluate = True
                chain_status = "EVALUATE"
                reason = ""
                cluster_id = cluster_for[key]
                representative = representative_for[key]

            representative_pdb = representative[0] if representative else ""
            representative_entity = representative[1] if representative else ""
            for chain_id in entity.chain_ids:
                if evaluate:
                    evaluate_chain_ids.append(chain_id)
                else:
                    masked_chain_ids.append(chain_id)
                chain_row = {
                    "PDB_ID": pdb_id,
                    "CHAIN_ID": chain_id,
                    "ENTITY_ID": entity.entity_id,
                    "RELEASE_DATE": entity.release_date.isoformat(),
                    "SEQUENCE_LENGTH": len(entity.search_sequence),
                    "EVALUATE": evaluate,
                    "CHAIN_STATUS": chain_status,
                    "REASON": reason,
                    "MATCH_PDB_ID": match.target_pdb_id if match else "",
                    "MATCH_ENTITY_ID": match.target_entity_id if match else "",
                    "IDENTITY": f"{match.identity:.6f}" if match else "",
                    "QUERY_COVERAGE": f"{match.query_coverage:.6f}" if match else "",
                    "TARGET_COVERAGE": f"{match.target_coverage:.6f}" if match else "",
                    "ALIGNMENT_SOURCE": match.alignment_source if match else "",
                    "INTERNAL_CLUSTER_ID": cluster_id,
                    "REPRESENTATIVE_PDB_ID": representative_pdb,
                    "REPRESENTATIVE_ENTITY_ID": representative_entity,
                }
                chain_rows.append(chain_row)
                json_chains[chain_id] = {
                    "entity_id": entity.entity_id,
                    "sequence_length": len(entity.search_sequence),
                    "evaluate": evaluate,
                    "status": chain_status,
                    "reason": reason,
                    "reference_match": (
                        {
                            "pdb_id": match.target_pdb_id,
                            "entity_id": match.target_entity_id,
                            "identity": match.identity,
                            "query_coverage": match.query_coverage,
                            "target_coverage": match.target_coverage,
                            "alignment_source": match.alignment_source,
                        }
                        if match
                        else None
                    ),
                    "internal_cluster_id": cluster_id or None,
                    "representative": (
                        {"pdb_id": representative_pdb, "entity_id": representative_entity}
                        if representative
                        else None
                    ),
                }

        if evaluate_chain_ids and masked_chain_ids:
            pdb_status = "KEEP_PARTIAL_CHAINS"
        elif evaluate_chain_ids:
            pdb_status = "KEEP_ALL_CHAINS"
        else:
            pdb_status = "DROP_NO_EVALUABLE_CHAINS"
        pdb_status_by_id[pdb_id] = pdb_status
        evaluate_chains_by_pdb[pdb_id] = evaluate_chain_ids
        masked_chains_by_pdb[pdb_id] = masked_chain_ids
        target_directory = (
            str(data_dir / "test" / pdb_id.lower()) if evaluate_chain_ids else ""
        )
        pdb_rows.append(
            {
                "PDB_ID": pdb_id,
                "RELEASE_DATE": entry.release.release_date.isoformat(),
                "PDB_STATUS": pdb_status,
                "TOTAL_RNA_CHAINS": len(evaluate_chain_ids) + len(masked_chain_ids),
                "EVALUATE_CHAIN_COUNT": len(evaluate_chain_ids),
                "MASKED_CHAIN_COUNT": len(masked_chain_ids),
                "EVALUATE_CHAIN_IDS": ",".join(evaluate_chain_ids),
                "MASKED_CHAIN_IDS": ",".join(masked_chain_ids),
                "TARGET_DIRECTORY": target_directory,
            }
        )
        mask_json["pdbs"][pdb_id] = {
            "release_date": entry.release.release_date.isoformat(),
            "pdb_status": pdb_status,
            "evaluate_chain_ids": evaluate_chain_ids,
            "masked_chain_ids": masked_chain_ids,
            "chains": json_chains,
        }

    write_tsv(
        report_dir / "test_chain_evaluation.tsv",
        REPORT_COLUMNS["test_chain_evaluation.tsv"],
        chain_rows,
    )
    write_tsv(
        report_dir / "test_pdb_evaluation.tsv",
        REPORT_COLUMNS["test_pdb_evaluation.tsv"],
        pdb_rows,
    )
    write_json(report_dir / "test_evaluation_mask.json", mask_json)

    final_test_ids = {
        pdb_id for pdb_id, chain_ids in evaluate_chains_by_pdb.items() if chain_ids
    }
    expected: dict[str, set[str]] = {
        "train": {pdb_id for pdb_id, split in assignments.items() if split == "train"},
        "val": {pdb_id for pdb_id, split in assignments.items() if split == "val"},
        "test": final_test_ids,
    }
    manifest: list[dict[str, Any]] = []
    for pdb_id, entry in sorted(entries.items()):
        split = assignments[pdb_id]
        if split in {"train", "val"}:
            status = "KEPT"
            final_split = split
            reason = ""
            evaluate_chain_ids = []
            masked_chain_ids = []
        else:
            status = pdb_status_by_id[pdb_id]
            final_split = "test" if pdb_id in final_test_ids else ""
            reason = (
                "All RNA chains are masked from evaluation"
                if status == "DROP_NO_EVALUABLE_CHAINS"
                else ""
            )
            evaluate_chain_ids = evaluate_chains_by_pdb[pdb_id]
            masked_chain_ids = masked_chains_by_pdb[pdb_id]
        target_directory = (
            str(data_dir / final_split / pdb_id.lower()) if final_split else ""
        )
        manifest.append(
            {
                "PDB_ID": pdb_id,
                "RELEASE_DATE": entry.release.release_date.isoformat(),
                "INITIAL_SPLIT": split,
                "FINAL_SPLIT": final_split,
                "FINAL_STATUS": status,
                "EXCLUSION_REASON": reason,
                "EVALUATE_CHAIN_IDS": ",".join(evaluate_chain_ids),
                "MASKED_CHAIN_IDS": ",".join(masked_chain_ids),
                "TARGET_DIRECTORY": target_directory,
            }
        )
    write_tsv(report_dir / "final_manifest.tsv", REPORT_COLUMNS["final_manifest.tsv"], manifest)

    actions = materialize_empty_directories(data_dir, expected, args.execute)
    write_tsv(report_dir / "mkdir_actions.tsv", REPORT_COLUMNS["mkdir_actions.tsv"], actions)
    final_counts = {split: len(ids) for split, ids in expected.items()}
    reference_masked_chain_count = sum(
        len(test_entity_by_key[key].chain_ids) for key in reference_masked_entities
    )
    internal_masked_chain_count = sum(
        len(test_entity_by_key[key].chain_ids) for key in internal_masked_entities
    )
    evaluated_entities = remaining_entity_keys - internal_masked_entities
    evaluated_chain_count = sum(
        len(test_entity_by_key[key].chain_ids) for key in evaluated_entities
    )
    pdb_status_counts = Counter(pdb_status_by_id.values())
    summary = {
        "status": "SUCCESS",
        "pipeline_version": PIPELINE_VERSION,
        "mode": "EXECUTE" if args.execute else "DRYRUN",
        "source_counts": counts,
        "excluded_pdb_ids": sorted(exclusions),
        "date_boundaries": {
            "train": f"release_date <= {args.train_end.isoformat()}",
            "val": f"{(args.train_end + date.resolution).isoformat()} <= release_date <= {args.val_end.isoformat()}",
            "test": f"release_date > {args.val_end.isoformat()}",
        },
        "split_before_dedup": dict(sorted(before_counts.items())),
        "dedup_thresholds": {
            "minimum_sequence_identity": args.min_seq_id,
            "minimum_query_coverage": args.min_query_cov,
            "minimum_target_coverage": args.min_target_cov,
            "mask_rule": "mask homologous RNA entity and all chains belonging to that entity",
            "pdb_directory_rule": "keep test PDB when at least one RNA chain remains evaluable",
        },
        "test_initial_counts": {
            "pdbs": before_counts["test"],
            "entities": len(test_entities),
            "chains": sum(len(entity.chain_ids) for entity in test_entities),
        },
        "reference_homology_mask": {
            "masked_entities": len(reference_masked_entities),
            "masked_chains": reference_masked_chain_count,
            "affected_pdbs": len(reference_affected_pdbs),
        },
        "internal_test_mask": {
            "masked_entities": len(internal_masked_entities),
            "masked_chains": internal_masked_chain_count,
            "affected_pdbs": len(internal_affected_pdbs),
            "entity_clusters": len(components),
        },
        "final_test_evaluation": {
            "pdbs_with_evaluable_chains": len(final_test_ids),
            "evaluable_entities": len(evaluated_entities),
            "evaluable_chains": evaluated_chain_count,
            "pdb_status_counts": dict(sorted(pdb_status_counts.items())),
        },
        "final_directory_counts": final_counts,
        "mmseqs_version": version,
        "report_directory": str(report_dir),
    }
    write_json(report_dir / "summary.json", summary)
    logger.log(f"Final directory counts: {final_counts}")
    logger.log(
        "Final test evaluation: "
        f"PDBs={len(final_test_ids)}, entities={len(evaluated_entities)}, "
        f"chains={evaluated_chain_count}, statuses={dict(pdb_status_counts)}"
    )
    return summary


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    report_root = args.report_root.expanduser().resolve()
    report_dir = make_report_dir(report_root, args.execute)
    logger = RunLogger(report_dir / "pipeline.log")
    config = {
        "pipeline_version": PIPELINE_VERSION,
        "script_path": str(Path(__file__).resolve()),
        "script_sha256": sha256_file(Path(__file__).resolve()),
        "argv": sys.argv if argv is None else [sys.argv[0], *argv],
        "cwd": str(Path.cwd()),
        "python": sys.version,
        "platform": platform.platform(),
        "started_at_utc": datetime.now(timezone.utc).isoformat(),
        "arguments": {
            key: (value.isoformat() if isinstance(value, date) else str(value) if isinstance(value, Path) else value)
            for key, value in vars(args).items()
        },
    }
    write_json(report_dir / "run_config.json", config)
    logger.log(f"Report directory: {report_dir}")
    logger.log("Mode: " + ("EXECUTE (create empty directories)" if args.execute else "DRYRUN (no Data writes)"))
    try:
        summary = run_pipeline(args, report_dir, logger)
        print(json.dumps(summary, ensure_ascii=False, indent=2), flush=True)
        return 0
    except Exception as exc:
        logger.log(f"FAILED: {type(exc).__name__}: {exc}")
        with (report_dir / "traceback.txt").open("w", encoding="utf-8") as handle:
            traceback.print_exc(file=handle)
        write_json(
            report_dir / "failure.json",
            {
                "status": "FAILED",
                "error_type": type(exc).__name__,
                "error": str(exc),
                "report_directory": str(report_dir),
            },
        )
        print(f"FAILED; see {report_dir}", file=sys.stderr, flush=True)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
